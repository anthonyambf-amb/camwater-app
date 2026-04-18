"""Export Excel - Génère des fichiers conformes à la structure Consolidation 2026."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from database import (get_db, STRUCTURE_CW, MOIS, CATEGORIES_VOLUMES,
                       CATEGORIES_GCO, CATEGORIES_ADM, CATEGORIES_BC, CATEGORIES_BFC, CATEGORIES_CW,
                       RUBRIQUES_ENC_CCIALE, RUBRIQUES_ENC_ADM, RUBRIQUES_ENC_BANQUES,
                       RUBRIQUES_ENC_TRVX_CCIALE, RUBRIQUES_ENC_TRVX_ADM, TVA_RATE)
from calculs import (calcul_ca_agence, calcul_dr, calcul_national,
                      calcul_cumul_dr, calcul_cumul_national,
                      get_agences_dr, get_branchements_par_dr,
                      classement_performances)

# Styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
NUM_FMT = '#,##0'
PCT_FMT = '0.00%'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_row(ws, row, max_col, fill=None, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.border = THIN_BORDER
        if col > 1:
            cell.number_format = NUM_FMT


def _write_agence_sheet(ws, agence_id, agence_nom, mois_debut, mois_fin, exercice):
    """Écrit une feuille complète pour une agence (volumes + CA + encaissements)."""
    ws.title = agence_nom[:31]
    cols = ["Rubrique"] + [MOIS[m - 1] for m in range(mois_debut, mois_fin + 1)] + ["Cumul"]
    max_col = len(cols)

    # ─── VOLUMES ───
    row = 1
    ws.cell(row=row, column=1, value=f"VOLUMES (m3) - {agence_nom}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    all_cats = CATEGORIES_VOLUMES + [CATEGORIES_GCO, CATEGORIES_ADM, CATEGORIES_BC, CATEGORIES_BFC] + CATEGORIES_CW
    row = 3
    for cat_name, pu in all_cats:
        ws.cell(row=row, column=1, value=cat_name)
        cumul = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            db = get_db()
            r = db.execute("SELECT valeur FROM volumes WHERE agence_id=? AND mois=? AND exercice=? AND categorie=?",
                           (agence_id, m, exercice, cat_name)).fetchone()
            db.close()
            val = r['valeur'] if r else 0
            ws.cell(row=row, column=i + 2, value=val)
            cumul += val
        ws.cell(row=row, column=max_col, value=cumul)
        _style_row(ws, row, max_col)
        row += 1

    # Total volumes
    ws.cell(row=row, column=1, value="TOTAL VOLUMES")
    for i, m in enumerate(range(mois_debut, mois_fin + 1)):
        data = calcul_ca_agence(agence_id, m, exercice)
        ws.cell(row=row, column=i + 2, value=data.get("total_volumes", 0))
    _style_row(ws, row, max_col, fill=TOTAL_FILL, font=TOTAL_FONT)
    row += 2

    # ─── CHIFFRE D'AFFAIRES ───
    ws.cell(row=row, column=1, value=f"CHIFFRE D'AFFAIRES (FCFA) - {agence_nom}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)
    row += 1

    ca_lines = [
        ("(A) CA VENTE EAU", "total_ve"),
        ("  dont CA auto (Vol x PU)", "total_ca_auto"),
        ("  dont CA spécifiques", "total_ca_spec"),
        ("  dont Locations compteurs", "total_locations"),
        ("(B) TRAVAUX REMBOURSABLES", "total_trvx_remb"),
        ("  Brchts neufs", "brchts_neufs"),
        ("  Pénalités", "penalites"),
        ("  Dévis ext.", "devis_ext"),
        ("  Autres travaux", "autres_trvx"),
        ("  dont Complément manuel", "complement"),
        ("  Fraudes", "fraudes_trvx"),
        ("CA GLOBAL (A)+(B)", "ca_global"),
    ]
    for label, key in ca_lines:
        ws.cell(row=row, column=1, value=label)
        cumul_val = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = calcul_ca_agence(agence_id, m, exercice)
            val = data.get(key, 0)
            ws.cell(row=row, column=i + 2, value=val)
            cumul_val += val
        ws.cell(row=row, column=max_col, value=cumul_val)
        is_total = key in ("total_ve", "total_trvx_remb", "ca_global")
        _style_row(ws, row, max_col,
                   fill=TOTAL_FILL if is_total else None,
                   font=TOTAL_FONT if is_total else None)
        row += 1
    row += 1

    # ─── ENCAISSEMENTS ───
    ws.cell(row=row, column=1, value=f"ENCAISSEMENTS (FCFA) - {agence_nom}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)
    row += 1

    enc_sections = [
        ("Caisse Commerciale", "cciale", RUBRIQUES_ENC_CCIALE),
        ("Caisse ADM", "adm", RUBRIQUES_ENC_ADM),
        ("Banques", "banques", RUBRIQUES_ENC_BANQUES),
        ("Travaux Cciale", "trvx_cciale", RUBRIQUES_ENC_TRVX_CCIALE),
        ("Travaux ADM", "trvx_adm", RUBRIQUES_ENC_TRVX_ADM),
    ]
    for sec_label, sec_key, rubriques in enc_sections:
        ws.cell(row=row, column=1, value=sec_label)
        _style_row(ws, row, max_col, fill=SECTION_FILL, font=TOTAL_FONT)
        row += 1
        for rub in rubriques:
            ws.cell(row=row, column=1, value=rub)
            cumul_val = 0
            for i, m in enumerate(range(mois_debut, mois_fin + 1)):
                db = get_db()
                r = db.execute("SELECT montant FROM encaissements WHERE agence_id=? AND mois=? AND exercice=? AND section=? AND rubrique=?",
                               (agence_id, m, exercice, sec_key, rub)).fetchone()
                db.close()
                val = r['montant'] if r else 0
                ws.cell(row=row, column=i + 2, value=val)
                cumul_val += val
            ws.cell(row=row, column=max_col, value=cumul_val)
            _style_row(ws, row, max_col)
            row += 1

    # Total encaissements
    ws.cell(row=row, column=1, value="TOTAL ENCAISSEMENTS")
    for i, m in enumerate(range(mois_debut, mois_fin + 1)):
        data = calcul_ca_agence(agence_id, m, exercice)
        ws.cell(row=row, column=i + 2, value=data.get("total_encaissements", 0))
    _style_row(ws, row, max_col, fill=TOTAL_FILL, font=TOTAL_FONT)

    ws.column_dimensions['A'].width = 30
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15


# ═══════════════════════════════════════════════════════════════
# Export principal : Consolidation
# ═══════════════════════════════════════════════════════════════

def export_consolidation(mois_debut, mois_fin, exercice=2026, site_type='national', site_id=None):
    """Exporte les données consolidées en fichier Excel."""
    wb = Workbook()
    wb.remove(wb.active)

    if site_type == 'agence':
        db = get_db()
        ag = db.execute("SELECT id, nom FROM agences WHERE id=?", (site_id,)).fetchone()
        db.close()
        if ag:
            ws = wb.create_sheet()
            _write_agence_sheet(ws, ag['id'], ag['nom'], mois_debut, mois_fin, exercice)
    elif site_type == 'dr':
        agences = get_agences_dr(site_id)
        for ag_id, ag_nom in agences:
            ws = wb.create_sheet()
            _write_agence_sheet(ws, ag_id, ag_nom, mois_debut, mois_fin, exercice)
        # Feuille synthèse DR
        ws = wb.create_sheet(title=f"Synthèse {site_id}")
        _write_synthese_dr(ws, site_id, mois_debut, mois_fin, exercice)
    else:
        # National : une feuille par DR + feuille nationale
        for dr_code in STRUCTURE_CW:
            ws = wb.create_sheet(title=f"Synthèse {dr_code}")
            _write_synthese_dr(ws, dr_code, mois_debut, mois_fin, exercice)
        ws = wb.create_sheet(title="Ensemble CAMWATER")
        _write_synthese_national(ws, mois_debut, mois_fin, exercice)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_synthese_dr(ws, dr_code, mois_debut, mois_fin, exercice):
    """Feuille synthèse agrégée pour une DR."""
    cols = ["Indicateur"] + [MOIS[m - 1] for m in range(mois_debut, mois_fin + 1)] + ["Cumul"]
    max_col = len(cols)

    row = 1
    ws.cell(row=row, column=1, value=f"SYNTHÈSE {dr_code}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    indicators = [
        ("Total Volumes (m3)", "total_volumes"),
        ("CA Vente Eau (A)", "total_ve"),
        ("CA Travaux Remb. (B)", "total_trvx_remb"),
        ("CA Global (A)+(B)", "ca_global"),
        ("Encaissements VE", "enc_vte_eau"),
        ("Encaissements Travaux", "enc_travaux"),
        ("Total Encaissements", "total_encaissements"),
        ("Paiements Électroniques", "enc_electroniques"),
        ("Facturation EF", "facturation_ef"),
        ("Recouvrement EF", "recouvrement_ef"),
    ]

    row = 3
    for label, key in indicators:
        ws.cell(row=row, column=1, value=label)
        cumul_val = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = calcul_dr(dr_code, m, exercice) or {}
            val = data.get(key, 0)
            ws.cell(row=row, column=i + 2, value=val)
            cumul_val += val
        ws.cell(row=row, column=max_col, value=cumul_val)
        is_total = key in ("ca_global", "total_encaissements")
        _style_row(ws, row, max_col,
                   fill=TOTAL_FILL if is_total else None,
                   font=TOTAL_FONT if is_total else None)
        row += 1

    # Taux (non cumulables)
    taux_lines = [
        ("Taux encaissement VE", "taux_enc_ve"),
        ("Taux encaissement Travaux", "taux_enc_trvx"),
        ("Taux encaissement Global", "taux_enc_global"),
        ("Taux recouvrement EF", "taux_recouvrement_ef"),
    ]
    for label, key in taux_lines:
        ws.cell(row=row, column=1, value=label)
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = calcul_dr(dr_code, m, exercice) or {}
            cell = ws.cell(row=row, column=i + 2, value=data.get(key, 0))
            cell.number_format = PCT_FMT
        # Cumul = taux sur la période complète
        cumul_data = calcul_cumul_dr(dr_code, mois_debut, mois_fin, exercice)
        cell = ws.cell(row=row, column=max_col, value=cumul_data.get(key, 0))
        cell.number_format = PCT_FMT
        _style_row(ws, row, max_col)
        row += 1

    ws.column_dimensions['A'].width = 30
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15


def _write_synthese_national(ws, mois_debut, mois_fin, exercice):
    """Feuille synthèse nationale."""
    cols = ["Indicateur"] + [MOIS[m - 1] for m in range(mois_debut, mois_fin + 1)] + ["Cumul"]
    max_col = len(cols)

    row = 1
    ws.cell(row=row, column=1, value="ENSEMBLE CAMWATER")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    indicators = [
        ("Total Volumes (m3)", "total_volumes"),
        ("CA Vente Eau (A)", "total_ve"),
        ("CA Travaux Remb. (B)", "total_trvx_remb"),
        ("CA Global (A)+(B)", "ca_global"),
        ("Encaissements VE", "enc_vte_eau"),
        ("Encaissements Travaux", "enc_travaux"),
        ("Total Encaissements", "total_encaissements"),
        ("Paiements Électroniques", "enc_electroniques"),
    ]

    row = 3
    for label, key in indicators:
        ws.cell(row=row, column=1, value=label)
        cumul_val = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = calcul_national(m, exercice)
            data.pop("details_dr", None)
            val = data.get(key, 0)
            ws.cell(row=row, column=i + 2, value=val)
            cumul_val += val
        ws.cell(row=row, column=max_col, value=cumul_val)
        is_total = key in ("ca_global", "total_encaissements")
        _style_row(ws, row, max_col,
                   fill=TOTAL_FILL if is_total else None,
                   font=TOTAL_FONT if is_total else None)
        row += 1

    ws.column_dimensions['A'].width = 30
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15


# ═══════════════════════════════════════════════════════════════
# Exports spécialisés (Budget, Fiscal, Reporting)
# ═══════════════════════════════════════════════════════════════

def export_budget(mois_debut, mois_fin, exercice=2026, site_type='national', site_id=None):
    """Export données budget en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Budget"
    cols = ["Rubrique"] + [MOIS[m - 1] for m in range(mois_debut, mois_fin + 1)] + ["Cumul"]
    max_col = len(cols)

    row = 1
    ws.cell(row=row, column=1, value="DONNÉES BUDGET")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    budget_keys = [
        ("Vente eau TTC", "vente_eau"),
        ("Travaux remboursables", "trvx_remb"),
        ("Recouvrement impayés", "recouvrement_impayes"),
        ("Fraude", "fraude"),
        ("Sinistre", "sinistre"),
        ("Pénalités", "penalites"),
        ("Locations TTC", "locations"),
        ("Branchements", "branchements"),
    ]

    row = 3
    for label, key in budget_keys:
        ws.cell(row=row, column=1, value=label)
        cumul_val = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = _get_site_data(site_type, site_id, m, exercice)
            val = data.get("budget", {}).get(key, 0)
            ws.cell(row=row, column=i + 2, value=val)
            cumul_val += val
        ws.cell(row=row, column=max_col, value=cumul_val)
        _style_row(ws, row, max_col)
        row += 1

    ws.column_dimensions['A'].width = 30
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_fiscal(mois_debut, mois_fin, exercice=2026, site_type='national', site_id=None):
    """Export données fiscales en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Fiscales"
    cols = ["Rubrique"] + [MOIS[m - 1] for m in range(mois_debut, mois_fin + 1)] + ["Cumul"]
    max_col = len(cols)

    row = 1
    ws.cell(row=row, column=1, value="DONNÉES FISCALES")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    fiscal_keys = [
        ("Tranche sociale", "tranche_sociale"),
        ("CA ADM HT", "ca_adm_ht"),
        ("CA Hors ADM HT", "ca_hors_adm_ht"),
        ("CA Total TTC", "ca_total_ttc"),
        ("Travaux TTC", "trvx_ttc"),
    ]

    row = 3
    for label, key in fiscal_keys:
        ws.cell(row=row, column=1, value=label)
        cumul_val = 0
        for i, m in enumerate(range(mois_debut, mois_fin + 1)):
            data = _get_site_data(site_type, site_id, m, exercice)
            val = data.get(key, 0)
            ws.cell(row=row, column=i + 2, value=val)
            cumul_val += val
        ws.cell(row=row, column=max_col, value=cumul_val)
        _style_row(ws, row, max_col)
        row += 1

    ws.column_dimensions['A'].width = 30
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_reporting(mois_debut, mois_fin, exercice=2026, site_type='national', site_id=None):
    """Export données reporting en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporting Commercial"
    cols = ["DR"] + ["CA Global", "Encaissements", "Volumes", "Taux Enc. VE",
                     "Taux Enc. Trvx", "Taux Enc. Global", "Taux Recouv. EF",
                     "Paiements Élec. %"]
    max_col = len(cols)

    row = 1
    period = f"{MOIS[mois_debut - 1]} à {MOIS[mois_fin - 1]} {exercice}"
    ws.cell(row=row, column=1, value=f"REPORTING COMMERCIAL - {period}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row = 2
    for c, label in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=label)
    _style_header(ws, row, max_col)

    row = 3
    from calculs import calcul_cumul_dr, calcul_cumul_national
    for dr_code in STRUCTURE_CW:
        d = calcul_cumul_dr(dr_code, mois_debut, mois_fin, exercice)
        ws.cell(row=row, column=1, value=dr_code)
        ws.cell(row=row, column=2, value=d.get("ca_global", 0))
        ws.cell(row=row, column=3, value=d.get("total_encaissements", 0))
        ws.cell(row=row, column=4, value=d.get("total_volumes", 0))
        ws.cell(row=row, column=5, value=d.get("taux_enc_ve", 0)).number_format = PCT_FMT
        ws.cell(row=row, column=6, value=d.get("taux_enc_trvx", 0)).number_format = PCT_FMT
        ws.cell(row=row, column=7, value=d.get("taux_enc_global", 0)).number_format = PCT_FMT
        ws.cell(row=row, column=8, value=d.get("taux_recouvrement_ef", 0)).number_format = PCT_FMT
        ws.cell(row=row, column=9, value=d.get("pct_paiements_elec", 0)).number_format = PCT_FMT
        _style_row(ws, row, max_col)
        row += 1

    # Total national
    d = calcul_cumul_national(mois_debut, mois_fin, exercice)
    ws.cell(row=row, column=1, value="ENSEMBLE CAMWATER")
    ws.cell(row=row, column=2, value=d.get("ca_global", 0))
    ws.cell(row=row, column=3, value=d.get("total_encaissements", 0))
    ws.cell(row=row, column=4, value=d.get("total_volumes", 0))
    ws.cell(row=row, column=5, value=d.get("taux_enc_ve", 0)).number_format = PCT_FMT
    ws.cell(row=row, column=6, value=d.get("taux_enc_trvx", 0)).number_format = PCT_FMT
    ws.cell(row=row, column=7, value=d.get("taux_enc_global", 0)).number_format = PCT_FMT
    ws.cell(row=row, column=8, value=d.get("taux_recouvrement_ef", 0)).number_format = PCT_FMT
    ws.cell(row=row, column=9, value=d.get("pct_paiements_elec", 0)).number_format = PCT_FMT
    _style_row(ws, row, max_col, fill=TOTAL_FILL, font=TOTAL_FONT)

    ws.column_dimensions['A'].width = 22
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _get_site_data(site_type, site_id, mois, exercice):
    """Helper : récupère les données d'un site pour un mois."""
    if site_type == 'agence':
        return calcul_ca_agence(int(site_id), mois, exercice)
    elif site_type == 'dr':
        return calcul_dr(site_id, mois, exercice) or {}
    else:
        data = calcul_national(mois, exercice)
        data.pop("details_dr", None)
        return data
