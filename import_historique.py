"""
Moteur d'import historique CAMWATER v2
======================================
Principes de conception :
  • 100 % éphémère : traitement exclusivement en mémoire (io.BytesIO).
    Aucun fichier source n'est jamais écrit sur disque.
  • Transactionnel : tout ou rien. En cas d'erreur sur une ligne,
    la transaction entière est annulée (rollback).
  • Idempotent : un UPSERT (INSERT OR REPLACE) garantit qu'un re-import
    écrase proprement les données existantes sans doublon.
  • Robuste : chaque fonction isole ses exceptions et les rapporte
    clairement plutôt que de propager un crash.
  • Flexible : détection automatique de format (colonnes, variantes de noms).
"""

import io
import re
import unicodedata
from typing import Any

import openpyxl

from database import (
    get_db, STRUCTURE_CW, MOIS,
    RUBRIQUES_ENC_CCIALE, RUBRIQUES_ENC_ADM,
    RUBRIQUES_ENC_BANQUES, RUBRIQUES_ENC_TRVX_CCIALE, RUBRIQUES_ENC_TRVX_ADM,
    TYPES_BRANCHEMENTS,
)

# ── Normalisation ──────────────────────────────────────────────────────────────

def _slug(s: str) -> str:
    """Normalise une chaîne pour la comparaison tolérante : minuscules, sans
    accents, sans caractères spéciaux, espaces → underscore."""
    if not isinstance(s, str):
        return ''
    nfd = unicodedata.normalize('NFD', s)
    ascii_ = nfd.encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '_', ascii_.lower()).strip('_')


def _num(v: Any) -> object:
    """Convertit une valeur cellule en float. Retourne None si non numérique."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        cleaned = re.sub(r'[^\d.,-]', '', v.replace(',', '.'))
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None
    return None


def _safe_int(v: Any, default: int = 0) -> int:
    n = _num(v)
    return int(n) if n is not None else default


# ── Index agences ──────────────────────────────────────────────────────────────

def _build_agence_index(db) -> dict:
    """Construit {slug_nom → id} et {slug_code_dr_nom → id} pour résolution flexible."""
    rows = db.execute(
        "SELECT a.id, a.nom, d.code as dr_code "
        "FROM agences a JOIN directions_regionales d ON a.dr_id = d.id"
    ).fetchall()
    idx: dict = {}
    for r in rows:
        key1 = _slug(r['nom'])
        idx[key1] = r['id']
        key2 = _slug(r['dr_code']) + '__' + _slug(r['nom'])
        idx[key2] = r['id']
        # variante : code DR seul si agence unique dans cette DR (utile pour petites DRs)
    return idx


def _resolve_agence(nom: str, dr: object, index: dict) -> object:
    if not nom:
        return None
    if dr:
        key = _slug(str(dr)) + '__' + _slug(str(nom))
        if key in index:
            return index[key]
    key = _slug(str(nom))
    return index.get(key)


# ── Parsing du mois ────────────────────────────────────────────────────────────

_MOIS_SLUGS = {_slug(m): i + 1 for i, m in enumerate(MOIS)}
_MOIS_ABBREV = {
    'jan': 1, 'janv': 1, 'feb': 2, 'fev': 2, 'fevr': 2, 'mar': 3, 'mars': 3,
    'avr': 4, 'apr': 4, 'mai': 5, 'may': 5, 'jun': 6, 'juin': 6,
    'jul': 7, 'juil': 7, 'aou': 8, 'aug': 8, 'aout': 8,
    'sep': 9, 'sept': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _parse_mois(val: Any) -> object:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if 1 <= v <= 12 else None
    s = _slug(str(val))
    if s in _MOIS_SLUGS:
        return _MOIS_SLUGS[s]
    # essai abréviation (3 premiers chars)
    for abbrev, num in _MOIS_ABBREV.items():
        if s.startswith(abbrev):
            return num
    # essai numérique
    try:
        v = int(re.sub(r'\D', '', str(val)))
        return v if 1 <= v <= 12 else None
    except ValueError:
        return None


# ── Noms de colonnes flexibles ─────────────────────────────────────────────────

_CAT_ALIASES: dict = {}  # slug → catégorie exacte — construit dynamiquement


def _build_cat_aliases(db) -> dict:
    cats = [r['categorie'] for r in
            db.execute("SELECT DISTINCT categorie FROM prix_unitaires").fetchall()]
    return {_slug(c): c for c in cats}


_ENC_SECTION_MAP = {
    'cciale': 'Cciale', 'commerciale': 'Cciale', 'ciale': 'Cciale',
    'adm': 'ADM', 'admin': 'ADM', 'administrative': 'ADM',
    'banques': 'Banques', 'banque': 'Banques',
    'trvx_cciale': 'Trvx Cciale', 'travaux_cciale': 'Trvx Cciale',
    'trvx_adm': 'Trvx ADM', 'travaux_adm': 'Trvx ADM',
}


# ══ Parsers par type de données ════════════════════════════════════════════════


def _parse_sheet_volumes(ws, agence_idx, cat_aliases, exercice) -> tuple:
    """Parse un onglet Volumes.
    Format attendu : ligne 1 = en-têtes [DR, Agence, Mois, <cat1>, <cat2>, ...]
    Retourne (lignes_valides, erreurs).
    """
    rows_valid = []
    errors = []
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    col_dr = col_ag = col_mois = None
    col_cats = []

    for i, h in enumerate(headers):
        s = _slug(h)
        if s in ('dr', 'direction', 'direction_regionale'):
            col_dr = i
        elif s in ('agence', 'centre', 'agence_centre'):
            col_ag = i
        elif s in ('mois', 'mois_1', 'periode', 'period'):
            col_mois = i
        else:
            cat = cat_aliases.get(s)
            if cat:
                col_cats.append((i, cat))

    if col_ag is None:
        errors.append("Colonne 'Agence' introuvable dans l'onglet Volumes")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None else None
        ag_val = row[col_ag] if col_ag is not None else None
        mois_val = row[col_mois] if col_mois is not None else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence inconnue '{ag_val}' (DR={dr_val})")
            continue

        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue

        for col_idx, cat in col_cats:
            val = _num(row[col_idx]) if col_idx < len(row) else None
            if val is not None:
                rows_valid.append((agence_id, mois, exercice, cat, val))

    return rows_valid, errors


def _parse_sheet_encaissements(ws, agence_idx, exercice) -> tuple:
    """Format : DR | Agence | Mois | Section | Rubrique | Montant"""
    rows_valid = []
    errors = []
    all_rubriques = (
        RUBRIQUES_ENC_CCIALE + RUBRIQUES_ENC_ADM + RUBRIQUES_ENC_BANQUES +
        RUBRIQUES_ENC_TRVX_CCIALE + RUBRIQUES_ENC_TRVX_ADM
    )
    rubrique_idx = {_slug(r): r for r in all_rubriques}

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {_slug(h): i for i, h in enumerate(headers)}

    col_dr = col.get('dr') or col.get('direction_regionale')
    col_ag = col.get('agence') or col.get('centre')
    col_mois = col.get('mois') or col.get('periode')
    col_sec = col.get('section') or col.get('type')
    col_rub = col.get('rubrique') or col.get('libelle')
    col_mt = col.get('montant') or col.get('valeur') or col.get('montant_fcfa')

    if col_ag is None or col_mt is None:
        errors.append("Colonnes minimales manquantes dans Encaissements (Agence, Montant)")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None and col_dr < len(row) else None
        ag_val = row[col_ag] if col_ag < len(row) else None
        mois_val = row[col_mois] if col_mois is not None and col_mois < len(row) else None
        sec_val = row[col_sec] if col_sec is not None and col_sec < len(row) else ''
        rub_val = row[col_rub] if col_rub is not None and col_rub < len(row) else ''
        mt_val = row[col_mt] if col_mt < len(row) else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence '{ag_val}' inconnue")
            continue
        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue
        montant = _num(mt_val)
        if montant is None:
            continue

        # Normaliser section
        sec_slug = _slug(str(sec_val or ''))
        section = _ENC_SECTION_MAP.get(sec_slug, str(sec_val or '').strip() or 'Cciale')

        # Normaliser rubrique
        rub_slug = _slug(str(rub_val or ''))
        rubrique = rubrique_idx.get(rub_slug, str(rub_val or '').strip())

        rows_valid.append((agence_id, mois, exercice, section, rubrique, montant))

    return rows_valid, errors


def _parse_sheet_ca_spec(ws, agence_idx, exercice) -> tuple:
    """Format : DR | Agence | Mois | Rubrique | Montant"""
    rows_valid = []
    errors = []
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {_slug(h): i for i, h in enumerate(headers)}

    col_dr = col.get('dr')
    col_ag = col.get('agence') or col.get('centre')
    col_mois = col.get('mois') or col.get('periode')
    col_rub = col.get('rubrique') or col.get('libelle')
    col_mt = col.get('montant') or col.get('valeur')

    if col_ag is None or col_mt is None:
        errors.append("Colonnes minimales manquantes dans CA_Spec (Agence, Montant)")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None and col_dr < len(row) else None
        ag_val = row[col_ag] if col_ag < len(row) else None
        mois_val = row[col_mois] if col_mois is not None and col_mois < len(row) else None
        rub_val = row[col_rub] if col_rub is not None and col_rub < len(row) else ''
        mt_val = row[col_mt] if col_mt < len(row) else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence '{ag_val}' inconnue")
            continue
        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue
        montant = _num(mt_val)
        if montant is None:
            continue

        rows_valid.append((agence_id, mois, exercice, str(rub_val or '').strip(), montant))

    return rows_valid, errors


def _parse_sheet_branchements(ws, agence_idx, exercice) -> tuple:
    """Format : DR | Agence | Mois | Type | Valeur
    Type : vendus / exécutés / pec / moratoire"""
    rows_valid = []
    errors = []
    type_idx = {_slug(t): t for t in TYPES_BRANCHEMENTS}

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {_slug(h): i for i, h in enumerate(headers)}

    col_dr = col.get('dr')
    col_ag = col.get('agence') or col.get('centre')
    col_mois = col.get('mois') or col.get('periode')
    col_type = col.get('type') or col.get('rubrique')
    col_val = col.get('valeur') or col.get('nombre') or col.get('quantite')

    if col_ag is None or col_val is None:
        errors.append("Colonnes minimales manquantes dans Branchements (Agence, Valeur)")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None and col_dr < len(row) else None
        ag_val = row[col_ag] if col_ag < len(row) else None
        mois_val = row[col_mois] if col_mois is not None and col_mois < len(row) else None
        type_val = row[col_type] if col_type is not None and col_type < len(row) else 'vendus'
        val_cell = row[col_val] if col_val < len(row) else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence '{ag_val}' inconnue")
            continue
        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue
        valeur = _num(val_cell)
        if valeur is None:
            continue

        type_norm = type_idx.get(_slug(str(type_val or '')), str(type_val or '').strip())
        rows_valid.append((agence_id, mois, exercice, type_norm, valeur))

    return rows_valid, errors


def _parse_sheet_impayes(ws, agence_idx, exercice) -> tuple:
    """Format : DR | Agence | Mois | Part_Actifs | GCO_Actifs | Part_Résiliés |
                GCO_Résiliés | BF_Actifs | BFC_Actifs | BF_Résiliés | BFC_Résiliés |
                Gestion_manuelle | Résiliés_créditeurs"""
    rows_valid = []
    errors = []

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {_slug(h): i for i, h in enumerate(headers)}

    col_dr = col.get('dr')
    col_ag = col.get('agence') or col.get('centre')
    col_mois = col.get('mois') or col.get('periode')

    _F = {
        'particuliers_actifs': ['particuliers_actifs', 'part_actifs', 'actifs_part'],
        'gco_actifs': ['gco_actifs', 'actifs_gco'],
        'particuliers_resilies': ['particuliers_resilies', 'part_resilies', 'resilies_part'],
        'gco_resilies': ['gco_resilies', 'resilies_gco'],
        'bf_actifs': ['bf_actifs', 'borne_fontaine_actifs'],
        'bfc_actifs': ['bfc_actifs'],
        'bf_resilies': ['bf_resilies'],
        'bfc_resilies': ['bfc_resilies'],
        'gestion_manuelle': ['gestion_manuelle', 'manuel'],
        'resiliers_crediteurs': ['resiliers_crediteurs', 'crediteurs'],
    }

    col_fields = {}
    for field, aliases in _F.items():
        for alias in aliases:
            if alias in col:
                col_fields[field] = col[alias]
                break

    if col_ag is None:
        errors.append("Colonne 'Agence' manquante dans Impayés")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None and col_dr < len(row) else None
        ag_val = row[col_ag] if col_ag < len(row) else None
        mois_val = row[col_mois] if col_mois is not None and col_mois < len(row) else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence '{ag_val}' inconnue")
            continue
        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue

        vals = {}
        for field, ci in col_fields.items():
            vals[field] = _num(row[ci]) if ci < len(row) else None

        rows_valid.append((
            agence_id, mois, exercice,
            vals.get('particuliers_actifs'), vals.get('gco_actifs'),
            vals.get('particuliers_resilies'), vals.get('gco_resilies'),
            vals.get('bf_actifs'), vals.get('bfc_actifs'),
            vals.get('bf_resilies'), vals.get('bfc_resilies'),
            vals.get('gestion_manuelle'), vals.get('resiliers_crediteurs'),
        ))

    return rows_valid, errors


def _parse_sheet_complements(ws, agence_idx, exercice) -> tuple:
    """Format : DR | Agence | Mois | Montant"""
    rows_valid = []
    errors = []
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {_slug(h): i for i, h in enumerate(headers)}

    col_dr = col.get('dr')
    col_ag = col.get('agence') or col.get('centre')
    col_mois = col.get('mois') or col.get('periode')
    col_mt = col.get('montant') or col.get('complement') or col.get('valeur')

    if col_ag is None or col_mt is None:
        errors.append("Colonnes minimales manquantes dans Compléments Travaux")
        return rows_valid, errors

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        dr_val = row[col_dr] if col_dr is not None and col_dr < len(row) else None
        ag_val = row[col_ag] if col_ag < len(row) else None
        mois_val = row[col_mois] if col_mois is not None and col_mois < len(row) else None
        mt_val = row[col_mt] if col_mt < len(row) else None

        agence_id = _resolve_agence(str(ag_val or ''), str(dr_val or ''), agence_idx)
        if not agence_id:
            errors.append(f"L{row_num} : agence '{ag_val}' inconnue")
            continue
        mois = _parse_mois(mois_val)
        if not mois:
            errors.append(f"L{row_num} : mois invalide '{mois_val}'")
            continue
        montant = _num(mt_val)
        if montant is None:
            continue

        rows_valid.append((agence_id, mois, exercice, montant))

    return rows_valid, errors


# ══ Mapping noms d'onglets ══════════════════════════════════════════════════════

_SHEET_DISPATCH = {
    'volumes': 'volumes',
    'volume': 'volumes',
    'vol': 'volumes',
    'encaissements': 'encaissements',
    'encaissement': 'encaissements',
    'enc': 'encaissements',
    'ca_specifiques': 'ca_spec',
    'ca_spec': 'ca_spec',
    'ca_complementaires': 'ca_spec',
    'specifiques': 'ca_spec',
    'branchements': 'branchements',
    'brcht': 'branchements',
    'bts': 'branchements',
    'impayes': 'impayes',
    'impaye': 'impayes',
    'imp': 'impayes',
    'complements_travaux': 'complements',
    'complements': 'complements',
    'complement_travaux': 'complements',
    'compl': 'complements',
}


# ══ Point d'entrée principal ═══════════════════════════════════════════════════

def importer_fichier(file_bytes: bytes, exercice: int) -> dict:
    """
    Importe les données historiques depuis un fichier Excel en mémoire.

    Paramètres :
        file_bytes  : contenu brut du fichier (jamais écrit sur disque)
        exercice    : année de l'exercice (ex. 2026)

    Retour :
        {
          "success": bool,
          "stats": {"volumes": N, "encaissements": N, ...},
          "errors": ["..."],
          "warnings": ["..."],
        }
    """
    stats = {'volumes': 0, 'encaissements': 0, 'ca_spec': 0,
             'branchements': 0, 'impayes': 0, 'complements': 0}
    errors: list = []
    warnings: list = []

    # ── Lecture en mémoire — aucun disque impliqué ─────────────────────────────
    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        return {
            'success': False,
            'stats': stats,
            'errors': [f"Fichier Excel invalide ou corrompu : {exc}"],
            'warnings': [],
        }

    db = get_db()
    try:
        agence_idx = _build_agence_index(db)
        cat_aliases = _build_cat_aliases(db)

        parsed: dict[str, list] = {k: [] for k in stats}
        sheet_errors: list = []

        for sheet_name in wb.sheetnames:
            key = _slug(sheet_name)
            sheet_type = _SHEET_DISPATCH.get(key)
            if not sheet_type:
                # Essai préfixe
                for slug, stype in _SHEET_DISPATCH.items():
                    if key.startswith(slug) or slug.startswith(key[:4]):
                        sheet_type = stype
                        break
            if not sheet_type:
                warnings.append(f"Onglet '{sheet_name}' non reconnu — ignoré")
                continue

            ws = wb[sheet_name]
            if sheet_type == 'volumes':
                rows, errs = _parse_sheet_volumes(ws, agence_idx, cat_aliases, exercice)
            elif sheet_type == 'encaissements':
                rows, errs = _parse_sheet_encaissements(ws, agence_idx, exercice)
            elif sheet_type == 'ca_spec':
                rows, errs = _parse_sheet_ca_spec(ws, agence_idx, exercice)
            elif sheet_type == 'branchements':
                rows, errs = _parse_sheet_branchements(ws, agence_idx, exercice)
            elif sheet_type == 'impayes':
                rows, errs = _parse_sheet_impayes(ws, agence_idx, exercice)
            elif sheet_type == 'complements':
                rows, errs = _parse_sheet_complements(ws, agence_idx, exercice)
            else:
                continue

            parsed[sheet_type].extend(rows)
            sheet_errors.extend([f"[{sheet_name}] {e}" for e in errs])

        errors.extend(sheet_errors)

        if not any(parsed.values()):
            return {
                'success': False, 'stats': stats,
                'errors': errors or ["Aucune donnée exploitable trouvée dans le fichier"],
                'warnings': warnings,
            }

        # ── Insertion transactionnelle (tout ou rien) ──────────────────────────
        db.execute('BEGIN')
        try:
            # Volumes
            for row in parsed['volumes']:
                db.execute(
                    "INSERT OR REPLACE INTO volumes "
                    "(agence_id, mois, exercice, categorie, valeur) VALUES (?,?,?,?,?)",
                    row,
                )
            stats['volumes'] = len(parsed['volumes'])

            # Encaissements
            for row in parsed['encaissements']:
                db.execute(
                    "INSERT OR REPLACE INTO encaissements "
                    "(agence_id, mois, exercice, section, rubrique, montant) VALUES (?,?,?,?,?,?)",
                    row,
                )
            stats['encaissements'] = len(parsed['encaissements'])

            # CA spécifiques
            for row in parsed['ca_spec']:
                db.execute(
                    "INSERT OR REPLACE INTO ca_specifiques "
                    "(agence_id, mois, exercice, rubrique, montant) VALUES (?,?,?,?,?)",
                    row,
                )
            stats['ca_spec'] = len(parsed['ca_spec'])

            # Branchements
            for row in parsed['branchements']:
                db.execute(
                    "INSERT OR REPLACE INTO branchements "
                    "(agence_id, mois, exercice, type, valeur) VALUES (?,?,?,?,?)",
                    row,
                )
            stats['branchements'] = len(parsed['branchements'])

            # Impayés
            for row in parsed['impayes']:
                db.execute(
                    """INSERT OR REPLACE INTO impayes
                    (agence_id, mois, exercice,
                     particuliers_actifs, gco_actifs, particuliers_resilies, gco_resilies,
                     bf_actifs, bfc_actifs, bf_resilies, bfc_resilies,
                     gestion_manuelle, resiliers_crediteurs)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    row,
                )
            stats['impayes'] = len(parsed['impayes'])

            # Compléments travaux
            for row in parsed['complements']:
                db.execute(
                    "INSERT OR REPLACE INTO complements_travaux "
                    "(agence_id, mois, exercice, montant) VALUES (?,?,?,?)",
                    row,
                )
            stats['complements'] = len(parsed['complements'])

            db.execute('COMMIT')

        except Exception as exc:
            db.execute('ROLLBACK')
            raise exc

    except Exception as exc:
        return {
            'success': False,
            'stats': stats,
            'errors': [f"Erreur lors de l'insertion : {exc}"] + errors,
            'warnings': warnings,
        }
    finally:
        db.close()
        wb.close()
        # file_bytes libéré par le GC (jamais persisté sur disque)

    total = sum(stats.values())
    return {
        'success': True,
        'stats': stats,
        'total_lignes': total,
        'errors': errors,
        'warnings': warnings,
        'message': (
            f"{total} enregistrement(s) importé(s) avec succès "
            f"(volumes:{stats['volumes']}, enc:{stats['encaissements']}, "
            f"ca:{stats['ca_spec']}, bts:{stats['branchements']}, "
            f"imp:{stats['impayes']}, compl:{stats['complements']})"
        ),
    }


# ══ Prévisualisation (sans écriture DB) ════════════════════════════════════════

def previsualiser_fichier(file_bytes: bytes, exercice: int) -> dict:
    """Analyse le fichier et retourne un résumé SANS toucher à la base.
    Identique à importer_fichier() mais s'arrête avant l'INSERT."""
    stats = {'volumes': 0, 'encaissements': 0, 'ca_spec': 0,
             'branchements': 0, 'impayes': 0, 'complements': 0}
    errors: list = []
    warnings: list = []

    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        return {'success': False, 'errors': [str(exc)], 'stats': stats, 'warnings': []}

    db = get_db()
    try:
        agence_idx = _build_agence_index(db)
        cat_aliases = _build_cat_aliases(db)

        for sheet_name in wb.sheetnames:
            key = _slug(sheet_name)
            sheet_type = _SHEET_DISPATCH.get(key)
            if not sheet_type:
                warnings.append(f"Onglet '{sheet_name}' non reconnu")
                continue
            ws = wb[sheet_name]
            if sheet_type == 'volumes':
                rows, errs = _parse_sheet_volumes(ws, agence_idx, cat_aliases, exercice)
            elif sheet_type == 'encaissements':
                rows, errs = _parse_sheet_encaissements(ws, agence_idx, exercice)
            elif sheet_type == 'ca_spec':
                rows, errs = _parse_sheet_ca_spec(ws, agence_idx, exercice)
            elif sheet_type == 'branchements':
                rows, errs = _parse_sheet_branchements(ws, agence_idx, exercice)
            elif sheet_type == 'impayes':
                rows, errs = _parse_sheet_impayes(ws, agence_idx, exercice)
            elif sheet_type == 'complements':
                rows, errs = _parse_sheet_complements(ws, agence_idx, exercice)
            else:
                continue
            stats[sheet_type] += len(rows)
            errors.extend([f"[{sheet_name}] {e}" for e in errs])
    finally:
        db.close()
        wb.close()

    return {
        'success': True,
        'stats': stats,
        'total_lignes': sum(stats.values()),
        'errors': errors,
        'warnings': warnings,
    }


# ══ Générateur du template Excel téléchargeable ════════════════════════════════

def generer_template_excel(exercice: int = 2026) -> bytes:
    """Génère un fichier Excel modèle prêt à remplir, en mémoire."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    cats = [r['categorie'] for r in
            db.execute("SELECT DISTINCT categorie FROM prix_unitaires ORDER BY id").fetchall()]
    db.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill('solid', fgColor='0D2B5E')
    HDR_FONT = Font(color='FFFFFF', bold=True, name='Raleway', size=10)
    HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    EXAMPLE_FONT = Font(color='5B7399', italic=True, size=9)
    thin = Side(style='thin', color='C8D9EE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _add_header(ws, headers):
        ws.row_dimensions[1].height = 36
        ws.append(headers)
        for i, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN
            cell.border = border
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _add_example(ws, row):
        ws.append(row)
        for i in range(1, len(row) + 1):
            cell = ws.cell(row=2, column=i)
            cell.font = EXAMPLE_FONT
            cell.border = border

    # ── Onglet Volumes ─────────────────────────────────────────────────────────
    ws_v = wb.create_sheet('Volumes')
    ws_v.sheet_properties.tabColor = '1B53AB'
    headers_v = ['DR', 'Agence', 'Mois'] + cats
    _add_header(ws_v, headers_v)
    _add_example(ws_v, ['DRYA', 'Ekounou', 'Janvier'] + [0.0] * len(cats))
    ws_v.freeze_panes = 'D2'

    # ── Onglet Encaissements ───────────────────────────────────────────────────
    ws_e = wb.create_sheet('Encaissements')
    ws_e.sheet_properties.tabColor = '5BACD8'
    _add_header(ws_e, ['DR', 'Agence', 'Mois', 'Section', 'Rubrique', 'Montant'])
    _add_example(ws_e, ['DRYA', 'Ekounou', 'Janvier', 'Cciale', 'Part. 1 & 2', 0.0])
    # Note d'aide sur les sections valides
    ws_e['H1'] = 'Sections valides: Cciale, ADM, Banques, Trvx Cciale, Trvx ADM'
    ws_e['H1'].font = Font(color='B91C1C', italic=True, size=8)

    # ── Onglet CA_Spec ─────────────────────────────────────────────────────────
    ws_c = wb.create_sheet('CA_Specifiques')
    ws_c.sheet_properties.tabColor = '219653'
    _add_header(ws_c, ['DR', 'Agence', 'Mois', 'Rubrique', 'Montant'])
    _add_example(ws_c, ['DRYA', 'Ekounou', 'Janvier', 'ENEO 871/873', 0.0])

    # ── Onglet Branchements ────────────────────────────────────────────────────
    ws_b = wb.create_sheet('Branchements')
    ws_b.sheet_properties.tabColor = '2E6CBF'
    _add_header(ws_b, ['DR', 'Agence', 'Mois', 'Type', 'Valeur'])
    _add_example(ws_b, ['DRYA', 'Ekounou', 'Janvier', 'vendus', 0])
    ws_b['G1'] = 'Types: vendus, exécutés, pec, moratoire'
    ws_b['G1'].font = Font(color='B91C1C', italic=True, size=8)

    # ── Onglet Impayés ─────────────────────────────────────────────────────────
    ws_i = wb.create_sheet('Impayes')
    ws_i.sheet_properties.tabColor = 'D68910'
    _add_header(ws_i, ['DR', 'Agence', 'Mois',
                        'Particuliers_Actifs', 'GCO_Actifs',
                        'Particuliers_Resilies', 'GCO_Resilies',
                        'BF_Actifs', 'BFC_Actifs', 'BF_Resilies', 'BFC_Resilies',
                        'Gestion_Manuelle', 'Resiliers_Crediteurs'])
    _add_example(ws_i, ['DRYA', 'Ekounou', 'Janvier'] + [0.0] * 10)

    # ── Onglet Compléments Travaux ─────────────────────────────────────────────
    ws_ct = wb.create_sheet('Complements_Travaux')
    ws_ct.sheet_properties.tabColor = 'B91C1C'
    _add_header(ws_ct, ['DR', 'Agence', 'Mois', 'Montant'])
    _add_example(ws_ct, ['DRYA', 'Ekounou', 'Janvier', 0.0])

    # ── Onglet README ──────────────────────────────────────────────────────────
    ws_r = wb.create_sheet('README', 0)  # en premier
    ws_r.sheet_properties.tabColor = '133672'
    ws_r['A1'] = f'CAMWATER — Template Import Historique — Exercice {exercice}'
    ws_r['A1'].font = Font(bold=True, size=13, color='0D2B5E')
    ws_r['A3'] = 'RÈGLES :'
    ws_r['A3'].font = Font(bold=True, color='B91C1C')
    instructions = [
        'Remplir uniquement les onglets concernés (les onglets vides sont ignorés).',
        'Colonne DR : code exact (DRYA, DRDA, DREN, DRN, DRA, DRC, DRS, DRE, DRO, DRNO, DRSO, DRL).',
        'Colonne Agence : nom exact de l\'agence (ex : Ekounou, Tsinga...).',
        'Colonne Mois : numéro (1-12) ou nom (Janvier, Février..., Jan, Fév...).',
        'Valeurs : chiffres uniquement, 0 si pas de données (ne pas laisser vide si la ligne est renseignée).',
        'Plusieurs mois peuvent figurer dans le même onglet.',
        'Un re-import écrase les données existantes (UPSERT — pas de doublon).',
        'Le fichier est détruit en mémoire après import — seules les données extraites sont conservées.',
    ]
    for i, inst in enumerate(instructions, start=4):
        ws_r[f'A{i}'] = f'  {i - 3}. {inst}'
        ws_r[f'A{i}'].font = Font(size=10)
    ws_r.column_dimensions['A'].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
