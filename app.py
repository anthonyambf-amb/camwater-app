"""Application Flask CAMWATER v2 - Données Commerciales."""
import os
import io
import logging
import traceback
from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for)
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename
from database import (init_db, get_db, STRUCTURE_CW, MOIS,
                       RUBRIQUES_ENC_CCIALE, RUBRIQUES_ENC_ADM, RUBRIQUES_ENC_BANQUES,
                       RUBRIQUES_ENC_TRVX_CCIALE, RUBRIQUES_ENC_TRVX_ADM,
                       TYPES_BRANCHEMENTS)
from calculs import (calcul_ca_agence, calcul_dr, calcul_national,
                      calcul_cumul_dr, calcul_cumul_national,
                      calcul_dashboard, get_agences_dr, classement_performances)
from export_excel import (export_consolidation, export_budget,
                          export_fiscal, export_reporting)
from monitoring import (generer_alertes, synthese_par_dr, get_params,
                        save_params, indicateurs_agence)
from import_historique import (importer_fichier, previsualiser_fichier,
                                generer_template_excel)

# ─── Logging ────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
)
logger = logging.getLogger('camwater')

app = Flask(__name__)

# ─── Configuration robustesse ────────────────────────────────────────
app.secret_key = os.environ.get(
    'CAMWATER_SECRET_KEY',
    'camwater-local-2026-secret-key'
)
# Taille maximale d'un upload : 32 Mo (protection contre les fichiers géants)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
# JSON compact + UTF-8 garanti
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False

# Dossier d'upload objectifs (seul cas où on écrit encore sur disque —
# les imports historiques, eux, sont 100 % éphémères BytesIO).
_DATA_DIR = os.environ.get(
    'CAMWATER_DATA_DIR',
    os.path.join(os.path.dirname(__file__), 'data')
)
UPLOAD_FOLDER = os.path.join(_DATA_DIR, 'objectifs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ─── Init DB au démarrage ────────────────────────────────────────────
init_db()

# ─── Gestionnaires d'erreurs globaux ────────────────────────────────
@app.errorhandler(400)
def bad_request(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Requête invalide', 'detail': str(e)}), 400
    return render_template('erreur.html', code=400,
                           message='Requête invalide'), 400

@app.errorhandler(403)
def forbidden(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Accès refusé'}), 403
    return render_template('erreur.html', code=403,
                           message='Accès non autorisé'), 403

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Ressource introuvable'}), 404
    return render_template('erreur.html', code=404,
                           message='Page introuvable'), 404

@app.errorhandler(RequestEntityTooLarge)
def too_large(e):
    return jsonify({
        'error': f'Fichier trop volumineux (max {app.config["MAX_CONTENT_LENGTH"] // 1024 // 1024} Mo)'
    }), 413

@app.errorhandler(500)
def server_error(e):
    logger.error('Erreur 500 : %s\n%s', e, traceback.format_exc())
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Erreur serveur interne — l\'équipe a été notifiée'}), 500
    return render_template('erreur.html', code=500,
                           message='Erreur interne du serveur'), 500

@app.errorhandler(Exception)
def unhandled(e):
    """Filet de sécurité absolu — aucune exception ne doit atteindre le client brut."""
    logger.error('Exception non gérée : %s\n%s', e, traceback.format_exc())
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Erreur inattendue — opération annulée'}), 500
    return render_template('erreur.html', code=500,
                           message='Une erreur inattendue est survenue'), 500


# ─── Contrôle d'accès par rôle ─────────────────────────────────────
# Rôles possibles dans la session :
#   'central'       : accès complet (saisie, consultations, dashboard, exports)
#   'agent_terrain' : accès limité à la saisie Recettes-Jour & Branchements-Jour
#   'direction'     : lecture seule (tableau de bord + consultations)
#
# Trois URL d'entrée distinctes à partager selon le public :
#   /central   → identification niveau central (saisie & consolidation)
#   /agences   → portail agents de terrain (sélection Recettes / Branchements)
#   /direction → accès direct des responsables au tableau de bord (lecture seule)

# URLs « commutation d'espace » : toujours accessibles (pour central/direction/anonyme)
# afin que la navigation entre espaces ne boucle jamais sur /dashboard.
# Les agents terrain restent quand eux verrouillés sur /terrain (aucune échappatoire).
SWITCH_PATHS = frozenset({'/', '/central', '/agences', '/direction'})


def _is_agent_terrain_allowed(path, method='GET'):
    """Retourne True si un agent terrain a le droit d'accéder à ce chemin."""
    if path.startswith('/static/') or path in ('/favicon.ico', '/robots.txt'):
        return True
    # NB : les SWITCH_PATHS (/, /central, /agences, /direction) sont *interdites*
    # aux agents terrain → toute tentative est captée par le guard et redirigée
    # vers /terrain.
    if path == '/terrain':
        return True
    if path in ('/recettes-jour', '/recettes-jour/saisie'):
        return True
    if path in ('/branchements-jour', '/branchements-jour/saisie'):
        return True
    if path in ('/api/recettes-jour/login', '/api/branchements-jour/login'):
        return True
    if path == '/api/recettes-jour' or path == '/api/branchements-jour':
        return True
    if path in ('/api/logout', '/api/operator', '/api/login-terrain', '/terrain/logout'):
        return True
    return False


def _is_direction_allowed(path, method='GET'):
    """Retourne True si un utilisateur 'direction' (lecture seule) a le droit d'accéder."""
    if method not in ('GET', 'HEAD', 'OPTIONS'):
        return False
    if path.startswith('/static/') or path in ('/favicon.ico', '/robots.txt'):
        return True
    # Pages en lecture seule (hors SWITCH_PATHS gérées séparément)
    if path == '/dashboard':
        return True
    if path == '/consultations' or path.startswith('/consultations/'):
        return True
    # Suivi des recettes quotidiennes (lecture seule pour la Direction)
    if path == '/recettes-jour/suivi':
        return True
    if path == '/api/recettes-jour/synthese':
        return True
    if path in ('/api/operator', '/api/logout'):
        return True
    if path == '/api/dashboard' or path.startswith('/api/consultation/'):
        return True
    if path == '/api/impayes/dashboard' or path == '/api/cumul/impayes':
        return True
    if path.startswith('/api/export/'):
        return True
    if path == '/api/objectifs' or path == '/api/statut':
        return True
    # Tableau de bord monitoring (lecture seule pour la Direction)
    if path == '/monitoring':
        return True
    if path == '/api/monitoring/alertes' or path == '/api/monitoring/indicateurs':
        return True
    if path == '/api/monitoring/params' and method == 'GET':
        return True
    return False


@app.before_request
def enforce_role_guard():
    """Applique le périmètre d'accès selon le rôle en session.

    Règles clefs :
    - Agents terrain : verrouillés. Toute URL hors périmètre → /terrain.
      Ils NE PEUVENT PAS revenir sur la page de choix des espaces.
    - Direction (lecture seule) : peut naviguer librement entre /, /central,
      /agences, /direction (URL de commutation). Pour le reste, GET de lecture
      uniquement. Les écritures et pages de saisie sont refusées.
    - Central / anonyme : pas de restriction ici.
    """
    path = request.path
    method = request.method
    role = session.get('role')

    # ── 1) Agents terrain : bouclés sur /terrain, y compris contre /, /central, etc.
    if role == 'agent_terrain':
        if _is_agent_terrain_allowed(path, method):
            return None
        if path.startswith('/api/'):
            return jsonify({"error": "Accès refusé",
                            "detail": "Cette ressource est réservée au niveau central."}), 403
        return redirect(url_for('terrain_portal'))

    # ── 2) URLs de commutation d'espace : toujours accessibles (central/direction/anonyme)
    #     C'est ce qui empêche la boucle infinie sur /dashboard.
    if path in SWITCH_PATHS and method == 'GET':
        return None

    # ── 3) Direction : lecture seule sur le reste du périmètre
    if role == 'direction':
        if _is_direction_allowed(path, method):
            return None
        if path.startswith('/api/'):
            return jsonify({"error": "Accès refusé",
                            "detail": "Lecture seule : opération non autorisée."}), 403
        return redirect(url_for('dashboard'))

    # ── 4) Central ou anonyme : pas de restriction
    return None


# ─── Helpers ────────────────────────────────────────────────────────

def get_all_agences():
    db = get_db()
    rows = db.execute("""
        SELECT a.id, a.nom, d.code as dr_code
        FROM agences a JOIN directions_regionales d ON a.dr_id = d.id
        ORDER BY d.code, a.nom
    """).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_all_drs():
    db = get_db()
    rows = db.execute("SELECT id, code, nom FROM directions_regionales ORDER BY code").fetchall()
    db.close()
    return [dict(r) for r in rows]


# ─── API Structure DRs / Agences (publique, sans auth) ──────────────

@app.route('/terrain-portal')
def terrain_portal_page():
    """Portail DR — page dédiée après identification depuis la vitrine."""
    return render_template('terrain_portal_page.html')


@app.route('/api/drs-agences')
def api_drs_agences():
    """Retourne la structure DRs → Agences pour le modal de connexion vitrine."""
    db = get_db()
    drs = db.execute(
        "SELECT id, code, nom FROM directions_regionales ORDER BY code"
    ).fetchall()
    result = []
    for dr in drs:
        agences = db.execute(
            "SELECT id, nom FROM agences WHERE dr_id=? ORDER BY nom", (dr['id'],)
        ).fetchall()
        result.append({
            'id': dr['id'], 'code': dr['code'], 'nom': dr['nom'],
            'agences': [{'id': a['id'], 'nom': a['nom']} for a in agences]
        })
    db.close()
    return jsonify(result)


# ─── Import Historique ──────────────────────────────────────────────

@app.route('/api/import-historique/template')
def api_import_template():
    """Télécharge le template Excel vierge (sans auth — pour diffusion)."""
    exercice = request.args.get('exercice', 2026, type=int)
    try:
        xls_bytes = generer_template_excel(exercice)
        return send_file(
            io.BytesIO(xls_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'CAMWATER_Template_Import_{exercice}.xlsx',
        )
    except Exception as exc:
        logger.error('Erreur génération template : %s', exc)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/import-historique/preview', methods=['POST'])
def api_import_preview():
    """Analyse le fichier uploadé et retourne un résumé SANS écrire en base.
    Si dr_id est fourni, restreint la résolution des agences à cette DR.
    Traitement 100 % éphémère : le fichier n'est jamais écrit sur disque."""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format requis : .xlsx ou .xls'}), 400
    exercice = request.form.get('exercice', 2026, type=int)
    dr_id = request.form.get('dr_id', type=int)

    try:
        file_bytes = f.stream.read()
        result = previsualiser_fichier(file_bytes, exercice, dr_id=dr_id)
        del file_bytes
        return jsonify(result)
    except Exception as exc:
        logger.error('Erreur preview import : %s\n%s', exc, traceback.format_exc())
        return jsonify({'error': f'Erreur analyse : {exc}'}), 500


@app.route('/api/import-historique', methods=['POST'])
def api_import_historique():
    """Import définitif des données historiques depuis un fichier Excel.
    Si dr_id est fourni : import restreint aux agences de cette DR (mode DR par DR).
    Le fichier est traité en mémoire pure et immédiatement détruit."""
    if session.get('role') != 'central':
        return jsonify({'error': 'Authentification requise (rôle central)'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format requis : .xlsx ou .xls'}), 400
    exercice = request.form.get('exercice', 2026, type=int)
    dr_id = request.form.get('dr_id', type=int)

    try:
        file_bytes = f.stream.read()
        result = importer_fichier(file_bytes, exercice, dr_id=dr_id)
        del file_bytes
        logger.info(
            'Import historique %s DR=%s : %s — %s lignes',
            exercice, dr_id or 'ALL',
            'OK' if result['success'] else 'ERR',
            result.get('total_lignes', 0),
        )
        return jsonify(result), 200 if result['success'] else 422
    except Exception as exc:
        logger.error('Erreur import historique : %s\n%s', exc, traceback.format_exc())
        return jsonify({'error': f'Erreur critique : {exc}', 'success': False}), 500


# ─── API Vitrine (publique, sans auth) ──────────────────────────────

@app.route('/api/vitrine-kpi')
def api_vitrine_kpi():
    """KPIs cumulatifs nationaux pour la page d'accueil vitrine — sans authentification."""
    import datetime
    exercice = request.args.get('exercice', 2026, type=int)
    mois_fin = min(datetime.date.today().month, 12)

    try:
        from calculs import calcul_cumul_national, get_branchements_national
        cumul = calcul_cumul_national(1, mois_fin, exercice)
        brch  = get_branchements_national(1, mois_fin, exercice)

        # Objectif encaissements national
        db = get_db()
        obj_row = db.execute("""
            SELECT COALESCE(SUM(montant), 0) as total FROM objectifs
            WHERE exercice=? AND scope_type='national' AND rubrique IN ('Encaissements','Encaissements facturation fraîche')
        """, (exercice,)).fetchone()
        objectif_enc = obj_row['total'] if obj_row else 0

        # Mois le plus récent avec données CA (pour indiquer la fraîcheur)
        last_row = db.execute("""
            SELECT MAX(mois) as m FROM volumes WHERE exercice=?
        """, (exercice,)).fetchone()
        dernier_mois = last_row['m'] if last_row and last_row['m'] else 0
        db.close()

        total_enc  = cumul.get('total_encaissements', 0) or 0
        objectif_p = objectif_enc * mois_fin / 12 if objectif_enc else 0
        taux_obj   = total_enc / objectif_p if objectif_p else 0

        return jsonify({
            'ca_global':           cumul.get('ca_global', 0) or 0,
            'total_encaissements': total_enc,
            'branchements_vendus': brch.get('vendus', 0) or 0,
            'taux_objectif':       taux_obj,
            'objectif_enc':        objectif_enc,
            'mois_fin':            mois_fin,
            'dernier_mois':        dernier_mois,
            'exercice':            exercice,
        })
    except Exception as e:
        return jsonify({'error': str(e), 'ca_global': 0, 'total_encaissements': 0,
                        'branchements_vendus': 0, 'taux_objectif': 0})


# ─── Pages ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Page d'accueil : présente les 3 entrées vers chaque espace.

    Les agents terrain sont gérés par le guard et ne peuvent pas atteindre cette
    page. Les utilisateurs 'central' et 'direction' voient les 3 cartes et
    peuvent changer d'espace à volonté.
    """
    return render_template('index.html')


@app.route('/central')
def central_login():
    """Entrée dédiée niveau central.

    Si l'utilisateur est déjà 'central', on garde sa session. Sinon on la
    nettoie pour éviter toute contamination depuis un autre rôle.
    """
    if session.get('role') != 'central':
        session.clear()
    return render_template('central_login.html')


@app.route('/agences')
def agences_portal():
    """Entrée dédiée agents de terrain : choix Recettes / Branchements.

    Toujours rendu avec une session vierge : l'agent terrain connecté est
    capturé par le guard en amont et renvoyé sur /terrain.
    """
    session.clear()
    return render_template('agences_portal.html')


@app.route('/direction')
def direction_portal():
    """Entrée dédiée responsables : tableau de bord en lecture seule, sans login."""
    if session.get('role') != 'direction':
        session.clear()
        session['role'] = 'direction'
    return redirect(url_for('dashboard'))


@app.route('/terrain')
def terrain_portal():
    """Ancienne route /terrain — redirigée vers /terrain-portal (nouvelle page dédiée)."""
    return redirect(url_for('terrain_portal_page'), code=301)


@app.route('/saisie/<fenetre>')
def saisie(fenetre):
    agences = get_all_agences()
    drs = get_all_drs()
    operateur = session.get('operateur', {})
    return render_template('saisie.html', fenetre=fenetre, agences=agences, drs=drs,
                           mois=MOIS, structure=STRUCTURE_CW,
                           structure_keys=list(STRUCTURE_CW.keys()),
                           rubriques_cciale=RUBRIQUES_ENC_CCIALE,
                           rubriques_adm=RUBRIQUES_ENC_ADM,
                           rubriques_banques=RUBRIQUES_ENC_BANQUES,
                           rubriques_trvx_cciale=RUBRIQUES_ENC_TRVX_CCIALE,
                           rubriques_trvx_adm=RUBRIQUES_ENC_TRVX_ADM,
                           types_branchements=TYPES_BRANCHEMENTS,
                           operateur=operateur)


@app.route('/consultations')
@app.route('/consultations/<fenetre>')
def consultations(fenetre='impayes_detail'):
    db = get_db()
    drs = db.execute(
        "SELECT id, code, nom FROM directions_regionales ORDER BY code"
    ).fetchall()
    db.close()
    return render_template('consultations.html', fenetre=fenetre,
                           mois=MOIS,
                           drs=drs,
                           structure_keys=list(STRUCTURE_CW.keys()),
                           structure=STRUCTURE_CW)


@app.route('/dashboard')
def dashboard():
    operateur = session.get('operateur', {})
    role = session.get('role') or 'anonyme'
    read_only = (role == 'direction')
    return render_template('dashboard.html', mois=MOIS,
                           drs=list(STRUCTURE_CW.keys()),
                           operateur=operateur,
                           role=role,
                           read_only=read_only)


@app.route('/objectifs')
@app.route('/objectifs/<fenetre>')
def objectifs_page(fenetre='objectifs'):
    # Redirection vers consultations
    return redirect(url_for('consultations', fenetre=fenetre))


# ─── API Authentification ──────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def api_login():
    import re as _re
    data = request.json
    nom = data.get('nom', '').strip()
    matricule = data.get('matricule', '').strip().upper()
    if not nom or not matricule:
        return jsonify({"error": "Nom et matricule requis"}), 400
    if not _re.match(r'^[0-9][A-Z][0-9][0-9]$', matricule):
        return jsonify({"error": "Format matricule invalide (ex: 1A23)"}), 400
    db = get_db()
    db.execute("""INSERT OR IGNORE INTO operateurs (nom, matricule) VALUES (?, ?)""",
               (nom, matricule))
    db.commit()
    db.close()
    # Login central : accès complet à l'application
    session['operateur'] = {"nom": nom, "matricule": matricule}
    session['role'] = 'central'
    return jsonify({"status": "ok", "nom": nom, "matricule": matricule, "role": "central"})


@app.route('/api/operator')
def api_operator():
    return jsonify({
        **session.get('operateur', {}),
        "role": session.get('role'),
    })


@app.route('/api/logout', methods=['POST', 'GET'])
def api_logout():
    """Déconnexion : purge complète de la session."""
    was_terrain = (session.get('role') == 'agent_terrain')
    session.clear()
    if request.method == 'GET':
        # Un agent terrain ne doit jamais voir la page des 3 espaces disponibles :
        # on le renvoie sur la page de connexion agents uniquement.
        if was_terrain:
            return redirect(url_for('agences_portal'))
        return redirect(url_for('index'))
    return jsonify({"status": "ok"})


@app.route('/terrain/logout', methods=['GET'])
def terrain_logout():
    """Déconnexion explicite pour un agent terrain → renvoie sur /agences."""
    session.clear()
    return redirect(url_for('agences_portal'))


# ─── API Statut Saisie (Brouillon / Définitif) ──────────────────

def check_locked(fenetre, agence_id, mois, exercice=2026):
    """Vérifie si une saisie est verrouillée (définitif)."""
    db = get_db()
    row = db.execute(
        "SELECT statut FROM saisie_statut WHERE fenetre=? AND agence_id=? AND mois=? AND exercice=?",
        (fenetre, int(agence_id), int(mois), exercice)).fetchone()
    db.close()
    return row is not None and row['statut'] == 'definitif'


@app.route('/api/statut', methods=['GET'])
def api_get_statut():
    fenetre = request.args.get('fenetre', '')
    agence_id = request.args.get('agence_id', 0, type=int)
    mois = request.args.get('mois', 0, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute(
        "SELECT statut, operateur_nom, operateur_matricule, date_statut FROM saisie_statut "
        "WHERE fenetre=? AND agence_id=? AND mois=? AND exercice=?",
        (fenetre, agence_id, mois, exercice)).fetchone()
    db.close()
    if row:
        return jsonify({"statut": row['statut'], "operateur_nom": row['operateur_nom'],
                        "operateur_matricule": row['operateur_matricule'],
                        "date_statut": row['date_statut']})
    return jsonify({"statut": None})


@app.route('/api/statut', methods=['POST'])
def api_save_statut():
    data = request.json
    fenetre = data.get('fenetre', '')
    agence_id = data.get('agence_id', 0)
    mois = data.get('mois', 0)
    exercice = data.get('exercice', 2026)
    statut = data.get('statut', 'brouillon')
    operateur_nom = data.get('operateur_nom', '')
    operateur_matricule = data.get('operateur_matricule', '')
    if check_locked(fenetre, agence_id, mois, exercice):
        return jsonify({"error": "Données déjà verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    db.execute("""INSERT INTO saisie_statut (fenetre, agence_id, mois, exercice, statut,
                      operateur_nom, operateur_matricule, date_statut)
                  VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
                  ON CONFLICT(fenetre, agence_id, mois, exercice)
                  DO UPDATE SET statut=excluded.statut, operateur_nom=excluded.operateur_nom,
                               operateur_matricule=excluded.operateur_matricule,
                               date_statut=excluded.date_statut""",
               (fenetre, int(agence_id), int(mois), exercice, statut,
                operateur_nom, operateur_matricule))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "statut": statut})


# ─── API Reset Agence (effacer toutes les données y compris définitif) ──

@app.route('/api/reset-agence', methods=['DELETE'])
def api_reset_agence():
    """Efface intégralement les données d'une agence pour un mois et un exercice donnés,
    y compris les enregistrements définitifs (saisie_statut)."""
    agence_id = request.args.get('agence_id', 0, type=int)
    mois = request.args.get('mois', 0, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    fenetre = request.args.get('fenetre', '')

    if not agence_id or not mois:
        return jsonify({"error": "agence_id et mois requis"}), 400

    db = get_db()
    deleted = 0

    # Tables à nettoyer selon la fenêtre (ou toutes si fenetre vide)
    if fenetre in ('volumes', ''):
        r = db.execute("DELETE FROM volumes WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('ca_ve', ''):
        r = db.execute("DELETE FROM ca_specifiques WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('ca_travaux', ''):
        r = db.execute("DELETE FROM complements_travaux WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('encaissements', ''):
        r = db.execute("DELETE FROM encaissements WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('branchements', ''):
        r = db.execute("DELETE FROM branchements WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('recettes', ''):
        r = db.execute("DELETE FROM recettes WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount
    if fenetre in ('impayes', ''):
        r = db.execute("DELETE FROM impayes WHERE agence_id=? AND mois=? AND exercice=?",
                       (agence_id, mois, exercice))
        deleted += r.rowcount

    # Supprimer aussi le statut de saisie (brouillon/définitif) pour cette fenêtre
    if fenetre:
        db.execute("DELETE FROM saisie_statut WHERE fenetre=? AND agence_id=? AND mois=? AND exercice=?",
                   (fenetre, agence_id, mois, exercice))
    else:
        # Toutes les fenêtres
        db.execute("DELETE FROM saisie_statut WHERE agence_id=? AND mois=? AND exercice=?",
                   (agence_id, mois, exercice))

    db.commit()
    db.close()
    return jsonify({"status": "ok", "deleted": deleted,
                    "message": f"Données réinitialisées ({deleted} enregistrements supprimés)"})


# ─── API Saisie : Volumes ──────────────────────────────────────────

@app.route('/api/volumes', methods=['GET'])
def api_get_volumes():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute("SELECT categorie, valeur FROM volumes WHERE agence_id=? AND mois=? AND exercice=?",
                      (agence_id, mois, exercice)).fetchall()
    db.close()
    return jsonify({r['categorie']: r['valeur'] for r in rows})


@app.route('/api/volumes', methods=['POST'])
def api_save_volumes():
    data = request.json
    if check_locked('volumes', data['agence_id'], data['mois'], data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    for cat, val in data['valeurs'].items():
        db.execute("""INSERT INTO volumes (agence_id, mois, exercice, categorie, valeur)
                      VALUES (?, ?, ?, ?, ?)
                      ON CONFLICT(agence_id, mois, exercice, categorie)
                      DO UPDATE SET valeur=excluded.valeur""",
                   (data['agence_id'], data['mois'], data.get('exercice', 2026), cat, val or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Saisie : CA Vente Eau (ex CA Spécifiques) ────────────────

@app.route('/api/ca_specifiques', methods=['GET'])
def api_get_ca_spec():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute("SELECT rubrique, montant FROM ca_specifiques WHERE agence_id=? AND mois=? AND exercice=?",
                      (agence_id, mois, exercice)).fetchall()
    db.close()
    return jsonify({r['rubrique']: r['montant'] for r in rows})


@app.route('/api/ca_specifiques', methods=['POST'])
def api_save_ca_spec():
    data = request.json
    if check_locked('ca_ve', data['agence_id'], data['mois'], data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    for rub, val in data['valeurs'].items():
        db.execute("""INSERT INTO ca_specifiques (agence_id, mois, exercice, rubrique, montant)
                      VALUES (?, ?, ?, ?, ?)
                      ON CONFLICT(agence_id, mois, exercice, rubrique)
                      DO UPDATE SET montant=excluded.montant""",
                   (data['agence_id'], data['mois'], data.get('exercice', 2026), rub, val or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Saisie : Encaissements ───────────────────────────────────

@app.route('/api/encaissements', methods=['GET'])
def api_get_enc():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute("SELECT section, rubrique, montant FROM encaissements WHERE agence_id=? AND mois=? AND exercice=?",
                      (agence_id, mois, exercice)).fetchall()
    db.close()
    return jsonify({f"{r['section']}|{r['rubrique']}": r['montant'] for r in rows})


@app.route('/api/encaissements', methods=['POST'])
def api_save_enc():
    data = request.json
    if check_locked('encaissements', data['agence_id'], data['mois'], data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    for key, val in data['valeurs'].items():
        section, rubrique = key.split('|', 1)
        db.execute("""INSERT INTO encaissements (agence_id, mois, exercice, section, rubrique, montant)
                      VALUES (?, ?, ?, ?, ?, ?)
                      ON CONFLICT(agence_id, mois, exercice, section, rubrique)
                      DO UPDATE SET montant=excluded.montant""",
                   (data['agence_id'], data['mois'], data.get('exercice', 2026), section, rubrique, val or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Saisie : CA Travaux (Compléments) ────────────────────────

@app.route('/api/complements', methods=['GET'])
def api_get_complements():
    agence_id = request.args.get('agence_id', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    # Compléments
    comp_rows = db.execute("SELECT mois, montant FROM complements_travaux WHERE agence_id=? AND exercice=?",
                           (agence_id, exercice)).fetchall()
    comps = {r['mois']: r['montant'] for r in comp_rows}
    # Encaissements (frais pose, vérif, mutation)
    enc_rows = db.execute("""
        SELECT mois,
            COALESCE(SUM(CASE WHEN rubrique='Frais pose cptrs' THEN montant ELSE 0 END), 0) as frais_pose,
            COALESCE(SUM(CASE WHEN rubrique='Frais vérif/étalon.' THEN montant ELSE 0 END), 0) as frais_verif,
            COALESCE(SUM(CASE WHEN rubrique='Mutation' THEN montant ELSE 0 END), 0) as mutation
        FROM encaissements
        WHERE agence_id=? AND exercice=? AND section='trvx_cciale'
            AND rubrique IN ('Frais pose cptrs','Frais vérif/étalon.','Mutation')
        GROUP BY mois
    """, (agence_id, exercice)).fetchall()
    db.close()

    result = {}
    for m in range(1, 13):
        enc_data = next((dict(r) for r in enc_rows if r['mois'] == m), {})
        result[m] = {
            "complement": comps.get(m, 0),
            "frais_pose": enc_data.get('frais_pose', 0),
            "frais_verif": enc_data.get('frais_verif', 0),
            "mutation": enc_data.get('mutation', 0),
        }
    return jsonify(result)


@app.route('/api/complements', methods=['POST'])
def api_save_complements():
    data = request.json
    if check_locked('ca_travaux', data['agence_id'], 0, data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    for mois_str, montant in data['valeurs'].items():
        db.execute("""INSERT INTO complements_travaux (agence_id, mois, exercice, montant)
                      VALUES (?, ?, ?, ?)
                      ON CONFLICT(agence_id, mois, exercice)
                      DO UPDATE SET montant=excluded.montant""",
                   (data['agence_id'], int(mois_str), data.get('exercice', 2026), montant or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Saisie : Branchements (par agence) ──────────────────────

@app.route('/api/branchements', methods=['GET'])
def api_get_branchements():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute("SELECT type, valeur FROM branchements WHERE agence_id=? AND mois=? AND exercice=?",
                      (agence_id, mois, exercice)).fetchall()
    db.close()
    return jsonify({r['type']: r['valeur'] for r in rows})


@app.route('/api/branchements', methods=['POST'])
def api_save_branchements():
    data = request.json
    if check_locked('branchements', data['agence_id'], data['mois'], data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    for btype, val in data['valeurs'].items():
        db.execute("""INSERT INTO branchements (agence_id, mois, exercice, type, valeur)
                      VALUES (?, ?, ?, ?, ?)
                      ON CONFLICT(agence_id, mois, exercice, type)
                      DO UPDATE SET valeur=excluded.valeur""",
                   (data['agence_id'], data['mois'], data.get('exercice', 2026), btype, val or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Saisie : Recettes (par agence) ──────────────────────────

@app.route('/api/recettes', methods=['GET'])
def api_get_recettes():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute("SELECT montant FROM recettes WHERE agence_id=? AND mois=? AND exercice=?",
                     (agence_id, mois, exercice)).fetchone()
    db.close()
    return jsonify({"montant": row['montant'] if row else 0})


@app.route('/api/recettes', methods=['POST'])
def api_save_recettes():
    data = request.json
    if check_locked('recettes', data['agence_id'], data['mois'], data.get('exercice', 2026)):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    db = get_db()
    db.execute("""INSERT INTO recettes (agence_id, mois, exercice, montant)
                  VALUES (?, ?, ?, ?)
                  ON CONFLICT(agence_id, mois, exercice)
                  DO UPDATE SET montant=excluded.montant""",
               (data['agence_id'], data['mois'], data.get('exercice', 2026), data.get('montant', 0)))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


# ─── API Cumul par agence (bulles récapitulatives saisie) ────────

@app.route('/api/cumul/volumes')
def api_cumul_volumes():
    """Cumul annuel volumes + CA auto pour une agence."""
    agence_id = request.args.get('agence_id', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute(
        "SELECT categorie, SUM(valeur) as vol FROM volumes WHERE agence_id=? AND exercice=? GROUP BY categorie",
        (agence_id, exercice)).fetchall()
    db.close()
    from database import CATEGORIES_PU
    total_vol = 0
    total_ca = 0
    for r in rows:
        v = r['vol'] or 0
        total_vol += v
        pu = CATEGORIES_PU.get(r['categorie'], 0)
        total_ca += v * pu
    return jsonify({"total_volumes": total_vol, "total_ca": total_ca})


@app.route('/api/cumul/ca')
def api_cumul_ca():
    """Cumul annuel CA vente eau + travaux + global pour une agence."""
    agence_id = request.args.get('agence_id', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    result = {"ca_ve": 0, "ca_trvx": 0, "ca_global": 0}
    try:
        data = calcul_ca_agence(agence_id, exercice=exercice)
        result["ca_ve"] = data.get("total_ve", 0) or 0
        result["ca_trvx"] = data.get("total_trvx_remb", 0) or 0
        result["ca_global"] = data.get("ca_global", 0) or 0
    except Exception:
        pass
    return jsonify(result)


@app.route('/api/cumul/encaissements')
def api_cumul_encaissements():
    """Cumul annuel encaissements pour une agence."""
    agence_id = request.args.get('agence_id', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute("""
        SELECT COALESCE(SUM(montant), 0) as enc_total
        FROM encaissements WHERE agence_id=? AND exercice=?
    """, (agence_id, exercice)).fetchone()
    db.close()
    return jsonify({"total": row['enc_total'] if row else 0})


@app.route('/api/cumul/impayes')
def api_cumul_impayes():
    """Dernier solde impayés enregistré pour une agence (mois le plus récent)."""
    agence_id = request.args.get('agence_id', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute("""
        SELECT particuliers_actifs, gco_actifs, particuliers_resilies, gco_resilies,
               bf_actifs, bfc_actifs, bf_resilies, bfc_resilies, gestion_manuelle
        FROM impayes WHERE agence_id=? AND exercice=?
        ORDER BY mois DESC LIMIT 1
    """, (agence_id, exercice)).fetchone()
    db.close()
    if not row:
        return jsonify({"actifs": 0, "resilies": 0, "total": 0})
    actifs = (row['particuliers_actifs'] or 0) + (row['gco_actifs'] or 0) + (row['bf_actifs'] or 0) + (row['bfc_actifs'] or 0)
    resilies = (row['particuliers_resilies'] or 0) + (row['gco_resilies'] or 0) + (row['bf_resilies'] or 0) + (row['bfc_resilies'] or 0) + (row['gestion_manuelle'] or 0)
    return jsonify({"actifs": actifs, "resilies": resilies, "total": actifs + resilies})


# ─── API Saisie : Impayés ─────────────────────────────────────────

@app.route('/api/impayes', methods=['GET'])
def api_get_impayes():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute(
        "SELECT * FROM impayes WHERE agence_id=? AND mois=? AND exercice=?",
        (agence_id, mois, exercice)).fetchone()
    db.close()
    if row:
        return jsonify({
            "particuliers_actifs": row['particuliers_actifs'],
            "gco_actifs": row['gco_actifs'],
            "particuliers_resilies": row['particuliers_resilies'],
            "gco_resilies": row['gco_resilies'],
            "bf_actifs": row['bf_actifs'],
            "bfc_actifs": row['bfc_actifs'],
            "bf_resilies": row['bf_resilies'],
            "bfc_resilies": row['bfc_resilies'],
            "gestion_manuelle": row['gestion_manuelle'],
            "resiliers_crediteurs": row['resiliers_crediteurs'],
        })
    return jsonify({})


@app.route('/api/impayes', methods=['POST'])
def api_save_impayes():
    data = request.json
    agence_id = data.get('agence_id')
    mois = data.get('mois')
    exercice = data.get('exercice', 2026)
    if check_locked('impayes', agence_id, mois, exercice):
        return jsonify({"error": "Données verrouillées (enregistrement définitif)"}), 403
    vals = data.get('valeurs', {})
    db = get_db()
    db.execute("""INSERT INTO impayes (agence_id, mois, exercice,
                      particuliers_actifs, gco_actifs, particuliers_resilies, gco_resilies,
                      bf_actifs, bfc_actifs, bf_resilies, bfc_resilies,
                      gestion_manuelle, resiliers_crediteurs)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                  ON CONFLICT(agence_id, mois, exercice)
                  DO UPDATE SET
                      particuliers_actifs=excluded.particuliers_actifs,
                      gco_actifs=excluded.gco_actifs,
                      particuliers_resilies=excluded.particuliers_resilies,
                      gco_resilies=excluded.gco_resilies,
                      bf_actifs=excluded.bf_actifs,
                      bfc_actifs=excluded.bfc_actifs,
                      bf_resilies=excluded.bf_resilies,
                      bfc_resilies=excluded.bfc_resilies,
                      gestion_manuelle=excluded.gestion_manuelle,
                      resiliers_crediteurs=excluded.resiliers_crediteurs""",
               (int(agence_id), int(mois), exercice,
                vals.get('particuliers_actifs', 0) or 0,
                vals.get('gco_actifs', 0) or 0,
                vals.get('particuliers_resilies', 0) or 0,
                vals.get('gco_resilies', 0) or 0,
                vals.get('bf_actifs', 0) or 0,
                vals.get('bfc_actifs', 0) or 0,
                vals.get('bf_resilies', 0) or 0,
                vals.get('bfc_resilies', 0) or 0,
                vals.get('gestion_manuelle', 0) or 0,
                vals.get('resiliers_crediteurs', 0) or 0))
    db.commit()
    db.close()
    return jsonify({"status": "ok"})


@app.route('/api/impayes/dashboard')
def api_impayes_dashboard():
    """Données impayés agrégées pour le tableau de bord."""
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', '')

    db = get_db()

    # Filtrage par site
    if site_type == 'dr' and site_id:
        agences_rows = db.execute(
            "SELECT a.id FROM agences a JOIN directions_regionales d ON a.dr_id=d.id WHERE d.code=?",
            (site_id,)).fetchall()
    else:
        agences_rows = db.execute("SELECT id FROM agences").fetchall()
    agence_ids = [r['id'] for r in agences_rows]

    if not agence_ids:
        db.close()
        return jsonify({"totaux": {}, "par_statut": {}, "par_categorie": {},
                        "par_dr": {}, "evolution_mensuelle": []})

    ph = ','.join('?' * len(agence_ids))

    # Dernier mois de la plage (état du portefeuille au mois fin)
    row = db.execute(f"""
        SELECT
            COALESCE(SUM(particuliers_actifs), 0) as particuliers_actifs,
            COALESCE(SUM(gco_actifs), 0) as gco_actifs,
            COALESCE(SUM(particuliers_resilies), 0) as particuliers_resilies,
            COALESCE(SUM(gco_resilies), 0) as gco_resilies,
            COALESCE(SUM(bf_actifs), 0) as bf_actifs,
            COALESCE(SUM(bfc_actifs), 0) as bfc_actifs,
            COALESCE(SUM(bf_resilies), 0) as bf_resilies,
            COALESCE(SUM(bfc_resilies), 0) as bfc_resilies,
            COALESCE(SUM(gestion_manuelle), 0) as gestion_manuelle,
            COALESCE(SUM(resiliers_crediteurs), 0) as resiliers_crediteurs
        FROM impayes
        WHERE agence_id IN ({ph}) AND exercice=? AND mois=?
    """, agence_ids + [exercice, mois_fin]).fetchone()

    totaux = dict(row) if row else {}

    # Calculs dérivés
    total_actifs = totaux.get('particuliers_actifs', 0) + totaux.get('gco_actifs', 0)
    total_resilies = totaux.get('particuliers_resilies', 0) + totaux.get('gco_resilies', 0)
    total_com_actifs = totaux.get('bf_actifs', 0) + totaux.get('bfc_actifs', 0)
    total_com_resilies = (totaux.get('bf_resilies', 0) + totaux.get('bfc_resilies', 0)
                          + totaux.get('gestion_manuelle', 0))
    total_impayes = total_actifs + total_resilies + total_com_actifs + total_com_resilies

    par_statut = {
        "Actifs": total_actifs + total_com_actifs,
        "Résiliés (Inactifs)": total_resilies + total_com_resilies,
    }

    par_categorie = {
        "Particuliers": totaux.get('particuliers_actifs', 0) + totaux.get('particuliers_resilies', 0),
        "GCO": totaux.get('gco_actifs', 0) + totaux.get('gco_resilies', 0),
        "Bornes Fontaines": totaux.get('bf_actifs', 0) + totaux.get('bf_resilies', 0),
        "B.F.C": totaux.get('bfc_actifs', 0) + totaux.get('bfc_resilies', 0),
    }

    # Par DR (au mois_fin)
    dr_rows = db.execute(f"""
        SELECT d.code as dr_code,
            COALESCE(SUM(i.particuliers_actifs), 0) + COALESCE(SUM(i.gco_actifs), 0)
            + COALESCE(SUM(i.bf_actifs), 0) + COALESCE(SUM(i.bfc_actifs), 0) as actifs,
            COALESCE(SUM(i.particuliers_resilies), 0) + COALESCE(SUM(i.gco_resilies), 0)
            + COALESCE(SUM(i.bf_resilies), 0) + COALESCE(SUM(i.bfc_resilies), 0)
            + COALESCE(SUM(i.gestion_manuelle), 0) as resilies,
            COALESCE(SUM(i.particuliers_actifs), 0) + COALESCE(SUM(i.gco_actifs), 0)
            + COALESCE(SUM(i.bf_actifs), 0) + COALESCE(SUM(i.bfc_actifs), 0)
            + COALESCE(SUM(i.particuliers_resilies), 0) + COALESCE(SUM(i.gco_resilies), 0)
            + COALESCE(SUM(i.bf_resilies), 0) + COALESCE(SUM(i.bfc_resilies), 0)
            + COALESCE(SUM(i.gestion_manuelle), 0) as total
        FROM impayes i
        JOIN agences a ON i.agence_id = a.id
        JOIN directions_regionales d ON a.dr_id = d.id
        WHERE i.agence_id IN ({ph}) AND i.exercice=? AND i.mois=?
        GROUP BY d.code ORDER BY d.code
    """, agence_ids + [exercice, mois_fin]).fetchall()

    par_dr = {r['dr_code']: {"actifs": r['actifs'], "resilies": r['resilies'],
                              "total": r['total']} for r in dr_rows}

    # Évolution mensuelle par 8 catégories clients (valeurs brutes).
    # On remonte d'un mois avant mois_debut pour pouvoir calculer la variation
    # % vs M-1 dès le premier mois de la plage.
    mois_start_ref = max(1, mois_debut - 1)
    evolution = []
    for m in range(mois_start_ref, mois_fin + 1):
        mrow = db.execute(f"""
            SELECT
                COALESCE(SUM(particuliers_actifs), 0) as particuliers_actifs,
                COALESCE(SUM(gco_actifs), 0) as gco_actifs,
                COALESCE(SUM(particuliers_resilies), 0) as particuliers_resilies,
                COALESCE(SUM(gco_resilies), 0) as gco_resilies,
                COALESCE(SUM(bf_actifs), 0) as bf_actifs,
                COALESCE(SUM(bfc_actifs), 0) as bfc_actifs,
                COALESCE(SUM(bf_resilies), 0) as bf_resilies,
                COALESCE(SUM(bfc_resilies), 0) as bfc_resilies,
                COALESCE(SUM(gestion_manuelle), 0) as gestion_manuelle
            FROM impayes
            WHERE agence_id IN ({ph}) AND exercice=? AND mois=?
        """, agence_ids + [exercice, m]).fetchone()
        d = dict(mrow) if mrow else {}
        total = (d.get('particuliers_actifs', 0) + d.get('gco_actifs', 0)
                 + d.get('particuliers_resilies', 0) + d.get('gco_resilies', 0)
                 + d.get('bf_actifs', 0) + d.get('bfc_actifs', 0)
                 + d.get('bf_resilies', 0) + d.get('bfc_resilies', 0)
                 + d.get('gestion_manuelle', 0))
        evolution.append({
            "mois": m,
            "in_range": m >= mois_debut,  # faux pour le mois de référence ajouté
            "total": total,
            "particuliers_actifs": d.get('particuliers_actifs', 0),
            "gco_actifs": d.get('gco_actifs', 0),
            "particuliers_resilies": d.get('particuliers_resilies', 0),
            "gco_resilies": d.get('gco_resilies', 0),
            "bf_actifs": d.get('bf_actifs', 0),
            "bfc_actifs": d.get('bfc_actifs', 0),
            "bf_resilies": d.get('bf_resilies', 0),
            "bfc_resilies": d.get('bfc_resilies', 0),
        })

    db.close()

    return jsonify({
        "totaux": {
            "total_actifs": total_actifs,
            "total_resilies": total_resilies,
            "total_com_actifs": total_com_actifs,
            "total_com_resilies": total_com_resilies,
            "total_impayes": total_impayes,
            "resiliers_crediteurs": totaux.get('resiliers_crediteurs', 0),
            "gestion_manuelle": totaux.get('gestion_manuelle', 0),
        },
        "par_statut": par_statut,
        "par_categorie": par_categorie,
        "par_dr": par_dr,
        "evolution_mensuelle": evolution,
    })


@app.route('/api/export/impayes')
def api_export_impayes():
    """Export des impayés au format Excel (même format que le fichier source)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', '')

    db = get_db()
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    total_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    col_headers = [
        "Agence", "Part. Actifs", "GCO Actifs", "Total Actifs",
        "Part. Résiliés", "GCO Résiliés", "Total Rés. Débiteurs",
        "B.F. Actifs", "B.F.C Actifs", "Total Com. Actifs",
        "B.F. Résiliés", "B.F.C Résiliés", "Gest. Manuelle", "Total Com. Rés.",
        "TOTAL IMPAYÉS", "Rés. Créditeurs"
    ]

    def write_sheet_for_dr(ws, dr_code, dr_agences, exercice, mois_debut, mois_fin):
        row_num = 1
        for m in range(mois_debut, mois_fin + 1):
            mois_label = MOIS[m - 1].upper() + f" {exercice}"
            ws.cell(row=row_num, column=1, value=mois_label).font = Font(bold=True, size=11)
            row_num += 1
            for ci, h in enumerate(col_headers, 1):
                cell = ws.cell(row=row_num, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row_num += 1

            totals = [0] * 15
            for ag_id, ag_nom in dr_agences:
                imp = db.execute(
                    "SELECT * FROM impayes WHERE agence_id=? AND mois=? AND exercice=?",
                    (ag_id, m, exercice)).fetchone()
                vals = [
                    imp['particuliers_actifs'] if imp else 0,
                    imp['gco_actifs'] if imp else 0,
                    0,  # total actifs (formula)
                    imp['particuliers_resilies'] if imp else 0,
                    imp['gco_resilies'] if imp else 0,
                    0,  # total resilies
                    imp['bf_actifs'] if imp else 0,
                    imp['bfc_actifs'] if imp else 0,
                    0,  # total com actifs
                    imp['bf_resilies'] if imp else 0,
                    imp['bfc_resilies'] if imp else 0,
                    imp['gestion_manuelle'] if imp else 0,
                    0,  # total com resilies
                    0,  # total impayes
                    imp['resiliers_crediteurs'] if imp else 0,
                ]
                # Computed
                vals[2] = vals[0] + vals[1]
                vals[5] = vals[3] + vals[4]
                vals[8] = vals[6] + vals[7]
                vals[12] = vals[9] + vals[10] + vals[11]
                vals[13] = vals[2] + vals[5] + vals[8] + vals[12]

                ws.cell(row=row_num, column=1, value=ag_nom).border = thin_border
                for ci, v in enumerate(vals, 2):
                    cell = ws.cell(row=row_num, column=ci, value=v)
                    cell.number_format = '#,##0'
                    cell.border = thin_border
                    totals[ci - 2] += v
                row_num += 1

            # Total DR row
            ws.cell(row=row_num, column=1, value=f"ENS {dr_code}").font = Font(bold=True)
            ws.cell(row=row_num, column=1).border = thin_border
            # Recompute totals for computed columns
            totals[2] = totals[0] + totals[1]
            totals[5] = totals[3] + totals[4]
            totals[8] = totals[6] + totals[7]
            totals[12] = totals[9] + totals[10] + totals[11]
            totals[13] = totals[2] + totals[5] + totals[8] + totals[12]
            for ci, v in enumerate(totals, 2):
                cell = ws.cell(row=row_num, column=ci, value=v)
                cell.font = Font(bold=True)
                cell.number_format = '#,##0'
                cell.fill = total_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.border = thin_border
            row_num += 2

    # Generate sheets per DR
    first_sheet = True
    for dr_code, agences_list in STRUCTURE_CW.items():
        if site_type == 'dr' and site_id and dr_code != site_id:
            continue
        if first_sheet:
            ws = wb.active
            ws.title = dr_code
            first_sheet = False
        else:
            ws = wb.create_sheet(dr_code)

        dr_agences = db.execute("""
            SELECT a.id, a.nom FROM agences a
            JOIN directions_regionales d ON a.dr_id = d.id
            WHERE d.code=? ORDER BY a.nom
        """, (dr_code,)).fetchall()
        dr_ag_list = [(r['id'], r['nom']) for r in dr_agences]
        write_sheet_for_dr(ws, dr_code, dr_ag_list, exercice, mois_debut, mois_fin)

    # RECAP MENSUEL sheet (national summary)
    if site_type == 'national':
        ws_recap = wb.create_sheet("RECAP MENSUEL", 0)
        row_num = 1
        for m in range(mois_debut, mois_fin + 1):
            mois_label = MOIS[m - 1].upper() + f" {exercice}"
            ws_recap.cell(row=row_num, column=1, value=mois_label).font = Font(bold=True, size=11)
            row_num += 1
            recap_headers = ["DR"] + col_headers[1:]
            for ci, h in enumerate(recap_headers, 1):
                cell = ws_recap.cell(row=row_num, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row_num += 1

            grand_totals = [0] * 15
            for dr_code in STRUCTURE_CW:
                dr_row = db.execute("""
                    SELECT
                        COALESCE(SUM(particuliers_actifs),0) as pa,
                        COALESCE(SUM(gco_actifs),0) as ga,
                        COALESCE(SUM(particuliers_resilies),0) as pr,
                        COALESCE(SUM(gco_resilies),0) as gr,
                        COALESCE(SUM(bf_actifs),0) as bfa,
                        COALESCE(SUM(bfc_actifs),0) as bfca,
                        COALESCE(SUM(bf_resilies),0) as bfr,
                        COALESCE(SUM(bfc_resilies),0) as bfcr,
                        COALESCE(SUM(gestion_manuelle),0) as gm,
                        COALESCE(SUM(resiliers_crediteurs),0) as rc
                    FROM impayes i
                    JOIN agences a ON i.agence_id = a.id
                    JOIN directions_regionales d ON a.dr_id = d.id
                    WHERE d.code=? AND i.exercice=? AND i.mois=?
                """, (dr_code, exercice, m)).fetchone()

                vals = [
                    dr_row['pa'], dr_row['ga'], 0,
                    dr_row['pr'], dr_row['gr'], 0,
                    dr_row['bfa'], dr_row['bfca'], 0,
                    dr_row['bfr'], dr_row['bfcr'], dr_row['gm'], 0,
                    0, dr_row['rc']
                ]
                vals[2] = vals[0] + vals[1]
                vals[5] = vals[3] + vals[4]
                vals[8] = vals[6] + vals[7]
                vals[12] = vals[9] + vals[10] + vals[11]
                vals[13] = vals[2] + vals[5] + vals[8] + vals[12]

                ws_recap.cell(row=row_num, column=1, value=dr_code).font = Font(bold=True)
                ws_recap.cell(row=row_num, column=1).border = thin_border
                for ci, v in enumerate(vals, 2):
                    cell = ws_recap.cell(row=row_num, column=ci, value=v)
                    cell.number_format = '#,##0'
                    cell.border = thin_border
                    grand_totals[ci - 2] += v
                row_num += 1

            # Grand total row
            ws_recap.cell(row=row_num, column=1, value="ENSEMBLE CAMWATER").font = Font(bold=True)
            ws_recap.cell(row=row_num, column=1).border = thin_border
            grand_totals[2] = grand_totals[0] + grand_totals[1]
            grand_totals[5] = grand_totals[3] + grand_totals[4]
            grand_totals[8] = grand_totals[6] + grand_totals[7]
            grand_totals[12] = grand_totals[9] + grand_totals[10] + grand_totals[11]
            grand_totals[13] = grand_totals[2] + grand_totals[5] + grand_totals[8] + grand_totals[12]
            for ci, v in enumerate(grand_totals, 2):
                cell = ws_recap.cell(row=row_num, column=ci, value=v)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.number_format = '#,##0'
                cell.fill = total_fill
                cell.border = thin_border
            row_num += 2

    db.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    label = site_id if site_type == 'dr' else 'CAMWATER'
    filename = f"SSP_IMPAYES_{label}_{MOIS[mois_debut-1]}-{MOIS[mois_fin-1]}_{exercice}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ─── API CA Total (détail agence) ─────────────────────────────────

@app.route('/api/ca_global/<int:agence_id>/<int:mois>')
def api_ca_detail(agence_id, mois):
    exercice = request.args.get('exercice', 2026, type=int)
    data = calcul_ca_agence(agence_id, mois, exercice)
    return jsonify(data)


# ─── API Dashboard ─────────────────────────────────────────────────

@app.route('/api/dashboard')
def api_dashboard():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', None)
    data = calcul_dashboard(mois_debut, mois_fin, exercice, site_type, site_id)
    return jsonify(data)


# ─── API Classement performances ──────────────────────────────────

@app.route('/api/classement')
def api_classement():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    rubrique = request.args.get('rubrique', 'Encaissements')
    data = classement_performances(mois_debut, mois_fin, exercice, rubrique)
    return jsonify(data)


# ─── API Exports Excel ────────────────────────────────────────────

@app.route('/api/export/consolidation')
def api_export_consolidation():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', None)
    data = export_consolidation(mois_debut, mois_fin, exercice, site_type, site_id)
    filename = f"Consolidation_{site_type}_{site_id or 'CAMWATER'}_{MOIS[mois_debut-1]}-{MOIS[mois_fin-1]}_{exercice}.xlsx"
    return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/export/budget')
def api_export_budget():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', None)
    data = export_budget(mois_debut, mois_fin, exercice, site_type, site_id)
    return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Budget_{exercice}.xlsx")


@app.route('/api/export/fiscal')
def api_export_fiscal():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', None)
    data = export_fiscal(mois_debut, mois_fin, exercice, site_type, site_id)
    return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Fiscal_{exercice}.xlsx")


@app.route('/api/export/reporting')
def api_export_reporting():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 2, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', None)
    data = export_reporting(mois_debut, mois_fin, exercice, site_type, site_id)
    return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Reporting_{exercice}.xlsx")


# ─── API Objectifs : Upload ───────────────────────────────────────

@app.route('/api/objectifs/upload', methods=['POST'])
def api_upload_objectifs():
    """Upload un fichier Excel d'objectifs.
    Formats acceptés :
      4 colonnes: scope_type | scope_ref | rubrique | valeur
      5 colonnes: scope_type | scope_ref | rubrique | mois | valeur
    scope_type peut être: 'camwater'/'national' ou un code DR (DRYA, DRA, etc.) ou 'agence'
    """
    if 'file' not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Format Excel requis (.xlsx)"}), 400

    # ── Lecture en mémoire — le fichier n'est JAMAIS écrit sur disque ──────────
    import openpyxl
    filename = secure_filename(f.filename)
    file_bytes = f.stream.read()

    # Codes DR connus (en lowercase pour comparaison)
    dr_codes_lower = {c.lower(): c for c in STRUCTURE_CW.keys()}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        del file_bytes  # destruction immédiate
        ws = wb.active
        db = get_db()
        count = 0
        exercice = request.form.get('exercice', 2026, type=int)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            col0 = str(row[0]).strip() if row[0] else ''
            col1 = str(row[1]).strip() if row[1] else ''
            rubrique = str(row[2]).strip() if row[2] else None

            if not rubrique:
                continue

            # Détecter le nombre de colonnes utiles
            ncols = len([c for c in row if c is not None and str(c).strip()])

            # Déterminer la valeur : dernière colonne numérique
            valeur = 0
            mois_val = None
            if ncols >= 5 and row[4] is not None:
                # 5 colonnes : scope | ref | rubrique | mois | valeur
                try:
                    mois_val = int(row[3]) if row[3] and 1 <= int(row[3]) <= 12 else None
                except (ValueError, TypeError):
                    mois_val = None
                try:
                    valeur = float(row[4])
                except (ValueError, TypeError):
                    valeur = 0
            elif ncols >= 4 and row[3] is not None:
                # 4 colonnes : scope | ref | rubrique | valeur
                try:
                    valeur = float(row[3])
                except (ValueError, TypeError):
                    valeur = 0
                mois_val = None

            # Résoudre scope_type et scope_id
            scope_type_raw = col0.lower().strip()
            scope_id = None

            if scope_type_raw in ('camwater', 'national', 'nat', ''):
                scope_type = 'national'
            elif scope_type_raw in dr_codes_lower:
                # Le scope_type est directement un code DR (drya, dra, etc.)
                scope_type = 'dr'
                dr_code_real = dr_codes_lower[scope_type_raw]
                r = db.execute("SELECT id FROM directions_regionales WHERE code=?", (dr_code_real,)).fetchone()
                scope_id = r['id'] if r else None
            elif scope_type_raw == 'dr' and col1:
                scope_type = 'dr'
                # col1 = code DR
                r = db.execute("SELECT id FROM directions_regionales WHERE code=?", (col1.upper(),)).fetchone()
                if not r:
                    r = db.execute("SELECT id FROM directions_regionales WHERE LOWER(code)=?", (col1.lower(),)).fetchone()
                scope_id = r['id'] if r else None
            elif scope_type_raw == 'agence' and col1:
                scope_type = 'agence'
                r = db.execute("SELECT id FROM agences WHERE nom=?", (col1,)).fetchone()
                scope_id = r['id'] if r else None
            else:
                # Tenter de deviner : si col0 ressemble à un code DR connu
                upper0 = col0.upper()
                r = db.execute("SELECT id FROM directions_regionales WHERE code=?", (upper0,)).fetchone()
                if r:
                    scope_type = 'dr'
                    scope_id = r['id']
                else:
                    scope_type = scope_type_raw or 'national'

            db.execute("""INSERT INTO objectifs (exercice, scope_type, scope_id, rubrique, mois, montant)
                          VALUES (?, ?, ?, ?, ?, ?)
                          ON CONFLICT(exercice, scope_type, scope_id, rubrique, mois)
                          DO UPDATE SET montant=excluded.montant""",
                       (exercice, scope_type, scope_id, rubrique, mois_val, valeur))
            count += 1

        # Enregistrer le fichier
        operateur = session.get('operateur', {})
        db.execute("""INSERT INTO objectifs_fichiers (nom_fichier, exercice, operateur_nom, operateur_matricule, nb_lignes_importees)
                      VALUES (?, ?, ?, ?, ?)""",
                   (filename, exercice, operateur.get('nom'), operateur.get('matricule'), count))
        db.commit()
        db.close()
        wb.close()
        return jsonify({"status": "ok", "lignes_importees": count, "fichier": filename})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/objectifs', methods=['GET'])
def api_get_objectifs():
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    rows = db.execute("""
        SELECT o.*, d.code as dr_code, a.nom as agence_nom
        FROM objectifs o
        LEFT JOIN directions_regionales d ON o.scope_type='dr' AND o.scope_id = d.id
        LEFT JOIN agences a ON o.scope_type='agence' AND o.scope_id = a.id
        WHERE o.exercice=?
        ORDER BY
            CASE o.scope_type WHEN 'national' THEN 0 WHEN 'dr' THEN 1 WHEN 'agence' THEN 2 ELSE 3 END,
            d.code, a.nom, o.rubrique
    """, (exercice,)).fetchall()
    fichiers = db.execute("SELECT * FROM objectifs_fichiers WHERE exercice=? ORDER BY date_upload DESC",
                          (exercice,)).fetchall()
    db.close()

    objectifs_list = []
    for r in rows:
        site_label = 'CAMWATER (National)'
        if r['scope_type'] == 'dr' and r['dr_code']:
            site_label = r['dr_code']
        elif r['scope_type'] == 'agence' and r['agence_nom']:
            site_label = r['agence_nom']
        elif r['scope_type'] not in ('national', 'dr', 'agence'):
            site_label = r['scope_type'].upper()

        # Valeur = montant, sauf si l'ancien format a mis la valeur dans mois
        valeur = r['montant'] if r['montant'] else (r['mois'] if r['mois'] and r['mois'] > 12 else 0)

        objectifs_list.append({
            "scope_type": r['scope_type'],
            "scope_id": r['scope_id'],
            "site_label": site_label,
            "rubrique": r['rubrique'],
            "mois": r['mois'],
            "valeur": valeur,
        })

    return jsonify({
        "objectifs": objectifs_list,
        "fichiers": [dict(r) for r in fichiers],
    })


# ─── API Objectifs : Suppression ─────────────────────────────────

@app.route('/api/objectifs/delete', methods=['DELETE'])
def api_delete_objectifs():
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    db.execute("DELETE FROM objectifs WHERE exercice=?", (exercice,))
    db.execute("DELETE FROM objectifs_fichiers WHERE exercice=?", (exercice,))
    db.commit()
    db.close()
    # Supprimer les fichiers physiques du dossier
    import glob
    for fpath in glob.glob(os.path.join(UPLOAD_FOLDER, '*')):
        try:
            os.remove(fpath)
        except Exception:
            pass
    return jsonify({"status": "ok", "message": f"Objectifs et fichiers supprimés pour l'exercice {exercice}"})


# ─── API Réalisations Antérieures ────────────────────────────────

@app.route('/api/realisations/upload', methods=['POST'])
def api_upload_realisations():
    """Upload réalisations N-1 depuis un fichier Excel.
    Format principal (5 colonnes) : DR | Agence | Rubrique | Période | Valeur
      - Période = numéro de mois (1-12) ou nom (Janvier, Février, etc.)
    Format legacy (multi-onglets) : auto-détecté si > 5 colonnes numériques.
    """
    if 'file' not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Format Excel requis (.xlsx)"}), 400
    exercice = request.form.get('exercice', 2025, type=int)

    # Codes DR connus
    dr_codes_upper = set(STRUCTURE_CW.keys())
    # Toutes les agences connues (nom → dr_code)
    all_agences_map = {}
    for dr_c, ag_list in STRUCTURE_CW.items():
        for ag in ag_list:
            all_agences_map[ag.lower()] = dr_c

    # Mapping de noms d'onglets/headers → rubrique normalisée
    RUBRIQUE_MAP = {
        'volumes': 'Volumes', 'volume': 'Volumes', 'vol': 'Volumes',
        'volumes facturés': 'Volumes', 'volumes factures': 'Volumes',
        'production': 'Volumes', 'vol factures': 'Volumes',
        'ca vente eau': 'CA Vente Eau', 'ca ve': 'CA Vente Eau', 'cave': 'CA Vente Eau',
        'chiffre affaires': 'CA Vente Eau', 'chiffre d\'affaires': 'CA Vente Eau',
        'ca travaux': 'CA Travaux', 'travaux': 'CA Travaux', 'ca trvx': 'CA Travaux',
        'ca branchements': 'CA Travaux', 'travaux remboursables': 'CA Travaux',
        'encaissements': 'Encaissements', 'enc': 'Encaissements', 'encaissement': 'Encaissements',
        'recouvrement': 'Encaissements',
        'impayés actifs': 'Impayés Actifs', 'impayes actifs': 'Impayés Actifs',
        'impayés résiliés': 'Impayés Résiliés', 'impayes resilies': 'Impayés Résiliés',
        'impayés': 'Impayés', 'impayes': 'Impayés', 'portefeuille impayés': 'Impayés',
        'branchements vendus': 'Branchements Vendus', 'bts vendus': 'Branchements Vendus',
        'branchements exécutés': 'Branchements Exécutés', 'bts exécutés': 'Branchements Exécutés',
        'branchements executes': 'Branchements Exécutés',
        'branchements': 'Branchements Vendus',
        'recettes': 'Recettes', 'recettes commerciales': 'Recettes',
    }

    def normalize_rubrique(name):
        if not name:
            return None
        key = name.strip().lower()
        return RUBRIQUE_MAP.get(key, name.strip())

    def parse_mois(val):
        """Convertir une valeur (numéro ou nom) en numéro de mois (1-12)."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            m = int(val)
            return m if 1 <= m <= 12 else None
        s = str(val).strip().lower()
        # Numéro direct
        try:
            m = int(float(s))
            return m if 1 <= m <= 12 else None
        except (ValueError, TypeError):
            pass
        # Nom de mois (début suffit)
        mois_noms = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                     'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        mois_noms_alt = ['janv', 'fev', 'mars', 'avr', 'mai', 'juin',
                         'juil', 'aout', 'sept', 'oct', 'nov', 'dec']
        for i, (mn, ma) in enumerate(zip(mois_noms, mois_noms_alt), 1):
            if s.startswith(mn[:3]) or s.startswith(ma[:3]):
                return i
        return None

    def detect_format(wb):
        """Détecter le format du fichier :
        - 'simple' si 5 colonnes type DR|Agence|Rubrique|Période|Valeur
        - 'legacy' si multi-colonnes mensuelles (ancien format)
        """
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if not row:
                continue
            # Compter les colonnes non-vides
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 3:
                # Vérifier les en-têtes pour le format simple
                headers_lower = [str(c).strip().lower() for c in row if c is not None]
                period_kw = ['période', 'periode', 'mois', 'period', 'month']
                has_period_header = any(any(pk in h for pk in period_kw) for h in headers_lower)
                # Si on trouve un en-tête "période/mois" et ≤ 6 colonnes non-vides → format simple
                if has_period_header and len(non_empty) <= 7:
                    return 'simple'
                # Si > 8 colonnes → probablement legacy avec colonnes mensuelles
                if len(non_empty) > 8:
                    return 'legacy'
        # Par défaut on tente le format simple (plus courant désormais)
        return 'simple'

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        db = get_db()

        # ── Cache agence_id en mémoire (évite des milliers de SELECT) ──
        agence_id_cache = {}
        for row_ag in db.execute("SELECT id, LOWER(nom) as nom_lower FROM agences").fetchall():
            agence_id_cache[row_ag['nom_lower']] = row_ag['id']

        def resolve_agence(agence_nom_raw):
            """Résout agence_id + dr_code depuis le cache mémoire."""
            if not agence_nom_raw:
                return None, None
            ag_lower = agence_nom_raw.lower().strip()
            ag_id = agence_id_cache.get(ag_lower)
            dr_from_ag = all_agences_map.get(ag_lower)
            return ag_id, dr_from_ag

        # ── Pré-scanner les rubriques du fichier (max 200 lignes) ──
        file_format = detect_format(wb)
        rubriques_in_file = set()
        for ws_scan in wb.worksheets:
            sheet_rub = normalize_rubrique(ws_scan.title)
            if sheet_rub and sheet_rub not in (ws_scan.title.strip(),):
                rubriques_in_file.add(sheet_rub)
            for row_scan in ws_scan.iter_rows(min_row=1, max_row=min(ws_scan.max_row or 50, 200), values_only=True):
                if not row_scan or len(row_scan) < 3:
                    continue
                cell2 = str(row_scan[2]).strip() if row_scan[2] else ''
                rub_norm = normalize_rubrique(cell2)
                if rub_norm:
                    rubriques_in_file.add(rub_norm)

        # Supprimer uniquement les rubriques qui seront ré-importées
        if rubriques_in_file:
            for rub_del in rubriques_in_file:
                db.execute("DELETE FROM realisations_anterieures WHERE exercice=? AND rubrique=?",
                           (exercice, rub_del))
        else:
            db.execute("DELETE FROM realisations_anterieures WHERE exercice=?", (exercice,))

        BATCH_SIZE = 2000
        batch = []
        count = 0

        def flush_batch():
            nonlocal batch
            if batch:
                db.executemany("""INSERT INTO realisations_anterieures
                    (exercice, dr_code, agence_nom, agence_id, rubrique, mois, valeur, total_annuel)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""", batch)
                batch = []

        if file_format == 'simple':
            # ═══ Format simplifié : DR | Agence | Rubrique | Période | Valeur ═══
            for ws in wb.worksheets:
                skip_header = True
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if not row or len(row) < 4:
                        continue

                    if skip_header:
                        first_val = str(row[0]).strip().lower() if row[0] else ''
                        if first_val in ('dr', 'direction', 'dr_code', 'code_dr', 'code dr', ''):
                            continue
                        last_cells = [c for c in row[-2:] if c is not None]
                        all_text = all(isinstance(c, str) or c is None for c in last_cells)
                        if all_text and len(last_cells) > 0:
                            skip_header = False
                            continue
                        skip_header = False

                    col_vals = [str(c).strip() if c is not None else '' for c in row[:6]]
                    dr_code_raw = col_vals[0] if len(col_vals) > 0 else ''
                    agence_nom = col_vals[1] if len(col_vals) > 1 else ''
                    rubrique_raw = col_vals[2] if len(col_vals) > 2 else ''

                    if len(row) >= 5 and row[4] is not None:
                        periode_raw = row[3]
                        valeur_raw = row[4]
                    elif len(row) >= 4:
                        periode_raw = None
                        valeur_raw = row[3]
                    else:
                        continue

                    try:
                        if isinstance(valeur_raw, (int, float)):
                            valeur = float(valeur_raw)
                        else:
                            v_str = str(valeur_raw).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
                            valeur = float(v_str)
                    except (ValueError, TypeError):
                        continue
                    if valeur == 0:
                        continue

                    mois_val = parse_mois(periode_raw)

                    dr_code = dr_code_raw.upper().strip()
                    if dr_code and dr_code not in dr_codes_upper:
                        for known in dr_codes_upper:
                            if known in dr_code or dr_code in known:
                                dr_code = known
                                break

                    rubrique = normalize_rubrique(rubrique_raw) or rubrique_raw.strip()

                    agence_id, dr_from_ag = resolve_agence(agence_nom)
                    if not dr_code and dr_from_ag:
                        dr_code = dr_from_ag

                    batch.append((exercice, dr_code or None, agence_nom.strip() or None,
                                  agence_id, rubrique, mois_val, valeur, 0))
                    count += 1

                    if len(batch) >= BATCH_SIZE:
                        flush_batch()

        else:
            # ═══ Format legacy : multi-onglets, colonnes mensuelles ═══
            def detect_site_type(cell_val):
                if not cell_val:
                    return None, None, None, None
                val = str(cell_val).strip()
                val_upper = val.upper()
                val_lower = val.lower()
                if val_lower in ('agence', 'site', 'centres', 'dr', ''):
                    return None, None, None, None
                if 'camwater' in val_lower or val_lower in ('total', 'total général', 'total general', 'national'):
                    return 'national', None, 'CAMWATER', None
                if val_upper in dr_codes_upper:
                    return 'dr', val_upper, val_upper, None
                for dc in dr_codes_upper:
                    if dc in val_upper and ('total' in val_lower or 's/t' in val_lower or 'sous' in val_lower):
                        return 'dr', dc, dc, None
                if val_lower in all_agences_map:
                    return 'agence', all_agences_map[val_lower], val, None
                return 'agence', None, val, None

            for ws in wb.worksheets:
                sheet_rubrique = normalize_rubrique(ws.title)
                current_rubrique = sheet_rubrique or 'Inconnu'
                current_dr = None
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if not row or not row[0]:
                        continue
                    cell_a = str(row[0]).strip() if row[0] else ''
                    has_numbers = any(isinstance(c, (int, float)) and c != 0 for c in row[1:14] if c is not None)
                    if not has_numbers:
                        maybe_rub = normalize_rubrique(cell_a)
                        if maybe_rub and maybe_rub != cell_a.strip():
                            current_rubrique = maybe_rub
                        continue
                    site_type, dr_code, site_name, _ = detect_site_type(cell_a)
                    if not site_type:
                        continue
                    if site_type == 'agence' and dr_code:
                        current_dr = dr_code
                    elif site_type == 'agence' and not dr_code and current_dr:
                        dr_code = current_dr
                    elif site_type == 'dr':
                        current_dr = dr_code

                    agence_id, _ = resolve_agence(site_name)

                    for m in range(1, 13):
                        val = 0
                        if len(row) > m and row[m] is not None:
                            try:
                                val = float(row[m])
                            except (ValueError, TypeError):
                                val = 0
                        if val != 0:
                            total_annuel = 0
                            if len(row) > 13 and row[13] is not None:
                                try:
                                    total_annuel = float(row[13])
                                except (ValueError, TypeError):
                                    total_annuel = 0
                            batch.append((exercice, dr_code, site_name, agence_id,
                                          current_rubrique, m, val, total_annuel))
                            count += 1

                            if len(batch) >= BATCH_SIZE:
                                flush_batch()

        flush_batch()
        db.commit()
        db.close()
        wb.close()
        return jsonify({"status": "ok", "lignes": count, "format": file_format})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/realisations')
def api_get_realisations():
    exercice = request.args.get('exercice', 2025, type=int)
    rubrique = request.args.get('rubrique', '')
    dr = request.args.get('dr', '')
    mois_filter = request.args.get('mois', '', type=str)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 500, type=int)
    per_page = min(per_page, 5000)  # Limiter à 5000 max par page

    db = get_db()

    # Clause WHERE commune
    where = " WHERE exercice=?"
    params = [exercice]
    if rubrique:
        where += " AND rubrique=?"
        params.append(rubrique)
    if dr:
        where += " AND dr_code=?"
        params.append(dr)
    if mois_filter:
        where += " AND mois=?"
        params.append(int(mois_filter))

    # Résumé rapide (toujours retourné)
    summary_rows = db.execute(
        "SELECT rubrique, COUNT(*) as n, SUM(valeur) as total FROM realisations_anterieures"
        + where + " GROUP BY rubrique ORDER BY rubrique", params).fetchall()
    summary = [{"rubrique": s['rubrique'], "count": s['n'], "total": s['total'] or 0} for s in summary_rows]
    total_count = sum(s['count'] for s in summary)

    # Données paginées
    offset = (page - 1) * per_page
    query = ("SELECT dr_code, agence_nom, rubrique, mois, valeur FROM realisations_anterieures"
             + where + " ORDER BY rubrique, dr_code, agence_nom, mois LIMIT ? OFFSET ?")
    rows = db.execute(query, params + [per_page, offset]).fetchall()
    db.close()

    lignes = []
    for r in rows:
        lignes.append({
            "dr_code": r['dr_code'] or '-',
            "agence": r['agence_nom'] or '-',
            "rubrique": r['rubrique'],
            "mois": r['mois'],
            "valeur": r['valeur'] or 0
        })

    return jsonify({
        "lignes": lignes,
        "summary": summary,
        "total_count": total_count,
        "page": page,
        "per_page": per_page,
        "total_pages": (total_count + per_page - 1) // per_page if per_page > 0 else 1,
    })


@app.route('/api/realisations/delete', methods=['DELETE'])
def api_delete_realisations():
    exercice = request.args.get('exercice', 2025, type=int)
    db = get_db()
    db.execute("DELETE FROM realisations_anterieures WHERE exercice=?", (exercice,))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "message": f"Réalisations supprimées pour l'exercice {exercice}"})


# ─── Recettes Quotidiennes (Suivi Journalier) ────────────────────

@app.route('/recettes-jour')
def recettes_jour_login_page():
    agences = get_all_agences()
    return render_template('recettes_jour_login.html', agences=agences,
                           structure_keys=list(STRUCTURE_CW.keys()))


@app.route('/recettes-jour/saisie')
def recettes_jour_saisie_page():
    return render_template('recettes_jour_saisie.html')


@app.route('/recettes-jour/suivi')
def recettes_jour_suivi_page():
    role = session.get('role')
    return render_template('recettes_jour_suivi.html',
                           drs=list(STRUCTURE_CW.keys()),
                           role=role,
                           read_only=(role == 'direction'))


@app.route('/api/recettes-jour/login', methods=['POST'])
def api_recettes_jour_login():
    import re as _re
    data = request.json
    matricule = data.get('matricule', '').strip().upper()
    agence_id = data.get('agence_id')
    date_saisie = data.get('date_saisie', '')
    if not all([matricule, agence_id, date_saisie]):
        return jsonify({"error": "Tous les champs sont requis"}), 400
    if not _re.match(r'^[0-9][A-Z][0-9][0-9]$', matricule):
        return jsonify({"error": "Format matricule invalide (ex: 1A23)"}), 400
    db = get_db()
    ag = db.execute("""
        SELECT a.id, a.nom, d.code as dr_code
        FROM agences a JOIN directions_regionales d ON a.dr_id = d.id
        WHERE a.id = ?
    """, (agence_id,)).fetchone()
    db.close()
    if not ag:
        return jsonify({"error": "Agence non trouvee"}), 404
    # Purge d'une éventuelle session centrale avant de passer en mode terrain
    session.pop('operateur', None)
    session['op_recettes_jour'] = {
        "matricule": matricule,
        "agence_id": ag['id'], "agence_nom": ag['nom'],
        "dr_code": ag['dr_code'], "date_saisie": date_saisie
    }
    # Restriction d'accès : agent terrain uniquement
    session['role'] = 'agent_terrain'
    return jsonify({"status": "ok", "agence_nom": ag['nom'], "dr_code": ag['dr_code']})


@app.route('/api/recettes-jour', methods=['GET'])
def api_get_recettes_jour():
    agence_id = request.args.get('agence_id', type=int)
    date_saisie = request.args.get('date', '')
    db = get_db()
    row = db.execute("SELECT * FROM recettes_jour WHERE agence_id=? AND date_saisie=?",
                     (agence_id, date_saisie)).fetchone()
    db.close()
    if row:
        return jsonify(dict(row))
    return jsonify({"agence_id": agence_id, "date_saisie": date_saisie})


@app.route('/api/recettes-jour', methods=['POST'])
def api_save_recettes_jour():
    data = request.json
    agence_id = data.get('agence_id')
    date_saisie = data.get('date_saisie')
    db = get_db()
    existing = db.execute("SELECT verrouille FROM recettes_jour WHERE agence_id=? AND date_saisie=?",
                          (agence_id, date_saisie)).fetchone()
    if existing and existing['verrouille']:
        db.close()
        return jsonify({"error": "Donnees deja verrouillees pour cette date"}), 403
    total = sum([
        data.get('caisse_commerciale', 0) or 0,
        data.get('cheques', 0) or 0,
        data.get('hors_sites', 0) or 0,
        data.get('virements', 0) or 0,
        data.get('paiements_electroniques', 0) or 0,
    ])
    db.execute("""INSERT INTO recettes_jour (agence_id, date_saisie, caisse_commerciale, cheques,
                      hors_sites, virements, paiements_electroniques, total,
                      type_piece, numero_piece, convoyeur, banque_depot,
                      operateur_nom, operateur_matricule, verrouille)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                  ON CONFLICT(agence_id, date_saisie)
                  DO UPDATE SET caisse_commerciale=excluded.caisse_commerciale,
                               cheques=excluded.cheques, hors_sites=excluded.hors_sites,
                               virements=excluded.virements,
                               paiements_electroniques=excluded.paiements_electroniques,
                               total=excluded.total, type_piece=excluded.type_piece,
                               numero_piece=excluded.numero_piece, convoyeur=excluded.convoyeur,
                               banque_depot=excluded.banque_depot,
                               operateur_nom=excluded.operateur_nom,
                               operateur_matricule=excluded.operateur_matricule,
                               verrouille=1""",
               (agence_id, date_saisie,
                data.get('caisse_commerciale', 0) or 0,
                data.get('cheques', 0) or 0,
                data.get('hors_sites', 0) or 0,
                data.get('virements', 0) or 0,
                data.get('paiements_electroniques', 0) or 0,
                total,
                data.get('type_piece', ''),
                data.get('numero_piece', ''),
                data.get('convoyeur', ''),
                data.get('banque_depot', ''),
                data.get('operateur_nom', ''),
                data.get('operateur_matricule', '')))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "total": total})


@app.route('/api/recettes-jour/synthese')
def api_recettes_jour_synthese():
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', '')
    db = get_db()
    if site_type == 'dr' and site_id:
        agences_rows = db.execute(
            "SELECT a.id FROM agences a JOIN directions_regionales d ON a.dr_id=d.id WHERE d.code=?",
            (site_id,)).fetchall()
    else:
        agences_rows = db.execute("SELECT id FROM agences").fetchall()
    agence_ids = [r['id'] for r in agences_rows]
    if not agence_ids:
        db.close()
        return jsonify({"total_agences": 0, "agences_renseignees": 0,
                        "totaux": {"caisse_commerciale": 0, "cheques": 0, "hors_sites": 0,
                                   "virements": 0, "paiements_electroniques": 0, "total": 0},
                        "details": [], "manquants": []})
    ph = ','.join('?' * len(agence_ids))
    rows = db.execute(f"""
        SELECT r.*, a.nom as agence_nom, d.code as dr_code
        FROM recettes_jour r
        JOIN agences a ON r.agence_id = a.id
        JOIN directions_regionales d ON a.dr_id = d.id
        WHERE r.agence_id IN ({ph})
          AND r.date_saisie BETWEEN ? AND ?
          AND r.verrouille = 1
        ORDER BY d.code, a.nom, r.date_saisie
    """, agence_ids + [date_debut, date_fin]).fetchall()
    totaux = {"caisse_commerciale": 0, "cheques": 0, "hors_sites": 0,
              "virements": 0, "paiements_electroniques": 0, "total": 0}
    details = []
    agences_renseignees = set()
    for r in rows:
        d = dict(r)
        details.append(d)
        agences_renseignees.add(d['agence_id'])
        for k in totaux:
            totaux[k] += d.get(k, 0) or 0
    all_agences = db.execute(f"""
        SELECT a.id, a.nom, d.code as dr_code
        FROM agences a JOIN directions_regionales d ON a.dr_id = d.id
        WHERE a.id IN ({ph}) ORDER BY d.code, a.nom
    """, agence_ids).fetchall()
    manquants = [{"id": a['id'], "nom": a['nom'], "dr_code": a['dr_code']}
                 for a in all_agences if a['id'] not in agences_renseignees]
    db.close()
    return jsonify({
        "total_agences": len(agence_ids),
        "agences_renseignees": len(agences_renseignees),
        "totaux": totaux,
        "details": details,
        "manquants": manquants
    })


@app.route('/api/recettes-jour/cumul')
def api_recettes_jour_cumul():
    agence_id = request.args.get('agence_id', type=int)
    db = get_db()
    row = db.execute(
        "SELECT COALESCE(SUM(total), 0) as cumul FROM recettes_jour "
        "WHERE agence_id=? AND verrouille=1", (agence_id,)).fetchone()
    db.close()
    return jsonify({"cumul": row['cumul'] if row else 0})


@app.route('/api/recettes-jour/convoyeurs')
def api_recettes_jour_convoyeurs():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT convoyeur FROM recettes_jour "
        "WHERE convoyeur IS NOT NULL AND convoyeur != '' ORDER BY convoyeur"
    ).fetchall()
    db.close()
    return jsonify([r['convoyeur'] for r in rows])


# ─── Branchements Quotidiens (Suivi Journalier) ────────────────────

@app.route('/branchements-jour')
def branchements_jour_login_page():
    agences = get_all_agences()
    return render_template('branchements_jour_login.html', agences=agences,
                           structure_keys=list(STRUCTURE_CW.keys()))


@app.route('/branchements-jour/saisie')
def branchements_jour_saisie_page():
    return render_template('branchements_jour_saisie.html')


@app.route('/branchements-jour/suivi')
def branchements_jour_suivi_page():
    return render_template('branchements_jour_suivi.html',
                           drs=list(STRUCTURE_CW.keys()))


@app.route('/api/branchements-jour/login', methods=['POST'])
def api_branchements_jour_login():
    import re as _re
    data = request.json
    matricule = data.get('matricule', '').strip().upper()
    agence_id = data.get('agence_id')
    date_saisie = data.get('date_saisie', '')
    if not all([matricule, agence_id, date_saisie]):
        return jsonify({"error": "Tous les champs sont requis"}), 400
    if not _re.match(r'^[0-9][A-Z][0-9][0-9]$', matricule):
        return jsonify({"error": "Format matricule invalide (ex: 1A23)"}), 400
    db = get_db()
    ag = db.execute("""
        SELECT a.id, a.nom, d.code as dr_code
        FROM agences a JOIN directions_regionales d ON a.dr_id = d.id
        WHERE a.id = ?
    """, (agence_id,)).fetchone()
    db.close()
    if not ag:
        return jsonify({"error": "Agence non trouvee"}), 404
    # Purge d'une éventuelle session centrale avant de passer en mode terrain
    session.pop('operateur', None)
    session['op_branchements_jour'] = {
        "matricule": matricule,
        "agence_id": ag['id'], "agence_nom": ag['nom'],
        "dr_code": ag['dr_code'], "date_saisie": date_saisie
    }
    # Restriction d'accès : agent terrain uniquement
    session['role'] = 'agent_terrain'
    return jsonify({"status": "ok", "agence_nom": ag['nom'], "dr_code": ag['dr_code']})


@app.route('/api/branchements-jour', methods=['GET'])
def api_get_branchements_jour():
    agence_id = request.args.get('agence_id', type=int)
    date_saisie = request.args.get('date', '')
    db = get_db()
    row = db.execute("SELECT * FROM branchements_jour WHERE agence_id=? AND date_saisie=?",
                     (agence_id, date_saisie)).fetchone()
    db.close()
    if row:
        return jsonify(dict(row))
    return jsonify({"agence_id": agence_id, "date_saisie": date_saisie})


@app.route('/api/branchements-jour', methods=['POST'])
def api_save_branchements_jour():
    data = request.json
    agence_id = data.get('agence_id')
    date_saisie = data.get('date_saisie')
    db = get_db()
    existing = db.execute("SELECT verrouille FROM branchements_jour WHERE agence_id=? AND date_saisie=?",
                          (agence_id, date_saisie)).fetchone()
    if existing and existing['verrouille']:
        db.close()
        return jsonify({"error": "Donnees deja verrouillees pour cette date"}), 403
    vendus = (data.get('vendus', 0) or 0)
    executes = (data.get('executes', 0) or 0)
    pec_machine = (data.get('pec_machine', 0) or 0)
    moratoire_val = (data.get('moratoire', 0) or 0)
    total = vendus + executes + pec_machine + moratoire_val
    db.execute("""INSERT INTO branchements_jour (agence_id, date_saisie,
                      vendus_d40, vendus_d20, vendus,
                      executes_d40, executes_d20, executes,
                      pec_d40, pec_d20, pec_machine,
                      moratoire_d40, moratoire_d20, moratoire,
                      total, observations,
                      operateur_nom, operateur_matricule, verrouille)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                  ON CONFLICT(agence_id, date_saisie)
                  DO UPDATE SET vendus_d40=excluded.vendus_d40, vendus_d20=excluded.vendus_d20,
                               vendus=excluded.vendus,
                               executes_d40=excluded.executes_d40, executes_d20=excluded.executes_d20,
                               executes=excluded.executes,
                               pec_d40=excluded.pec_d40, pec_d20=excluded.pec_d20,
                               pec_machine=excluded.pec_machine,
                               moratoire_d40=excluded.moratoire_d40, moratoire_d20=excluded.moratoire_d20,
                               moratoire=excluded.moratoire,
                               total=excluded.total, observations=excluded.observations,
                               operateur_nom=excluded.operateur_nom,
                               operateur_matricule=excluded.operateur_matricule,
                               verrouille=1""",
               (agence_id, date_saisie,
                data.get('vendus_d40', 0) or 0, data.get('vendus_d20', 0) or 0, vendus,
                data.get('executes_d40', 0) or 0, data.get('executes_d20', 0) or 0, executes,
                data.get('pec_d40', 0) or 0, data.get('pec_d20', 0) or 0, pec_machine,
                data.get('moratoire_d40', 0) or 0, data.get('moratoire_d20', 0) or 0, moratoire_val,
                total, data.get('observations', ''),
                data.get('operateur_nom', ''),
                data.get('operateur_matricule', '')))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "total": total})


@app.route('/api/branchements-jour/synthese')
def api_branchements_jour_synthese():
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', '')
    db = get_db()
    if site_type == 'dr' and site_id:
        agences_rows = db.execute(
            "SELECT a.id FROM agences a JOIN directions_regionales d ON a.dr_id=d.id WHERE d.code=?",
            (site_id,)).fetchall()
    else:
        agences_rows = db.execute("SELECT id FROM agences").fetchall()
    agence_ids = [r['id'] for r in agences_rows]
    if not agence_ids:
        db.close()
        return jsonify({"total_agences": 0, "agences_renseignees": 0,
                        "totaux": {"vendus": 0, "executes": 0, "pec_machine": 0,
                                   "moratoire": 0, "total": 0},
                        "details": [], "manquants": []})
    ph = ','.join('?' * len(agence_ids))
    rows = db.execute(f"""
        SELECT b.*, a.nom as agence_nom, d.code as dr_code
        FROM branchements_jour b
        JOIN agences a ON b.agence_id = a.id
        JOIN directions_regionales d ON a.dr_id = d.id
        WHERE b.agence_id IN ({ph})
          AND b.date_saisie BETWEEN ? AND ?
          AND b.verrouille = 1
        ORDER BY d.code, a.nom, b.date_saisie
    """, agence_ids + [date_debut, date_fin]).fetchall()
    totaux = {"vendus": 0, "executes": 0, "pec_machine": 0,
              "moratoire": 0, "total": 0}
    details = []
    agences_renseignees = set()
    for r in rows:
        d = dict(r)
        details.append(d)
        agences_renseignees.add(d['agence_id'])
        for k in totaux:
            totaux[k] += d.get(k, 0) or 0
    all_agences = db.execute(f"""
        SELECT a.id, a.nom, d.code as dr_code
        FROM agences a JOIN directions_regionales d ON a.dr_id = d.id
        WHERE a.id IN ({ph}) ORDER BY d.code, a.nom
    """, agence_ids).fetchall()
    manquants = [{"id": a['id'], "nom": a['nom'], "dr_code": a['dr_code']}
                 for a in all_agences if a['id'] not in agences_renseignees]
    db.close()
    return jsonify({
        "total_agences": len(agence_ids),
        "agences_renseignees": len(agences_renseignees),
        "totaux": totaux,
        "details": details,
        "manquants": manquants
    })


@app.route('/api/branchements-jour/cumul')
def api_branchements_jour_cumul():
    agence_id = request.args.get('agence_id', type=int)
    db = get_db()
    row = db.execute(
        "SELECT COALESCE(SUM(total), 0) as cumul FROM branchements_jour "
        "WHERE agence_id=? AND verrouille=1", (agence_id,)).fetchone()
    db.close()
    return jsonify({"cumul": row['cumul'] if row else 0})


@app.route('/api/branchements-jour/dashboard-stats')
def api_branchements_jour_dashboard_stats():
    """Stats branchements quotidiens pour le tableau de bord commercial."""
    import datetime
    site_type = request.args.get('site_type', 'national')
    site_id = request.args.get('site_id', '')
    db = get_db()

    # Agences filtrees
    if site_type == 'dr' and site_id:
        agences_rows = db.execute(
            "SELECT a.id FROM agences a JOIN directions_regionales d ON a.dr_id=d.id WHERE d.code=?",
            (site_id,)).fetchall()
    else:
        agences_rows = db.execute("SELECT id FROM agences").fetchall()
    agence_ids = [r['id'] for r in agences_rows]

    if not agence_ids:
        db.close()
        return jsonify({"totaux": {}, "par_dr": {}, "evolution": [], "centres": "0 / 0"})

    ph = ','.join('?' * len(agence_ids))
    today = datetime.date.today()
    mois_debut = today.replace(day=1).isoformat()
    mois_fin = today.isoformat()

    # Totaux du mois en cours
    row = db.execute(f"""
        SELECT COUNT(DISTINCT agence_id) as nb_centres,
               COALESCE(SUM(vendus),0) as vendus,
               COALESCE(SUM(executes),0) as executes,
               COALESCE(SUM(pec_machine),0) as pec_machine,
               COALESCE(SUM(moratoire),0) as moratoire,
               COALESCE(SUM(total),0) as total
        FROM branchements_jour
        WHERE agence_id IN ({ph}) AND verrouille=1
          AND date_saisie BETWEEN ? AND ?
    """, agence_ids + [mois_debut, mois_fin]).fetchone()

    # Par DR (mois en cours)
    dr_rows = db.execute(f"""
        SELECT d.code as dr_code,
               COALESCE(SUM(b.vendus),0) as vendus,
               COALESCE(SUM(b.executes),0) as executes,
               COALESCE(SUM(b.pec_machine),0) as pec_machine,
               COALESCE(SUM(b.moratoire),0) as moratoire
        FROM branchements_jour b
        JOIN agences a ON b.agence_id = a.id
        JOIN directions_regionales d ON a.dr_id = d.id
        WHERE b.agence_id IN ({ph}) AND b.verrouille=1
          AND b.date_saisie BETWEEN ? AND ?
        GROUP BY d.code ORDER BY d.code
    """, agence_ids + [mois_debut, mois_fin]).fetchall()

    # Evolution quotidienne (30 derniers jours)
    date_30j = (today - datetime.timedelta(days=30)).isoformat()
    evol_rows = db.execute(f"""
        SELECT date_saisie,
               COALESCE(SUM(vendus),0) as vendus,
               COALESCE(SUM(executes),0) as executes,
               COALESCE(SUM(pec_machine),0) as pec_machine
        FROM branchements_jour
        WHERE agence_id IN ({ph}) AND verrouille=1
          AND date_saisie >= ?
        GROUP BY date_saisie ORDER BY date_saisie
    """, agence_ids + [date_30j]).fetchall()

    db.close()

    return jsonify({
        "centres": f"{row['nb_centres']} / {len(agence_ids)}",
        "totaux": {
            "vendus": row['vendus'], "executes": row['executes'],
            "pec_machine": row['pec_machine'], "moratoire": row['moratoire'],
            "total": row['total']
        },
        "par_dr": {r['dr_code']: {"vendus": r['vendus'], "executes": r['executes'],
                                   "pec_machine": r['pec_machine'], "moratoire": r['moratoire']}
                   for r in dr_rows},
        "evolution": [{"date": r['date_saisie'], "vendus": r['vendus'],
                       "executes": r['executes'], "pec_machine": r['pec_machine']}
                      for r in evol_rows]
    })


# ─── API Consultation : Paiements Électroniques ─────────────────

@app.route('/api/paiements-elec/upload', methods=['POST'])
def api_upload_paiements_elec():
    if 'file' not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Format Excel requis (.xlsx)"}), 400
    exercice = request.form.get('exercice', 2026, type=int)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        db = get_db()
        # Supprimer les anciennes données pour cet exercice
        db.execute("DELETE FROM paiements_electroniques WHERE exercice=?", (exercice,))
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            dr_code = str(row[0]).strip() if row[0] else None
            agence_nom = str(row[1]).strip() if row[1] else None
            mois_val = int(row[2]) if row[2] else None
            mode = str(row[3]).strip() if len(row) > 3 and row[3] else 'Non précisé'
            montant = float(row[4]) if len(row) > 4 and row[4] else (float(row[3]) if len(row) > 3 and row[3] and not mode else 0)
            # Résoudre agence_id
            agence_id = None
            if agence_nom:
                r = db.execute("SELECT id FROM agences WHERE nom=?", (agence_nom,)).fetchone()
                agence_id = r['id'] if r else None
            db.execute("""INSERT INTO paiements_electroniques
                          (agence_id, dr_code, agence_nom, mois, exercice, mode_paiement, montant)
                          VALUES (?, ?, ?, ?, ?, ?, ?)""",
                       (agence_id, dr_code, agence_nom, mois_val, exercice, mode, montant))
            count += 1
        db.commit()
        db.close()
        wb.close()
        return jsonify({"status": "ok", "lignes": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/paiements-elec')
def api_get_paiements_elec():
    exercice = request.args.get('exercice', 2026, type=int)
    dr = request.args.get('dr', '')
    db = get_db()
    if dr:
        rows = db.execute(
            "SELECT * FROM paiements_electroniques WHERE exercice=? AND dr_code=? ORDER BY mois, agence_nom",
            (exercice, dr)).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM paiements_electroniques WHERE exercice=? ORDER BY dr_code, mois, agence_nom",
            (exercice,)).fetchall()
    db.close()
    lignes = [{"dr_code": r['dr_code'], "agence": r['agence_nom'], "mois": MOIS[r['mois']-1] if r['mois'] else '-',
               "mode": r['mode_paiement'], "montant": r['montant']} for r in rows]
    return jsonify({"lignes": lignes})


# ─── API Consultation : Impayés Détail par DR ───────────────────

@app.route('/api/consultation/impayes-detail')
def api_consultation_impayes_detail():
    mois = request.args.get('mois', 1, type=int)
    dr_filter = request.args.get('dr', '')
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()

    query = """
        SELECT d.code as dr_code, a.nom as agence_nom,
               COALESCE(i.particuliers_actifs,0) + COALESCE(i.gco_actifs,0) +
               COALESCE(i.bf_actifs,0) + COALESCE(i.bfc_actifs,0) as actifs,
               COALESCE(i.particuliers_resilies,0) + COALESCE(i.gco_resilies,0) +
               COALESCE(i.bf_resilies,0) + COALESCE(i.bfc_resilies,0) +
               COALESCE(i.gestion_manuelle,0) as resilies
        FROM agences a
        JOIN directions_regionales d ON a.dr_id = d.id
        LEFT JOIN impayes i ON i.agence_id = a.id AND i.mois = ? AND i.exercice = ?
    """
    params = [mois, exercice]
    if dr_filter:
        query += " WHERE d.code = ?"
        params.append(dr_filter)
    query += " ORDER BY d.code, a.nom"
    rows = db.execute(query, params).fetchall()
    db.close()

    lignes = []
    current_dr = None
    dr_actifs = 0
    dr_resilies = 0
    for r in rows:
        if current_dr and current_dr != r['dr_code']:
            lignes.append({"dr_code": current_dr, "agence": "Sous-total " + current_dr,
                           "actifs": dr_actifs, "resilies": dr_resilies, "is_subtotal": True})
            dr_actifs = 0
            dr_resilies = 0
        current_dr = r['dr_code']
        actifs = r['actifs'] or 0
        resilies = r['resilies'] or 0
        dr_actifs += actifs
        dr_resilies += resilies
        lignes.append({"dr_code": r['dr_code'], "agence": r['agence_nom'],
                       "actifs": actifs, "resilies": resilies, "is_subtotal": False})
    if current_dr:
        lignes.append({"dr_code": current_dr, "agence": "Sous-total " + current_dr,
                       "actifs": dr_actifs, "resilies": dr_resilies, "is_subtotal": True})

    return jsonify({"lignes": lignes})


# ─── Helpers Consultation : extraction des données du modèle ────────

def _extract_reporting(dr_code, d):
    """Extrait les champs reporting d'un résultat de calcul (calcul_cumul_dr/national)."""
    return {
        "dr_code": dr_code,
        # Volumes
        "total_volumes": d.get("total_volumes", 0),
        # Section (A) CA Vente Eau
        "total_ve": d.get("total_ve", 0),
        "total_ca_auto": d.get("total_ca_auto", 0),
        "total_ca_spec": d.get("total_ca_spec", 0),
        "total_locations": d.get("total_locations", 0),
        # Section (B) Travaux Remboursables
        "total_trvx_remb": d.get("total_trvx_remb", 0),
        "brchts_neufs": d.get("brchts_neufs", 0),
        "penalites": d.get("penalites", 0),
        "devis_ext": d.get("devis_ext", 0),
        "autres_trvx": d.get("autres_trvx", 0),
        "fraudes_trvx": d.get("fraudes_trvx", 0),
        # CA Global
        "ca_global": d.get("ca_global", 0),
        # Encaissements détaillés
        "enc_cciale": d.get("enc_cciale", 0),
        "enc_adm_enc": d.get("enc_adm_enc", 0),
        "enc_banques": d.get("enc_banques", 0),
        "enc_vte_eau": d.get("enc_vte_eau", 0),
        "enc_travaux": d.get("enc_travaux", 0),
        "enc_asc": d.get("enc_asc", 0),
        "total_encaissements": d.get("total_encaissements", 0),
        # Taux
        "taux_enc_ve": d.get("taux_enc_ve", 0),
        "taux_enc_trvx": d.get("taux_enc_trvx", 0),
        "taux_enc_global": d.get("taux_enc_global", 0),
        # Recouvrement EF
        "facturation_ef": d.get("facturation_ef", 0),
        "recouvrement_ef": d.get("recouvrement_ef", 0),
        "taux_recouvrement_ef": d.get("taux_recouvrement_ef", 0),
    }


def _extract_fiscal(dr_code, d):
    """Extrait les champs fiscaux d'un résultat de calcul."""
    ca_adm_ht = d.get("ca_adm_ht", 0)
    ca_hors_adm_ht = d.get("ca_hors_adm_ht", 0)
    ca_total_ttc = d.get("ca_total_ttc", 0)
    trvx_ttc = d.get("trvx_ttc", 0)
    return {
        "dr_code": dr_code,
        "tranche_sociale": d.get("tranche_sociale", 0),
        "ca_adm_ht": ca_adm_ht,
        "ca_hors_adm_ht": ca_hors_adm_ht,
        "ca_total_ttc": ca_total_ttc,
        "trvx_ttc": trvx_ttc,
        "total_ttc": ca_total_ttc + trvx_ttc,
    }


def _extract_budget(dr_code, d):
    """Extrait les données budget d'un résultat de calcul."""
    b = d.get("budget", {}) or {}
    vals = {
        "vente_eau": b.get("vente_eau", 0),
        "trvx_remb": b.get("trvx_remb", 0),
        "recouvrement_impayes": b.get("recouvrement_impayes", 0),
        "fraude": b.get("fraude", 0),
        "sinistre": b.get("sinistre", 0),
        "penalites": b.get("penalites", 0),
        "locations": b.get("locations", 0),
        "branchements": b.get("branchements", 0),
    }
    vals["dr_code"] = dr_code
    vals["total"] = sum(v for k, v in vals.items() if k != "dr_code")
    return vals


# ─── API Consultation : Reporting ────────────────────────────────

@app.route('/api/consultation/reporting')
def api_consultation_reporting():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 3, type=int)
    exercice = request.args.get('exercice', 2026, type=int)

    # Calcul DR par DR via le moteur de calcul (fidèle au modèle)
    drs_data = []
    for dr_code in STRUCTURE_CW:
        d = calcul_cumul_dr(dr_code, mois_debut, mois_fin, exercice)
        drs_data.append(_extract_reporting(dr_code, d))

    # Total national
    nat = calcul_cumul_national(mois_debut, mois_fin, exercice)
    national = _extract_reporting("ENSEMBLE CAMWATER", nat)

    # Impayés par DR (snapshot dernier mois de la période)
    db = get_db()
    imp_rows = db.execute("""
        SELECT d.code as dr_code,
               COALESCE(SUM(i.particuliers_actifs),0)+COALESCE(SUM(i.gco_actifs),0)+
               COALESCE(SUM(i.bf_actifs),0)+COALESCE(SUM(i.bfc_actifs),0) as actifs,
               COALESCE(SUM(i.particuliers_resilies),0)+COALESCE(SUM(i.gco_resilies),0)+
               COALESCE(SUM(i.bf_resilies),0)+COALESCE(SUM(i.bfc_resilies),0)+
               COALESCE(SUM(i.gestion_manuelle),0) as resilies
        FROM impayes i
        JOIN agences a ON i.agence_id = a.id
        JOIN directions_regionales d ON a.dr_id = d.id
        WHERE i.mois = ? AND i.exercice = ?
        GROUP BY d.code
    """, (mois_fin, exercice)).fetchall()
    db.close()
    imp_by_dr = {r['dr_code']: {"actifs": r['actifs'] or 0, "resilies": r['resilies'] or 0} for r in imp_rows}

    total_imp_actifs = 0
    total_imp_resilies = 0
    for dr_item in drs_data:
        imp = imp_by_dr.get(dr_item["dr_code"], {"actifs": 0, "resilies": 0})
        dr_item["impayes_actifs"] = imp["actifs"]
        dr_item["impayes_resilies"] = imp["resilies"]
        dr_item["impayes_total"] = imp["actifs"] + imp["resilies"]
        total_imp_actifs += imp["actifs"]
        total_imp_resilies += imp["resilies"]
    national["impayes_actifs"] = total_imp_actifs
    national["impayes_resilies"] = total_imp_resilies
    national["impayes_total"] = total_imp_actifs + total_imp_resilies

    return jsonify({
        "drs": drs_data,
        "national": national,
        "periode": {"mois_debut": mois_debut, "mois_fin": mois_fin, "exercice": exercice},
    })


# ─── API Consultation : Fiscal ───────────────────────────────────

@app.route('/api/consultation/fiscal')
def api_consultation_fiscal():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 3, type=int)
    exercice = request.args.get('exercice', 2026, type=int)

    drs_data = []
    for dr_code in STRUCTURE_CW:
        d = calcul_cumul_dr(dr_code, mois_debut, mois_fin, exercice)
        drs_data.append(_extract_fiscal(dr_code, d))

    nat = calcul_cumul_national(mois_debut, mois_fin, exercice)
    national = _extract_fiscal("ENSEMBLE CAMWATER", nat)

    return jsonify({
        "drs": drs_data,
        "national": national,
        "periode": {"mois_debut": mois_debut, "mois_fin": mois_fin, "exercice": exercice},
    })


# ─── API Consultation : Budget ───────────────────────────────────

@app.route('/api/consultation/budget')
def api_consultation_budget():
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 3, type=int)
    exercice = request.args.get('exercice', 2026, type=int)

    drs_data = []
    for dr_code in STRUCTURE_CW:
        d = calcul_cumul_dr(dr_code, mois_debut, mois_fin, exercice)
        drs_data.append(_extract_budget(dr_code, d))

    nat = calcul_cumul_national(mois_debut, mois_fin, exercice)
    national = _extract_budget("ENSEMBLE CAMWATER", nat)

    # Objectifs (nationaux + DR) proratisés sur la période
    nb_mois = mois_fin - mois_debut + 1
    db = get_db()
    nat_obj_rows = db.execute("""
        SELECT rubrique, COALESCE(SUM(montant),0) as total FROM objectifs
        WHERE exercice=? AND scope_type='national' GROUP BY rubrique
    """, (exercice,)).fetchall()
    objectifs_national = {r['rubrique']: (r['total'] or 0) * nb_mois / 12 for r in nat_obj_rows}

    objectifs_dr = {}
    for dr_code in STRUCTURE_CW:
        dr_row = db.execute("SELECT id FROM directions_regionales WHERE code=?", (dr_code,)).fetchone()
        if not dr_row:
            objectifs_dr[dr_code] = {}
            continue
        rows = db.execute("""
            SELECT rubrique, COALESCE(SUM(montant),0) as total FROM objectifs
            WHERE exercice=? AND scope_type='dr' AND scope_id=? GROUP BY rubrique
        """, (exercice, dr_row['id'])).fetchall()
        objectifs_dr[dr_code] = {r['rubrique']: (r['total'] or 0) * nb_mois / 12 for r in rows}
    db.close()

    return jsonify({
        "drs": drs_data,
        "national": national,
        "objectifs_dr": objectifs_dr,
        "objectifs_national": objectifs_national,
        "nb_mois": nb_mois,
        "periode": {"mois_debut": mois_debut, "mois_fin": mois_fin, "exercice": exercice},
    })


# ─── Page Paiements Électroniques ─────────────────────────────────

@app.route('/paiements-electroniques')
def paiements_electroniques_page():
    operateur = session.get('operateur', {})
    return render_template('paiements_electroniques.html', mois=MOIS,
                           drs=list(STRUCTURE_CW.keys()),
                           operateur=operateur)


# ─── API Paiements Électroniques : Upload CSV ────────────────────

@app.route('/api/paiements-elec/upload-csv', methods=['POST'])
def api_upload_paiements_elec_csv():
    """Upload un fichier CSV HIST de paiements électroniques.
    Format HIST : 13 colonnes, séparateur point-virgule, PAS d'en-tête.
    Colonnes (index 0-12) :
      0=?, 1=date, 2=code_operateur (UTOM/UTMM/UTEU/UTEE/UTYM/CAMO),
      3=?, 4=PL (point de livraison), 5=?, 6=nom_client,
      7=montant_principal (Vente Eau), 8=pénalité, 9=?, 10=?, 11=?, 12=?
    Le PL permet de déduire la DR (2 premiers chiffres) et le Centre (4 premiers chiffres).
    """
    import csv
    import re

    if 'file' not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.csv', '.txt')):
        return jsonify({"error": "Format CSV requis (.csv)"}), 400
    exercice = request.form.get('exercice', 2026, type=int)

    # ═══ Mapping code opérateur → nom opérateur ═══
    OPERATOR_MAP = {
        'UTOM': 'ORANGE MONEY',
        'UTMM': 'MOBILE MONEY',
        'UTEU': 'EXPRESS UNION',
        'UTEE': 'EXPRESS EXCHANGE',
        'UTYM': 'YOOMEE MONEY',
        'CAMO': 'CAMPOST MONEY',
    }

    # ═══ Mapping code DR (2 premiers chiffres du PL) → code DR ═══
    DR_MAP = {
        '01': 'DRDA', '02': 'DRL', '03': 'DRYA',
        '04': 'DRA',   # Adamaoua (+ quelques cas DRC/DRN)
        '05': 'DRC',   # Centre (+ quelques cas DRA/DRS)
        '06': 'DREN', '07': 'DRO',
        '08': 'DRSO', '09': 'DRNO',
        '10': 'DRE', '11': 'DRN', '12': 'DRS',
    }

    # ═══ Mapping code agence (4 premiers chiffres du PL) → centre ═══
    # Extrait des formules IF de l'Excel "Enc Elec souche de travail"
    CENTRE_MAP = {
        # DRDA (01xx)
        '0101': 'Koumassi', '0102': 'Deido', '0103': 'Bassa',
        '0104': 'Bonamoussadi', '0105': 'Bonaberi', '0106': 'Nyalla',
        # DRL (02xx)
        '0201': 'Nkongsamba', '0202': 'Edea', '0203': 'Loum',
        '0204': 'Mbanga', '0205': 'Njombé', '0206': 'Manjo',
        '0207': 'Penja', '0208': 'Yabassi', '0209': 'Ngambè',
        '0210': 'Pouma', '0211': 'Dizangué', '0212': 'Dibang',
        # DRYA (03xx)
        '0301': 'Ekounou', '0302': 'Tsinga', '0303': 'Mvog Ada',
        '0304': 'Etoudi', '0305': 'Obili', '0306': 'Soa',
        '0307': 'Mbankomo', '0308': 'Olembé', '0309': 'Biteng',
        '0310': 'Messamendongo',
        # DRA (04xx) — Adamaoua
        '0401': 'Ngaoundéré', '0402': 'Meiganga', '0403': 'Tibati',
        '0404': 'Banyo', '0406': 'Mbé',
        # DRC (05xx) — Centre
        '0501': 'Mbalmayo', '0502': 'Bafia', '0503': 'Mfou',
        '0504': 'Obala', '0505': 'Nanga Eboko', '0506': 'Eseka',
        '0507': 'Akonolinga', '0508': 'Monatélé', '0509': 'Mbandjock',
        '0510': 'Okola', '0511': 'Ndikinimeki', '0512': 'Akono',
        '0513': 'Ayos', '0514': 'Batchenga', '0515': 'Bikok',
        '0516': 'Bokito', '0517': 'Evodoula', '0518': 'Makak',
        '0519': 'Makenene', '0520': 'Matomb', '0521': 'Ngoumou',
        '0522': 'Ombessa', '0523': "Sa'a", '0524': 'Yoko',
        # DREN (06xx) — Extrême-Nord
        '0601': 'Maroua', '0602': 'Mokolo', '0603': 'Yagoua',
        '0604': 'Kousseri', '0605': 'Mora', '0606': 'Kaélé',
        '0607': 'Doukoula', '0608': 'Makari', '0609': 'Kolofata',
        '0610': 'Maga', '0611': 'Koza', '0612': 'Bogo',
        '0613': 'Tokombéré',
        # DRO (07xx) — Ouest
        '0701': 'Bafoussam', '0702': 'Bafang', '0703': 'Bandjoun',
        '0704': 'Bangangté', '0705': 'Tonga', '0706': 'Bazou',
        '0707': 'Dschang', '0708': 'Foumban', '0709': 'Foumbot',
        '0710': 'Mbouda', '0711': 'Melong', '0712': 'Kekem',
        '0713': 'Bankim', '0714': 'Bamendjou', '0715': 'Baham',
        '0716': 'Bansoa', '0717': 'Bana', '0718': 'Bangoua',
        '0719': 'Bayangam', '0720': 'Bangou', '0721': 'Batié',
        # DRSO (08xx) — Sud-Ouest
        '0801': 'Limbé', '0802': 'Buea', '0803': 'Kumba',
        '0804': 'Tiko', '0805': 'Mamfé', '0806': 'Mundemba',
        '0807': 'Muyuka', '0808': 'Nguti',
        # DRNO (09xx) — Nord-Ouest
        '0901': 'Bamenda', '0902': 'Batibo', '0903': 'Fundong',
        '0904': 'Jakiri', '0905': 'Njikijem', '0906': 'Mbengwi',
        '0907': 'Ndop', '0908': 'Njinikom', '0909': 'Nkambè',
        '0910': 'Wum',
        # DRE (10xx) — Est
        '1001': 'Bertoua', '1002': 'Abong Mbang', '1003': 'Batouri',
        '1004': 'Belabo', '1005': 'Yokadouma', '1006': 'Dimako',
        '1007': 'Lomié',
        # DRN (11xx) — Nord
        '1101': 'Garoua', '1102': 'Guider', '1103': 'Figuil',
        '1104': 'Mayo Oulo', '1105': 'Pitoa',
        # DRS (12xx) — Sud
        '1201': 'Ebolowa', '1202': 'Kribi', '1203': 'Sangmélima',
        '1204': 'Ambam', '1205': 'Campo', '1206': 'Djoum',
        '1207': 'Meyomessala', '1208': 'Zoétélé',
    }

    # Mapping agence connue → agence_id (pré-chargé)
    all_agences_map = {}
    for dr_c, ag_list in STRUCTURE_CW.items():
        for ag in ag_list:
            all_agences_map[ag.lower()] = (dr_c, ag)

    try:
        raw = f.read()
        text = None
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                text = raw.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            return jsonify({"error": "Impossible de décoder le fichier"}), 400

        # Déterminer si le fichier a un en-tête ou non
        lines = text.strip().split('\n')
        if not lines:
            return jsonify({"error": "Fichier vide"}), 400

        # Auto-détection séparateur
        first_line = lines[0]
        if first_line.count(';') >= 5:
            sep = ';'
        elif first_line.count('\t') >= 5:
            sep = '\t'
        elif first_line.count(',') >= 5:
            sep = ','
        else:
            sep = ';'  # Défaut pour les fichiers HIST

        # Vérifier si la première ligne est un en-tête (contient du texte non-numérique)
        first_cols = first_line.split(sep)
        has_header = False
        if len(first_cols) >= 5:
            # Si les colonnes 7 et 8 ne sont pas numériques → c'est un en-tête
            try:
                if len(first_cols) > 7:
                    float(first_cols[7].strip().replace(' ', '').replace(',', '.'))
                has_header = False
            except (ValueError, IndexError):
                has_header = True

        reader = csv.reader(lines, delimiter=sep)
        if has_header:
            next(reader, None)  # Sauter l'en-tête

        db = get_db()
        db.execute("DELETE FROM paiements_elec_csv WHERE exercice=?", (exercice,))

        # Pré-charger les IDs des agences
        agence_id_cache = {}
        for arow in db.execute("SELECT id, nom FROM agences").fetchall():
            agence_id_cache[arow['nom'].lower()] = arow['id']

        count = 0
        errors = 0
        operateurs_set = set()
        batch = []

        for row in reader:
            if not row or len(row) < 9:
                continue

            # Extraire les colonnes selon le format HIST
            # Col 2 (index 2) = code opérateur
            # Col 4 (index 4) = PL (point de livraison)
            # Col 6 (index 6) = nom client
            # Col 7 (index 7) = montant principal (Vente Eau)
            # Col 8 (index 8) = pénalité
            code_op = row[2].strip().upper() if len(row) > 2 else ''
            pl_raw = row[4].strip() if len(row) > 4 else ''
            nom_client = row[6].strip() if len(row) > 6 else ''
            date_trans = row[1].strip() if len(row) > 1 else ''

            # Montant principal (Vente Eau)
            vente_eau = 0
            try:
                ve_str = row[7].strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
                vente_eau = float(ve_str) if ve_str else 0
            except (ValueError, TypeError, IndexError):
                pass

            # Pénalité
            penalite = 0
            try:
                pen_str = row[8].strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
                penalite = float(pen_str) if pen_str else 0
            except (ValueError, TypeError, IndexError):
                pass

            montant_total = vente_eau + penalite
            if montant_total == 0:
                continue

            # Résoudre l'opérateur
            operateur_nom = OPERATOR_MAP.get(code_op, code_op or 'Non précisé')
            operateurs_set.add(operateur_nom)

            # Extraire le PL : supprimer espaces et caractères non numériques en début
            pl_clean = re.sub(r'[^0-9]', '', pl_raw)[:10]  # Garder max 10 chiffres

            # Code DR (2 premiers chiffres du PL)
            dr_code = ''
            code_agence_4 = ''
            centre = ''
            agence_id = None

            if len(pl_clean) >= 2:
                prefix_dr = pl_clean[:2]
                dr_code = DR_MAP.get(prefix_dr, '')

            if len(pl_clean) >= 4:
                code_agence_4 = pl_clean[:4]
                centre = CENTRE_MAP.get(code_agence_4, '')

                # Résoudre agence_id
                if centre:
                    agence_id = agence_id_cache.get(centre.lower())
                    # Si le centre existe, forcer la DR depuis la structure
                    match = all_agences_map.get(centre.lower())
                    if match:
                        dr_code = match[0]

            # Extraire le mois depuis la date (format DD/MM/YYYY ou DD-MM-YYYY ou YYYYMMDD)
            mois_val = None
            if date_trans:
                m_fr = re.match(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', date_trans)
                m_iso = re.match(r'(\d{4})[/\-.]?(\d{2})[/\-.]?(\d{2})', date_trans)
                if m_fr:
                    mois_val = int(m_fr.group(2))
                elif m_iso:
                    mois_val = int(m_iso.group(2))
                if mois_val and (mois_val < 1 or mois_val > 12):
                    mois_val = None

            batch.append((exercice, date_trans, operateur_nom, pl_raw, nom_client,
                          montant_total, centre, dr_code, agence_id, '', '',
                          mois_val, '', f.filename,
                          vente_eau, penalite, pl_clean, code_agence_4))
            count += 1

            # Insertion par lots de 2000
            if len(batch) >= 2000:
                db.executemany("""INSERT INTO paiements_elec_csv
                    (exercice, date_transaction, operateur, reference_client, nom_client,
                     montant, centre, dr_code, agence_id, statut, comptabilise,
                     mois, ligne_brute, fichier_source,
                     vente_eau, penalite, pl_code, code_agence)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", batch)
                batch = []

        if batch:
            db.executemany("""INSERT INTO paiements_elec_csv
                (exercice, date_transaction, operateur, reference_client, nom_client,
                 montant, centre, dr_code, agence_id, statut, comptabilise,
                 mois, ligne_brute, fichier_source,
                 vente_eau, penalite, pl_code, code_agence)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", batch)

        db.commit()
        db.close()

        return jsonify({
            "status": "ok",
            "lignes_importees": count,
            "operateurs_detectes": sorted(list(operateurs_set)),
            "nb_operateurs": len(operateurs_set),
            "format": "HIST CSV (13 colonnes, point-virgule)",
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── API Paiements Électroniques : Résultats agrégés ─────────────

@app.route('/api/paiements-elec/resultats')
def api_paiements_elec_resultats():
    """Résultats agrégés format Enc centres : par DR > Centre > Opérateur.
    Chaque opérateur a une ventilation VE (Vente Eau) + Pénalité + Total."""
    exercice = request.args.get('exercice', 2026, type=int)
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 12, type=int)
    dr_filter = request.args.get('dr', '')

    db = get_db()

    # Vérifier s'il y a des données validées
    valid_row = db.execute(
        "SELECT date_validation FROM paiements_elec_definitif WHERE exercice=? LIMIT 1",
        (exercice,)).fetchone()
    est_valide = valid_row is not None

    # Filtrage par mois
    where_mois = "AND (mois BETWEEN ? AND ? OR mois IS NULL)"
    params_base = [exercice, mois_debut, mois_fin]

    if dr_filter:
        where_dr = "AND dr_code=?"
        params_base.append(dr_filter)
    else:
        where_dr = ""

    # Liste des opérateurs distincts
    ops_rows = db.execute(f"""
        SELECT DISTINCT operateur FROM paiements_elec_csv
        WHERE exercice=? {where_mois} {where_dr}
        ORDER BY operateur
    """, params_base).fetchall()
    operateurs = [r['operateur'] for r in ops_rows]

    # Données agrégées par DR > Centre > Opérateur avec VE/Pénalité
    agg_rows = db.execute(f"""
        SELECT dr_code, centre, operateur,
               COALESCE(SUM(vente_eau), 0) as total_ve,
               COALESCE(SUM(penalite), 0) as total_pen,
               SUM(montant) as total,
               COUNT(*) as nb
        FROM paiements_elec_csv
        WHERE exercice=? {where_mois} {where_dr}
        GROUP BY dr_code, centre, operateur
        ORDER BY dr_code, centre, operateur
    """, params_base).fetchall()

    # Construire la structure Enc centres
    from collections import OrderedDict
    centres_data = OrderedDict()
    for r in agg_rows:
        dr = r['dr_code'] or 'N/A'
        centre = r['centre'] or 'Non précisé'
        key = (dr, centre)
        if key not in centres_data:
            centres_data[key] = {'par_operateur': {}, 'total': 0, 'total_ve': 0, 'total_pen': 0, 'nb': 0}
        centres_data[key]['par_operateur'][r['operateur']] = {
            've': r['total_ve'], 'pen': r['total_pen'], 'total': r['total']
        }
        centres_data[key]['total'] += r['total']
        centres_data[key]['total_ve'] += r['total_ve']
        centres_data[key]['total_pen'] += r['total_pen']
        centres_data[key]['nb'] += r['nb']

    # Construire les lignes avec sous-totaux par DR
    lignes = []
    current_dr = None
    dr_subtotals = {}  # op → {ve, pen, total}
    dr_total = 0

    def make_subtotal_line(dr_code, subtotals, total_val):
        return {
            'dr_code': dr_code, 'centre': 'Sous-total ' + dr_code,
            'par_operateur': subtotals, 'total': total_val,
            'is_subtotal': True
        }

    for (dr, centre), cdata in centres_data.items():
        if current_dr and dr != current_dr:
            lignes.append(make_subtotal_line(current_dr, dr_subtotals, dr_total))
            dr_subtotals = {}
            dr_total = 0

        current_dr = dr
        lignes.append({
            'dr_code': dr, 'centre': centre,
            'par_operateur': cdata['par_operateur'],
            'total': cdata['total'], 'is_subtotal': False
        })
        for op, vals in cdata['par_operateur'].items():
            if op not in dr_subtotals:
                dr_subtotals[op] = {'ve': 0, 'pen': 0, 'total': 0}
            dr_subtotals[op]['ve'] += vals['ve']
            dr_subtotals[op]['pen'] += vals['pen']
            dr_subtotals[op]['total'] += vals['total']
        dr_total += cdata['total']

    # Dernier sous-total
    if current_dr and dr_subtotals:
        lignes.append(make_subtotal_line(current_dr, dr_subtotals, dr_total))

    # Totaux par opérateur
    par_op_total = {}
    montant_total = 0
    total_ve_global = 0
    total_pen_global = 0
    nb_total = 0
    for cdata in centres_data.values():
        for op, vals in cdata['par_operateur'].items():
            if op not in par_op_total:
                par_op_total[op] = {'ve': 0, 'pen': 0, 'total': 0}
            par_op_total[op]['ve'] += vals['ve']
            par_op_total[op]['pen'] += vals['pen']
            par_op_total[op]['total'] += vals['total']
        montant_total += cdata['total']
        total_ve_global += cdata['total_ve']
        total_pen_global += cdata['total_pen']
        nb_total += cdata['nb']

    nb_centres = len(set(c for _, c in centres_data.keys()))

    db.close()

    return jsonify({
        "operateurs": operateurs,
        "lignes": lignes,
        "par_operateur": par_op_total,
        "montant_total": montant_total,
        "total_ve": total_ve_global,
        "total_pen": total_pen_global,
        "nb_transactions": nb_total,
        "nb_centres": nb_centres,
        "est_valide": est_valide,
        "date_validation": valid_row['date_validation'] if valid_row else None,
    })


# ─── API Paiements Électroniques : Validation définitive ─────────

@app.route('/api/paiements-elec/valider', methods=['POST'])
def api_paiements_elec_valider():
    """Enregistrement définitif : copier les données agrégées dans la table definitif."""
    data = request.json
    exercice = data.get('exercice', 2026)
    mois_debut = data.get('mois_debut', 1)
    mois_fin = data.get('mois_fin', 12)
    operateur_session = session.get('operateur', {})

    db = get_db()

    # Vérifier qu'il y a des données
    count = db.execute("SELECT COUNT(*) as c FROM paiements_elec_csv WHERE exercice=?",
                       (exercice,)).fetchone()['c']
    if count == 0:
        db.close()
        return jsonify({"error": "Aucune donnée à valider"}), 400

    # Supprimer les anciennes validations pour cette période
    db.execute("""DELETE FROM paiements_elec_definitif
                  WHERE exercice=? AND mois_debut=? AND mois_fin=?""",
               (exercice, mois_debut, mois_fin))

    # Agréger et insérer
    agg_rows = db.execute("""
        SELECT dr_code, agence_id, centre as agence_nom, operateur,
               SUM(montant) as montant_total, COUNT(*) as nb_transactions
        FROM paiements_elec_csv
        WHERE exercice=? AND (mois BETWEEN ? AND ? OR mois IS NULL)
        GROUP BY dr_code, agence_id, centre, operateur
    """, (exercice, mois_debut, mois_fin)).fetchall()

    for r in agg_rows:
        db.execute("""INSERT INTO paiements_elec_definitif
            (exercice, mois_debut, mois_fin, dr_code, agence_id, agence_nom,
             operateur, montant_total, nb_transactions, operateur_nom, operateur_matricule)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (exercice, mois_debut, mois_fin, r['dr_code'], r['agence_id'],
             r['agence_nom'], r['operateur'], r['montant_total'], r['nb_transactions'],
             operateur_session.get('nom', ''), operateur_session.get('matricule', '')))

    db.commit()
    db.close()

    return jsonify({"status": "ok", "message": f"Données validées définitivement ({len(agg_rows)} lignes agrégées)"})


# ─── API Paiements Électroniques : Suppression ───────────────────

@app.route('/api/paiements-elec/delete-csv', methods=['DELETE'])
def api_delete_paiements_elec_csv():
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    db.execute("DELETE FROM paiements_elec_csv WHERE exercice=?", (exercice,))
    db.execute("DELETE FROM paiements_elec_definitif WHERE exercice=?", (exercice,))
    db.commit()
    db.close()
    return jsonify({"status": "ok", "message": f"Données paiements électroniques supprimées pour {exercice}"})


# ─── API Paiements Électroniques : Export Excel ──────────────────

@app.route('/api/paiements-elec/export-excel')
def api_export_paiements_elec_excel():
    """Export au format Excel - style Enc centres avec VE / Pénalité / Total par opérateur."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    exercice = request.args.get('exercice', 2026, type=int)
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 12, type=int)
    dr_filter = request.args.get('dr', '')

    db = get_db()

    where_mois = "AND (mois BETWEEN ? AND ? OR mois IS NULL)"
    params = [exercice, mois_debut, mois_fin]
    where_dr = ""
    if dr_filter:
        where_dr = "AND dr_code=?"
        params.append(dr_filter)

    # Opérateurs
    ops = [r['operateur'] for r in db.execute(f"""
        SELECT DISTINCT operateur FROM paiements_elec_csv
        WHERE exercice=? {where_mois} {where_dr} ORDER BY operateur
    """, params).fetchall()]

    # Données avec VE/Pénalité
    rows = db.execute(f"""
        SELECT dr_code, centre, operateur,
               COALESCE(SUM(vente_eau), 0) as total_ve,
               COALESCE(SUM(penalite), 0) as total_pen,
               SUM(montant) as total
        FROM paiements_elec_csv
        WHERE exercice=? {where_mois} {where_dr}
        GROUP BY dr_code, centre, operateur ORDER BY dr_code, centre
    """, params).fetchall()

    db.close()

    # Pivot : chaque cellule = {ve, pen, total}
    from collections import OrderedDict
    pivot = OrderedDict()
    for r in rows:
        key = (r['dr_code'] or 'N/A', r['centre'] or 'Non précisé')
        if key not in pivot:
            pivot[key] = {}
        pivot[key][r['operateur']] = {
            've': r['total_ve'], 'pen': r['total_pen'], 'total': r['total']
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enc Centres"

    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="003D82", end_color="003D82", fill_type="solid")
    ve_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    pen_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    total_fill = PatternFill(start_color="002B5C", end_color="002B5C", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # En-tête ligne 1 : DR | Centre | [Op1 (merge 3 cols)] ... | TOTAL
    col = 1
    for h in ['DR', 'Centre']:
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    for op in ops:
        cell = ws.cell(row=1, column=col, value=op)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        # Sous-en-têtes VE | Pén | Total
        for si, (sub_label, sub_fill) in enumerate([('VE', ve_fill), ('Pén.', pen_fill), ('Total', hdr_fill)]):
            sc = ws.cell(row=2, column=col + si, value=sub_label)
            sc.font = Font(bold=True, size=9, color="FFFFFF")
            sc.fill = sub_fill
            sc.alignment = Alignment(horizontal='center')
            sc.border = thin_border
        col += 3

    # Colonne TOTAL GENERAL
    cell = ws.cell(row=1, column=col, value='TOTAL')
    cell.font = hdr_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    row_num = 3
    current_dr = None
    dr_sums = {}   # op → {ve, pen, total}
    grand_sums = {}

    def write_subtotal(ws, row_n, label, sums, ops_list, fill):
        ws.cell(row=row_n, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_n, column=1).border = thin_border
        ws.cell(row=row_n, column=2, value=f"Sous-total {label}").font = Font(bold=True)
        ws.cell(row=row_n, column=2).border = thin_border
        c = 3
        st_total = 0
        for op in ops_list:
            vals = sums.get(op, {'ve': 0, 'pen': 0, 'total': 0})
            for v in [vals['ve'], vals['pen'], vals['total']]:
                cl = ws.cell(row=row_n, column=c, value=v)
                cl.number_format = '#,##0'
                cl.font = Font(bold=True)
                cl.fill = fill
                cl.border = thin_border
                c += 1
            st_total += vals['total']
        cl = ws.cell(row=row_n, column=c, value=st_total)
        cl.number_format = '#,##0'
        cl.font = Font(bold=True)
        cl.fill = fill
        cl.border = thin_border
        return row_n + 1

    for (dr, centre), op_data in pivot.items():
        if current_dr and dr != current_dr:
            row_num = write_subtotal(ws, row_num, current_dr, dr_sums, ops, subtotal_fill)
            dr_sums = {}

        current_dr = dr
        ws.cell(row=row_num, column=1, value=dr).border = thin_border
        ws.cell(row=row_num, column=2, value=centre).border = thin_border
        c = 3
        line_total = 0
        for op in ops:
            vals = op_data.get(op, {'ve': 0, 'pen': 0, 'total': 0})
            if isinstance(vals, (int, float)):
                vals = {'ve': 0, 'pen': 0, 'total': vals}
            for v in [vals['ve'], vals['pen'], vals['total']]:
                cl = ws.cell(row=row_num, column=c, value=v)
                cl.number_format = '#,##0'
                cl.border = thin_border
                c += 1
            # Accumuler
            if op not in dr_sums:
                dr_sums[op] = {'ve': 0, 'pen': 0, 'total': 0}
            if op not in grand_sums:
                grand_sums[op] = {'ve': 0, 'pen': 0, 'total': 0}
            dr_sums[op]['ve'] += vals['ve']
            dr_sums[op]['pen'] += vals['pen']
            dr_sums[op]['total'] += vals['total']
            grand_sums[op]['ve'] += vals['ve']
            grand_sums[op]['pen'] += vals['pen']
            grand_sums[op]['total'] += vals['total']
            line_total += vals['total']
        cl = ws.cell(row=row_num, column=c, value=line_total)
        cl.number_format = '#,##0'
        cl.font = Font(bold=True)
        cl.border = thin_border
        row_num += 1

    # Dernier sous-total
    if current_dr and dr_sums:
        row_num = write_subtotal(ws, row_num, current_dr, dr_sums, ops, subtotal_fill)

    # Grand total
    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = thin_border
    ws.cell(row=row_num, column=2, value="ENSEMBLE CAMWATER").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row_num, column=2).fill = total_fill
    ws.cell(row=row_num, column=2).border = thin_border
    c = 3
    g_total = 0
    for op in ops:
        vals = grand_sums.get(op, {'ve': 0, 'pen': 0, 'total': 0})
        for v in [vals['ve'], vals['pen'], vals['total']]:
            cl = ws.cell(row=row_num, column=c, value=v)
            cl.number_format = '#,##0'
            cl.font = Font(bold=True, color="FFFFFF")
            cl.fill = total_fill
            cl.border = thin_border
            c += 1
        g_total += vals['total']
    cl = ws.cell(row=row_num, column=c, value=g_total)
    cl.number_format = '#,##0'
    cl.font = Font(bold=True, color="FFFFFF", size=11)
    cl.fill = total_fill
    cl.border = thin_border

    # Auto-width
    total_cols = 2 + len(ops) * 3 + 1
    for ci in range(1, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    mois_label = f"{MOIS[mois_debut-1]}-{MOIS[mois_fin-1]}" if mois_debut != mois_fin else MOIS[mois_debut-1]
    filename = f"Enc_Centres_PE_{mois_label}_{exercice}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ─── API Dashboard : Opérateurs Paiements Électroniques ──────────

@app.route('/api/paiements-elec/dashboard-operators')
def api_paiements_elec_dashboard_operators():
    """Synthèse des paiements électroniques par opérateur pour le tableau de bord."""
    exercice = request.args.get('exercice', 2026, type=int)
    mois_debut = request.args.get('mois_debut', 1, type=int)
    mois_fin = request.args.get('mois_fin', 12, type=int)

    db = get_db()

    # Chercher d'abord dans les données validées (definitif)
    def_rows = db.execute("""
        SELECT operateur, SUM(montant_total) as total, SUM(nb_transactions) as nb
        FROM paiements_elec_definitif
        WHERE exercice=? AND mois_debut<=? AND mois_fin>=?
        GROUP BY operateur ORDER BY total DESC
    """, (exercice, mois_fin, mois_debut)).fetchall()

    if def_rows:
        source = 'definitif'
        rows = def_rows
    else:
        # Sinon utiliser les données brutes CSV
        rows = db.execute("""
            SELECT operateur, SUM(montant) as total, COUNT(*) as nb
            FROM paiements_elec_csv
            WHERE exercice=? AND (mois BETWEEN ? AND ? OR mois IS NULL)
            GROUP BY operateur ORDER BY total DESC
        """, (exercice, mois_debut, mois_fin)).fetchall()
        source = 'csv'

    db.close()

    grand_total = sum(r['total'] for r in rows)
    operators = []
    for r in rows:
        operators.append({
            "operateur": r['operateur'],
            "montant": r['total'],
            "nb_transactions": r['nb'],
            "taux": r['total'] / grand_total if grand_total > 0 else 0,
        })

    return jsonify({
        "operators": operators,
        "montant_total": grand_total,
        "source": source,
    })


# ════════════════════════════════════════════════════════════════════════════
# MODULE MONITORING — Alertes Automatisées
# ════════════════════════════════════════════════════════════════════════════

@app.route('/monitoring')
def monitoring_page():
    operateur = session.get('operateur', {})
    role = session.get('role') or 'anonyme'
    read_only = (role == 'direction')
    return render_template('monitoring.html',
                           mois=MOIS,
                           drs=list(STRUCTURE_CW.keys()),
                           operateur=operateur,
                           role=role,
                           read_only=read_only)


@app.route('/api/monitoring/alertes')
def api_monitoring_alertes():
    mois = request.args.get('mois', 1, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    scope = request.args.get('scope', 'national')
    scope_id = request.args.get('scope_id', None)
    try:
        data = generer_alertes(mois, exercice, scope, scope_id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e), 'alertes': [], 'nb_critiques': 0,
                        'nb_warnings': 0, 'total': 0}), 500


@app.route('/api/monitoring/indicateurs')
def api_monitoring_indicateurs():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', 1, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    if not agence_id:
        return jsonify({'error': 'agence_id requis'}), 400
    data = indicateurs_agence(agence_id, mois, exercice)
    return jsonify(data)


@app.route('/api/monitoring/synthese-dr')
def api_monitoring_synthese_dr():
    mois = request.args.get('mois', 1, type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    data = synthese_par_dr(mois, exercice)
    return jsonify(data)


@app.route('/api/monitoring/params', methods=['GET'])
def api_monitoring_params_get():
    exercice = request.args.get('exercice', 2026, type=int)
    params = get_params(exercice)
    # Retourner avec libellés
    db = get_db()
    rows = db.execute(
        "SELECT cle, valeur, libelle FROM monitoring_params WHERE exercice=?", (exercice,)
    ).fetchall()
    db.close()
    return jsonify({
        'params': params,
        'details': [{'cle': r['cle'], 'valeur': r['valeur'], 'libelle': r['libelle']} for r in rows],
    })


@app.route('/api/monitoring/params', methods=['POST'])
def api_monitoring_params_post():
    data = request.json
    exercice = data.get('exercice', 2026)
    updates = data.get('params', {})
    if not updates:
        return jsonify({'error': 'Aucun paramètre fourni'}), 400
    save_params(updates, exercice)
    return jsonify({'status': 'ok'})


# ─── API Saisie Monitoring : Facturation Abonnés (Indicateur 1) ─────────────

@app.route('/api/facturation-abonnes', methods=['GET'])
def api_get_facturation_abonnes():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute(
        "SELECT abonnes_actifs, abonnes_factures FROM facturation_abonnes "
        "WHERE agence_id=? AND mois=? AND exercice=?",
        (agence_id, mois, exercice)
    ).fetchone()
    db.close()
    if row:
        return jsonify({'abonnes_actifs': row['abonnes_actifs'],
                        'abonnes_factures': row['abonnes_factures']})
    return jsonify({'abonnes_actifs': 0, 'abonnes_factures': 0})


@app.route('/api/facturation-abonnes', methods=['POST'])
def api_save_facturation_abonnes():
    data = request.json
    agence_id = data.get('agence_id')
    mois = data.get('mois')
    exercice = data.get('exercice', 2026)
    if not agence_id or not mois:
        return jsonify({'error': 'agence_id et mois requis'}), 400
    db = get_db()
    db.execute("""INSERT INTO facturation_abonnes (agence_id, mois, exercice, abonnes_actifs, abonnes_factures)
                  VALUES (?, ?, ?, ?, ?)
                  ON CONFLICT(agence_id, mois, exercice)
                  DO UPDATE SET abonnes_actifs=excluded.abonnes_actifs,
                                abonnes_factures=excluded.abonnes_factures""",
               (int(agence_id), int(mois), exercice,
                data.get('abonnes_actifs', 0) or 0,
                data.get('abonnes_factures', 0) or 0))
    db.commit()
    db.close()
    return jsonify({'status': 'ok'})


# ─── API Saisie Monitoring : Branchements Délais (Indicateur 4) ─────────────

@app.route('/api/branchements-delais', methods=['GET'])
def api_get_branchements_delais():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute(
        "SELECT total_devis_payes, dans_15j, delai_moyen_jours FROM branchements_delais "
        "WHERE agence_id=? AND mois=? AND exercice=?",
        (agence_id, mois, exercice)
    ).fetchone()
    db.close()
    if row:
        return jsonify({'total_devis_payes': row['total_devis_payes'],
                        'dans_15j': row['dans_15j'],
                        'delai_moyen_jours': row['delai_moyen_jours']})
    return jsonify({'total_devis_payes': 0, 'dans_15j': 0, 'delai_moyen_jours': None})


@app.route('/api/branchements-delais', methods=['POST'])
def api_save_branchements_delais():
    data = request.json
    agence_id = data.get('agence_id')
    mois = data.get('mois')
    exercice = data.get('exercice', 2026)
    if not agence_id or not mois:
        return jsonify({'error': 'agence_id et mois requis'}), 400
    db = get_db()
    db.execute("""INSERT INTO branchements_delais (agence_id, mois, exercice, total_devis_payes, dans_15j, delai_moyen_jours)
                  VALUES (?, ?, ?, ?, ?, ?)
                  ON CONFLICT(agence_id, mois, exercice)
                  DO UPDATE SET total_devis_payes=excluded.total_devis_payes,
                                dans_15j=excluded.dans_15j,
                                delai_moyen_jours=excluded.delai_moyen_jours""",
               (int(agence_id), int(mois), exercice,
                data.get('total_devis_payes', 0) or 0,
                data.get('dans_15j', 0) or 0,
                data.get('delai_moyen_jours') or None))
    db.commit()
    db.close()
    return jsonify({'status': 'ok'})


# ─── API Saisie Monitoring : Réémissions Factures (Indicateur 6) ────────────

@app.route('/api/reemissions', methods=['GET'])
def api_get_reemissions():
    agence_id = request.args.get('agence_id', type=int)
    mois = request.args.get('mois', type=int)
    exercice = request.args.get('exercice', 2026, type=int)
    db = get_db()
    row = db.execute(
        "SELECT nb_factures_emises, nb_reemissions FROM reemissions_factures "
        "WHERE agence_id=? AND mois=? AND exercice=?",
        (agence_id, mois, exercice)
    ).fetchone()
    db.close()
    if row:
        return jsonify({'nb_factures_emises': row['nb_factures_emises'],
                        'nb_reemissions': row['nb_reemissions']})
    return jsonify({'nb_factures_emises': 0, 'nb_reemissions': 0})


@app.route('/api/reemissions', methods=['POST'])
def api_save_reemissions():
    data = request.json
    agence_id = data.get('agence_id')
    mois = data.get('mois')
    exercice = data.get('exercice', 2026)
    if not agence_id or not mois:
        return jsonify({'error': 'agence_id et mois requis'}), 400
    db = get_db()
    db.execute("""INSERT INTO reemissions_factures (agence_id, mois, exercice, nb_factures_emises, nb_reemissions)
                  VALUES (?, ?, ?, ?, ?)
                  ON CONFLICT(agence_id, mois, exercice)
                  DO UPDATE SET nb_factures_emises=excluded.nb_factures_emises,
                                nb_reemissions=excluded.nb_reemissions""",
               (int(agence_id), int(mois), exercice,
                data.get('nb_factures_emises', 0) or 0,
                data.get('nb_reemissions', 0) or 0))
    db.commit()
    db.close()
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    # host='0.0.0.0' : écoute sur toutes les interfaces réseau.
    # Permet l'accès depuis d'autres machines du réseau local (agents de terrain
    # sur le même réseau) et depuis internet (via un tunnel type Cloudflare
    # Tunnel / ngrok / déploiement cloud). Voir README_DEPLOIEMENT.md.
    import os as _os
    host = _os.environ.get('CAMWATER_HOST', '0.0.0.0')
    port = int(_os.environ.get('CAMWATER_PORT', '5050'))
    debug = _os.environ.get('CAMWATER_DEBUG', '1') == '1'
    app.run(host=host, port=port, debug=debug)
