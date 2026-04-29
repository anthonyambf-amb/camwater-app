"""
Moteur d'import des fichiers Excel CAMWATER au format historique (matriciel)
============================================================================

Ce module gère les fichiers Excel HISTORIQUES de la CAMWATER au format réel :

  ▸ CA <DR> <année>.xls
       Une feuille par centre (Koumassi, Deido, ...) pour les volumes en m³,
       puis une feuille par centre suffixée 1 (Koumassi1, Deido1, ...) pour
       le chiffre d'affaires en FCFA. Lignes = rubriques, colonnes = mois.

  ▸ encais <DR> <année>.xls
       Une feuille par centre. Lignes = rubriques d'encaissement,
       colonnes = mois.

  ▸ SSP IMPAYÉS <année>.xls
       Une feuille `RECAP MENSUEL` (synthèse par DR), puis une feuille par
       DR avec des blocs mensuels (lignes = agences).

Distinct du module `import_historique.py` qui, lui, gère un template
standardisé avec les colonnes DR | Agence | Mois | Rubrique | Valeur.
Ici, on parse les fichiers métier réels — sans rien retoucher à la main.

Principes :
  • 100 % éphémère : tout en mémoire (io.BytesIO), aucune écriture disque.
  • Transactionnel : tout ou rien.
  • Idempotent : INSERT OR REPLACE — un re-import écrase proprement.
  • Tolérance accents/casse/typos via slug + alias.
  • Sélection mois : l'utilisateur choisit ce qui est écrit en base.
"""
from __future__ import annotations

import io
import re
import unicodedata
from typing import Any

import xlrd  # pour les .xls anciens
import openpyxl  # pour les .xlsx

from database import (get_db, STRUCTURE_CW, MOIS,
                       RUBRIQUES_ENC_CCIALE)

# ─── Normalisation ────────────────────────────────────────────────────────────

def _slug(s: Any) -> str:
    """Normalise pour comparaison tolérante : minuscules, sans accents,
    sans caractères spéciaux. Préserve les opérateurs ≤/≥/</> via 'lte'/'gte'/'lt'/'gt'
    afin de distinguer 'Particuliers <= 10m3' de 'Particuliers > 10m3'."""
    if s is None:
        return ''
    if not isinstance(s, str):
        s = str(s)
    # Préserver la sémantique des opérateurs avant la suppression des non-alphanum
    s = (s.replace('<=', ' lte ').replace('≤', ' lte ')
          .replace('>=', ' gte ').replace('≥', ' gte ')
          .replace('<', ' lt ').replace('>', ' gt '))
    nfd = unicodedata.normalize('NFD', s)
    ascii_ = nfd.encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '_', ascii_.lower()).strip('_')


def _num(v: Any) -> Any:
    """Convertit en float ou retourne None."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        cleaned = re.sub(r'[^\d.,-]', '', v.replace(',', '.'))
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None
    return None


# ─── Mapping mois → numéro ────────────────────────────────────────────────────

# Header des colonnes dans les fichiers : « JANVIER », « JANV. », « JANV »...
_MOIS_HEADER_MAP = {
    'janvier': 1, 'janv': 1, 'jan': 1,
    'fevrier': 2, 'fev': 2, 'feb': 2, 'fevr': 2,
    'mars': 3, 'mar': 3,
    'avril': 4, 'avr': 4, 'apr': 4,
    'mai': 5, 'may': 5,
    'juin': 6, 'jun': 6,
    'juillet': 7, 'juil': 7, 'jul': 7,
    'aout': 8, 'aou': 8, 'aug': 8,
    'septembre': 9, 'sept': 9, 'sep': 9,
    'octobre': 10, 'oct': 10,
    'novembre': 11, 'nov': 11,
    'decembre': 12, 'dec': 12,
}


def _parse_mois_header(label: Any) -> int | None:
    s = _slug(label)
    if not s:
        return None
    if s in _MOIS_HEADER_MAP:
        return _MOIS_HEADER_MAP[s]
    for prefix, num in _MOIS_HEADER_MAP.items():
        if s.startswith(prefix):
            return num
    return None


# ─── Workbook abstraction (xls + xlsx) ────────────────────────────────────────

class _Workbook:
    """Wrapper unifié pour xlrd (.xls) et openpyxl (.xlsx)."""

    def __init__(self, file_bytes: bytes):
        self._is_xlsx = self._detect_xlsx(file_bytes)
        if self._is_xlsx:
            self._wb = openpyxl.load_workbook(
                filename=io.BytesIO(file_bytes), read_only=True, data_only=True
            )
            self.sheet_names = self._wb.sheetnames
        else:
            self._wb = xlrd.open_workbook(file_contents=file_bytes)
            self.sheet_names = self._wb.sheet_names()

    @staticmethod
    def _detect_xlsx(b: bytes) -> bool:
        # ZIP magic = .xlsx ; OLE = .xls
        return b[:4] == b'PK\x03\x04'

    def sheet(self, name: str) -> '_Sheet':
        if self._is_xlsx:
            return _Sheet(self._wb[name], is_xlsx=True)
        return _Sheet(self._wb.sheet_by_name(name), is_xlsx=False)

    def close(self):
        try:
            if self._is_xlsx:
                self._wb.close()
            else:
                self._wb.release_resources()
        except Exception:
            pass


class _Sheet:
    def __init__(self, sh, is_xlsx: bool):
        self._sh = sh
        self._is_xlsx = is_xlsx
        if is_xlsx:
            self.nrows = sh.max_row or 0
            self.ncols = sh.max_column or 0
        else:
            self.nrows = sh.nrows
            self.ncols = sh.ncols

    def cell(self, r: int, c: int) -> Any:
        """Lit la cellule (0-indexed)."""
        try:
            if self._is_xlsx:
                return self._sh.cell(row=r + 1, column=c + 1).value
            return self._sh.cell_value(r, c)
        except Exception:
            return None

    def row(self, r: int) -> list:
        return [self.cell(r, c) for c in range(self.ncols)]


# ─── Index agences DB ─────────────────────────────────────────────────────────

def _build_agence_index(db, dr_code: str | None = None) -> dict:
    """Construit {slug_nom → id, slug_dr+nom → id} pour résolution flexible."""
    if dr_code:
        rows = db.execute(
            "SELECT a.id, a.nom, d.code AS dr_code "
            "FROM agences a JOIN directions_regionales d ON a.dr_id=d.id "
            "WHERE d.code=?", (dr_code,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT a.id, a.nom, d.code AS dr_code "
            "FROM agences a JOIN directions_regionales d ON a.dr_id=d.id"
        ).fetchall()
    idx: dict = {}
    for r in rows:
        idx[_slug(r['nom'])] = r['id']
        idx[_slug(r['dr_code']) + '__' + _slug(r['nom'])] = r['id']
    return idx


# Alias spécifiques aux noms de feuilles non standards
_AGENCE_ALIASES = {
    'b_ssdi': 'bonamoussadi',
    'bonamoussadi1': 'bonamoussadi',
    'bonaberi1': 'bonaberi',
    'bonaberi': 'bonaberi',
    'koumassi1': 'koumassi',
    'deido1': 'deido',
    'bassa1': 'bassa',
    'nyalla1': 'nyalla',
}


def _resolve_agence(name: str, agence_idx: dict, dr_code: str | None = None) -> int | None:
    """Tente plusieurs stratégies pour mapper un nom de feuille sur une agence."""
    if not name:
        return None
    s = _slug(name)
    # 1) Alias direct
    s_norm = _AGENCE_ALIASES.get(s, s)
    # 2) Strip suffixe « 1 » (CA sheets type "Koumassi1")
    if s_norm.endswith('1') and len(s_norm) > 1:
        s_norm = s_norm.rstrip('1') or s_norm
    # 3) Lookup direct
    if s_norm in agence_idx:
        return agence_idx[s_norm]
    if dr_code:
        key = _slug(dr_code) + '__' + s_norm
        if key in agence_idx:
            return agence_idx[key]
    return None


# ─── Détection DR + année depuis nom de fichier ───────────────────────────────

_DR_CODES = list(STRUCTURE_CW.keys())


def detect_dr_from_filename(filename: str) -> str | None:
    """Extrait le code DR depuis le nom de fichier (ex: 'CA DRDA 2024.xls')."""
    s = filename.upper()
    # Recherche par ordre de longueur décroissante (DRSO avant DRS, DRNO avant DRN)
    for code in sorted(_DR_CODES, key=len, reverse=True):
        if re.search(rf'\b{code}\b', s):
            return code
    return None


def detect_year_from_filename(filename: str) -> int | None:
    m = re.search(r'(20[12]\d)', filename)
    return int(m.group(1)) if m else None


# ─── Mappings rubriques (étendus avec aliases) ────────────────────────────────

# Volumes → catégories DB (table volumes)
# Les slugs incluent 'lte' / 'gt' pour préserver la distinction <= / >
_VOL_RUBRIQUES = {
    'particuliers_lte_10m3': 'Particuliers <= 10m3',
    'particuliers_lte_10': 'Particuliers <= 10m3',
    # Particuliers > 10m3 : 2 occurrences (T2, T3) — distinguées par ordre
    'particuliers_gt_10m3': '__PART_GT_10__',
    'particuliers_gt_10': '__PART_GT_10__',
    'sonel_870': 'Sonel 870',
    'sonel_871': 'Sonel 871',
    'sonel_872': 'Sonel 872',
    'soel_872': 'Sonel 872',  # typo source
    'sonel_873': 'Sonel 873',
    'cadre_eneo_890': 'Cadre Eneo 890',
    'cadre_eneo_891': 'Cadre Eneo 891',
    'cadre_eneo_892': 'Cadre Eneo 892',
    'cadre_eneo_880': 'Cadre Eneo 880',
    'cadre_eneo_881': 'Cadre Eneo 881',
    'b_f_payantes': 'B.F. Payantes',
    'bf_payantes': 'B.F. Payantes',
    'depasst_agts_cde': 'Dépasst agts CDE',
    'depasst_adm_cde': 'Dépasst Adm. CDE',
    'ventes_directes': 'Ventes Directes',
    # Rappel : 2 occurrences (T1=293, T2=364) — distinguées par ordre
    'rappel': '__RAPPEL__',
    'sinistres': 'Sinistres',
    'partants_lte_10m3': 'Partants <= 10m3',
    'partants_lte_10': 'Partants <= 10m3',
    'partants_gt_10m3': 'Partants > 10m3',
    'partants_gt_10': 'Partants > 10m3',
    'fraudes': 'Fraudes',
    'gco': 'GCO',
    'administrations': 'Administrations',
    'batts_communaux': 'Bâtiments communaux',
    'batiments_communaux': 'Bâtiments communaux',
    'b_f_c': 'B.F.C',
    'bfc': 'B.F.C',
    'val_part_agent_cde': 'Val Part Agent CDE',
    'val_services_cde': 'Val Services CDE',
    'val_part_adm_cde': 'Val Part Adm CDE',
    'val_part_agts_camw': 'Val Part agts Camwater',
    'val_part_agts_camwater': 'Val Part agts Camwater',
}

# Lignes à IGNORER (totaux dans le fichier source)
_VOL_SKIP = {
    'total_particuliers', 'total_part_gco', 'total_part_+_gco',
    'total_communes', 'total_cde', 'total', 'totaux',
}

# Rubriques CA spécifiques (rentals — pas de volumes associés)
# Vont dans `ca_specifiques`
_CA_SPEC_RUBRIQUES = {
    'locat_copteur_part': 'Loc. Compteur Particuliers',
    'locat_copteur_part_': 'Loc. Compteur Particuliers',
    'locat_compteur_part': 'Loc. Compteur Particuliers',
    'loc_cptr_part': 'Loc. Compteur Particuliers',
    'locat_compteur_gco': 'Loc. Compteur GCO',
    'locat_cptr_adm': 'Loc. Compteur ADM',
    'location_cptr_bc': 'Loc. Compteur BC',
    'locat_cptr_bfc': 'Loc. Compteur BFC',
    'locat_cptr': 'Loc. Compteur',
}

# Rubriques encaissements → mapping vers RUBRIQUES_ENC_CCIALE de database.py
_ENC_RUBRIQUES = {
    'part_1_2': 'Part. 1 & 2',
    'enc_electroniques': 'Enc. Électroniques',
    'impayes_cde': 'Impayés CDE',
    'gco': 'GCO',
    'enc_cheques': 'Enc. Chèques',
    'anticipat': 'Anticipations',
    'anticipations': 'Anticipations',
    'hors_site': 'Hors site',
    'clts_douteux_cde': 'Clts douteux CDE',
    'clts_douteux_cw': 'Clts douteux CW',
    'resil_a_imp': 'Résil. à imp.',
    'fact_cfd': 'Fact. CFD',
    'frais_imp_tiers': 'Frais imp tiers',
    'beac': 'BEAC',
    'aes_sonel': 'Aes sonel',
    'scdp': 'SCDP',
    'asecna': 'ASECNA',
    'camtel': 'CAMTEL',
    'campost': 'CAMPOST',
    'crtv': 'CRTV',
    'camrail': 'CAMRAIL',
    'universite': 'Université',
    'chu': 'CHU',
    'communes': 'Communes',
    'tva_eau': 'TVA eau',
    'arrondi_eau': 'Arrondi eau',
}

# Champs impayés → colonnes table impayes
_IMP_FIELDS = {
    'particuliers': 'particuliers_actifs',
    'gco': 'gco_actifs',
    'total_actifs': '__skip__',  # calculé
    'particulier': 'particuliers_resilies',  # 2e occurrence "Particulier" = résiliés
    'g_c_o': 'gco_resilies',
    'total_resilies_deb': '__skip__',
    'total_resilies_debiteurs': '__skip__',
    'bornes_fontaines': 'bf_actifs',
    'bornes_fontaines_c': 'bfc_actifs',
    'total_communes_act': '__skip__',
    'total_communes_actifs': '__skip__',
    'borne_fontaine': 'bf_resilies',
    'borne_fontaine_com': 'bfc_resilies',
    'gestion_manuelle': 'gestion_manuelle',
    'total_communes_res': '__skip__',
    'total_communes_resilies': '__skip__',
    'total_impayes': '__skip__',
    'resiliers_crediteur': 'resiliers_crediteurs',
    'resiliers_crediteurs': 'resiliers_crediteurs',
}


# ─── PARSER 1 : Volumes + CA (fichiers « CA <DR> <année>.xls ») ───────────────

def _find_month_header_row(sheet: _Sheet) -> tuple[int, dict]:
    """Trouve la ligne d'en-tête contenant les noms de mois.
    Retourne (numero_ligne, {col_index: numero_mois}).
    """
    for r in range(min(15, sheet.nrows)):
        col_to_mois: dict = {}
        for c in range(sheet.ncols):
            mois = _parse_mois_header(sheet.cell(r, c))
            if mois:
                col_to_mois[c] = mois
        if len(col_to_mois) >= 6:  # au moins 6 mois reconnus = c'est la bonne ligne
            return r, col_to_mois
    return -1, {}


def _is_volumes_sheet(sheet_name: str) -> bool:
    """Détecte si c'est une feuille de volumes (pas la version CA)."""
    s = _slug(sheet_name)
    # Les feuilles CA ont le suffixe '1' ou contiennent ' ca '
    if s.endswith('1') and not s.endswith('11'):  # garde DRDA1 mais exclut un éventuel '11'
        return False
    if 'ca' in s.split('_'):
        return False
    return True


def _is_dr_recap_sheet(sheet_name: str, dr_code: str) -> bool:
    """Détecte si c'est la feuille récapitulative DR (à ignorer pour l'import agences)."""
    s = _slug(sheet_name)
    dr_s = _slug(dr_code)
    return s == dr_s or s == dr_s + '1' or s == dr_s + '_ca' or s.startswith(dr_s + '_')


def _parse_matrix_sheet(sheet: _Sheet, rubrique_mapping: dict,
                         skip_set: set | None = None) -> dict:
    """Parse une feuille au format : ligne=rubrique, colonnes=mois.
    Retourne {rubrique_db: {mois_int: valeur, ...}, ...}.
    Gère les rubriques à doubles occurrences (Particuliers > 10m → T2/T3).
    """
    skip_set = skip_set or set()
    header_row, col_to_mois = _find_month_header_row(sheet)
    if header_row < 0:
        return {}

    result: dict = {}
    seen_part_gt_10 = 0
    seen_rappel = 0

    for r in range(header_row + 1, sheet.nrows):
        label = sheet.cell(r, 0)
        if not label:
            continue
        s = _slug(label)
        if s in skip_set:
            continue

        # Résolution rubrique : match EXACT uniquement (sécurité ↔ rigueur)
        target = rubrique_mapping.get(s)
        if not target:
            continue

        # Cas spéciaux : doubles occurrences
        if target == '__PART_GT_10__':
            seen_part_gt_10 += 1
            target = ('Particuliers > 10m3 T2' if seen_part_gt_10 == 1
                      else 'Particuliers > 10m3 T3')
        elif target == '__RAPPEL__':
            seen_rappel += 1
            target = 'Rappel T1' if seen_rappel == 1 else 'Rappel T2'

        if target in skip_set:
            continue

        # Lecture des cellules mensuelles
        mois_data: dict = {}
        for col_idx, mois_num in col_to_mois.items():
            v = _num(sheet.cell(r, col_idx))
            if v is not None:
                mois_data[mois_num] = v
        if mois_data:
            # Si la rubrique est déjà présente, on additionne (cas de doublons)
            if target in result:
                for m, v in mois_data.items():
                    result[target][m] = result[target].get(m, 0) + v
            else:
                result[target] = mois_data

    return result


# ─── PARSER 2 : Encaissements ─────────────────────────────────────────────────

def _parse_encaissements_sheet(sheet: _Sheet) -> dict:
    """Parse une feuille d'encaissements (matrice rubrique × mois)."""
    return _parse_matrix_sheet(sheet, _ENC_RUBRIQUES, set())


# ─── PARSER 3 : Impayés (format à blocs mensuels) ─────────────────────────────

def _parse_impayes_dr_sheet(sheet: _Sheet, agence_idx: dict, dr_code: str) -> dict:
    """Parse une feuille DR du fichier impayés.
    Format : blocs mensuels successifs, chaque bloc commence par une cellule
    contenant le numéro de série Excel d'une date (ex 45292 = 31/01/2024).
    Lignes du bloc : agences + ENS DRDA. Colonnes : champs impayés.

    Retourne {agence_id: {mois_int: {champ: valeur}}}.
    """
    result: dict = {}
    # Détection des blocs : on cherche les lignes où col0 est un nombre
    # entre 30000 et 80000 (plage des dates Excel modernes)
    bloc_starts: list = []
    for r in range(sheet.nrows):
        v = sheet.cell(r, 0)
        if isinstance(v, (int, float)) and 30000 < v < 80000:
            bloc_starts.append((r, int(v)))

    if not bloc_starts:
        return result

    # Convertir le serial date en mois (Excel epoch = 1900-01-01)
    # serial 45292 = 31/01/2024 → mois 1
    import datetime
    excel_epoch = datetime.date(1899, 12, 30)  # corrigé pour bug Lotus 1900
    bloc_starts_with_mois: list = []
    for row_idx, serial in bloc_starts:
        try:
            d = excel_epoch + datetime.timedelta(days=serial)
            bloc_starts_with_mois.append((row_idx, d.month))
        except Exception:
            continue

    for i, (start_row, mois) in enumerate(bloc_starts_with_mois):
        # Le bloc va du start_row+2 (header) jusqu'au prochain bloc - 1
        next_row = (bloc_starts_with_mois[i + 1][0]
                    if i + 1 < len(bloc_starts_with_mois) else sheet.nrows)
        # Trouver la ligne d'en-tête (Unités | Particuliers | GCO | Total Actifs | ...)
        header_row = -1
        for r in range(start_row, min(start_row + 4, sheet.nrows)):
            v = sheet.cell(r, 0)
            if v and 'unit' in _slug(v):
                header_row = r
                break
        if header_row < 0:
            continue

        # Mapper colonnes → champ DB
        col_to_field: dict = {}
        seen_particulier = 0
        seen_gco = 0
        for c in range(1, sheet.ncols):
            label = _slug(sheet.cell(header_row, c))
            if not label:
                continue
            field = _IMP_FIELDS.get(label)
            if not field:
                # Désambiguïsation par occurrence : "particulier" 1ère = actifs (mais s'appelle "particuliers"), 2e = résiliés
                # Et "g_c_o" 2e occurrence = résiliés
                if label.startswith('particulier'):
                    seen_particulier += 1
                    field = ('particuliers_actifs' if seen_particulier == 1
                             else 'particuliers_resilies')
                elif label == 'gco' or label == 'g_c_o':
                    seen_gco += 1
                    field = 'gco_actifs' if seen_gco == 1 else 'gco_resilies'
            if field and field != '__skip__':
                col_to_field[c] = field

        # Lire les lignes agences
        for r in range(header_row + 1, next_row):
            agence_label = sheet.cell(r, 0)
            if not agence_label:
                continue
            sl = _slug(agence_label)
            # Ignorer lignes de total
            if sl.startswith('ens_') or sl.startswith('total'):
                continue
            agence_id = _resolve_agence(agence_label, agence_idx, dr_code)
            if not agence_id:
                continue
            if agence_id not in result:
                result[agence_id] = {}
            if mois not in result[agence_id]:
                result[agence_id][mois] = {}
            for col_idx, field in col_to_field.items():
                v = _num(sheet.cell(r, col_idx))
                if v is not None:
                    result[agence_id][mois][field] = v

    return result


# ═══ POINT D'ENTRÉE 1 : DÉTECTION ══════════════════════════════════════════════

def detect_file_type(filename: str, file_bytes: bytes) -> dict:
    """Détecte le type de fichier (volumes_ca / encaissements / impayes)
    + DR + année + agences trouvées. Sans toucher à la base.

    Retourne {
        'type': 'volumes_ca' | 'encaissements' | 'impayes' | 'unknown',
        'dr': 'DRDA' | None,
        'annee': 2024 | None,
        'agences_excel': [...],
        'feuilles': [...],
        'message': str,
    }
    """
    fl = filename.lower()
    name_only = fl.replace('.xlsx', '').replace('.xls', '')

    # Détection par nom de fichier
    if name_only.startswith('ca ') or 'ca ' in name_only and 'encais' not in name_only:
        ftype = 'volumes_ca'
    elif 'encais' in name_only:
        ftype = 'encaissements'
    elif 'imp' in name_only or 'ssp' in name_only:
        ftype = 'impayes'
    else:
        ftype = 'unknown'

    dr = detect_dr_from_filename(filename)
    annee = detect_year_from_filename(filename)

    try:
        wb = _Workbook(file_bytes)
        feuilles = wb.sheet_names
        wb.close()
    except Exception as exc:
        return {
            'type': 'unknown',
            'dr': dr, 'annee': annee,
            'agences_excel': [],
            'feuilles': [],
            'message': f"Erreur lecture du fichier : {exc}",
        }

    # Pour volumes/encaissements : trouver les feuilles agences (pas la récap DR)
    agences_excel: list = []
    if ftype in ('volumes_ca', 'encaissements'):
        for sn in feuilles:
            if dr and _is_dr_recap_sheet(sn, dr):
                continue
            if _slug(sn) in ('feuil1', 'feuille1', 'sheet1'):
                continue
            # Pour volumes_ca, garder uniquement les volumes (pas les CA suffixées 1)
            if ftype == 'volumes_ca' and not _is_volumes_sheet(sn):
                continue
            agences_excel.append(sn)

    return {
        'type': ftype,
        'dr': dr,
        'annee': annee,
        'agences_excel': agences_excel,
        'feuilles': feuilles,
        'message': 'OK',
    }


# ═══ POINT D'ENTRÉE 2 : PRÉVISUALISATION ═══════════════════════════════════════

def preview_legacy(file_bytes: bytes, file_type: str, exercice: int,
                    dr_code: str | None = None) -> dict:
    """Analyse le fichier sans écrire en base.
    Retourne un résumé : nb agences trouvées, mois disponibles, exemples.
    """
    db = get_db()
    try:
        agence_idx = _build_agence_index(db, dr_code=dr_code)
    finally:
        db.close()

    try:
        wb = _Workbook(file_bytes)
    except Exception as exc:
        return {'success': False, 'error': f"Fichier illisible : {exc}"}

    summary = {
        'success': True,
        'type': file_type,
        'dr': dr_code,
        'exercice': exercice,
        'agences_traitees': [],
        'agences_inconnues': [],
        'mois_detectes': set(),
        'nb_cellules_estime': 0,
        'errors': [],
        'warnings': [],
    }

    try:
        if file_type == 'volumes_ca':
            for sname in wb.sheet_names:
                if dr_code and _is_dr_recap_sheet(sname, dr_code):
                    continue
                if _slug(sname) in ('feuil1', 'feuille1', 'sheet1'):
                    continue
                is_vol = _is_volumes_sheet(sname)
                # Agence ?
                agence_id = _resolve_agence(sname, agence_idx, dr_code)
                if not agence_id:
                    summary['agences_inconnues'].append(sname)
                    continue
                sh = wb.sheet(sname)
                mapping = _VOL_RUBRIQUES if is_vol else {**_VOL_RUBRIQUES, **_CA_SPEC_RUBRIQUES}
                data = _parse_matrix_sheet(sh, mapping, _VOL_SKIP)
                summary['agences_traitees'].append({
                    'feuille': sname,
                    'agence_id': agence_id,
                    'kind': 'volumes' if is_vol else 'ca',
                    'rubriques': len(data),
                    'lignes_data': sum(len(v) for v in data.values()),
                })
                for v in data.values():
                    summary['mois_detectes'].update(v.keys())
                    summary['nb_cellules_estime'] += len(v)

        elif file_type == 'encaissements':
            for sname in wb.sheet_names:
                if dr_code and _is_dr_recap_sheet(sname, dr_code):
                    continue
                if _slug(sname) in ('feuil1', 'feuille1', 'sheet1'):
                    continue
                agence_id = _resolve_agence(sname, agence_idx, dr_code)
                if not agence_id:
                    summary['agences_inconnues'].append(sname)
                    continue
                sh = wb.sheet(sname)
                data = _parse_encaissements_sheet(sh)
                summary['agences_traitees'].append({
                    'feuille': sname,
                    'agence_id': agence_id,
                    'kind': 'encaissements',
                    'rubriques': len(data),
                    'lignes_data': sum(len(v) for v in data.values()),
                })
                for v in data.values():
                    summary['mois_detectes'].update(v.keys())
                    summary['nb_cellules_estime'] += len(v)

        elif file_type == 'impayes':
            # 1 feuille par DR. On parse la feuille du dr_code (si fourni) ou toutes.
            dr_targets = [dr_code] if dr_code else _DR_CODES
            for dr_t in dr_targets:
                # Trouver feuille correspondante
                feuille_dr = None
                for sn in wb.sheet_names:
                    if _slug(sn) == _slug(dr_t):
                        feuille_dr = sn
                        break
                if not feuille_dr:
                    continue
                sh = wb.sheet(feuille_dr)
                # Recharger l'index agences pour CETTE DR si on traite plusieurs DR
                if not dr_code:
                    db = get_db()
                    try:
                        agence_idx_dr = _build_agence_index(db, dr_code=dr_t)
                    finally:
                        db.close()
                else:
                    agence_idx_dr = agence_idx
                data = _parse_impayes_dr_sheet(sh, agence_idx_dr, dr_t)
                for ag_id, mois_data in data.items():
                    summary['agences_traitees'].append({
                        'feuille': feuille_dr,
                        'agence_id': ag_id,
                        'kind': 'impayes',
                        'mois_couverts': list(mois_data.keys()),
                        'nb_champs': sum(len(c) for c in mois_data.values()),
                    })
                    summary['mois_detectes'].update(mois_data.keys())
                    summary['nb_cellules_estime'] += sum(len(c) for c in mois_data.values())
        else:
            summary['errors'].append(f"Type de fichier non supporté : {file_type}")

    finally:
        wb.close()

    summary['mois_detectes'] = sorted(summary['mois_detectes'])
    return summary


# ═══ POINT D'ENTRÉE 3 : IMPORT EFFECTIF ═══════════════════════════════════════

def import_legacy(file_bytes: bytes, file_type: str, exercice: int,
                   dr_code: str | None = None,
                   mois_to_keep: list | None = None) -> dict:
    """Import définitif. Écrit dans la base avec INSERT OR REPLACE.

    mois_to_keep : liste des mois à importer (1..12). Si None, tous les mois.
    """
    if mois_to_keep is None:
        mois_to_keep = list(range(1, 13))
    mois_set = set(mois_to_keep)

    db = get_db()
    stats = {'volumes': 0, 'ca_specifiques': 0, 'encaissements': 0,
             'impayes': 0, 'agences': 0, 'mois': sorted(mois_set)}
    errors: list = []
    warnings: list = []

    try:
        wb = _Workbook(file_bytes)
    except Exception as exc:
        db.close()
        return {'success': False, 'error': f"Fichier illisible : {exc}",
                'stats': stats}

    try:
        agence_idx = _build_agence_index(db, dr_code=dr_code)
        db.execute('BEGIN')

        if file_type == 'volumes_ca':
            for sname in wb.sheet_names:
                if dr_code and _is_dr_recap_sheet(sname, dr_code):
                    continue
                if _slug(sname) in ('feuil1', 'feuille1', 'sheet1'):
                    continue
                agence_id = _resolve_agence(sname, agence_idx, dr_code)
                if not agence_id:
                    warnings.append(f"Feuille '{sname}' : agence non reconnue, ignorée")
                    continue
                sh = wb.sheet(sname)
                is_vol = _is_volumes_sheet(sname)
                mapping = _VOL_RUBRIQUES if is_vol else {**_VOL_RUBRIQUES, **_CA_SPEC_RUBRIQUES}
                data = _parse_matrix_sheet(sh, mapping, _VOL_SKIP)
                for rubrique, mois_data in data.items():
                    for mois, val in mois_data.items():
                        if mois not in mois_set:
                            continue
                        if rubrique in _CA_SPEC_RUBRIQUES.values():
                            db.execute(
                                "INSERT OR REPLACE INTO ca_specifiques "
                                "(agence_id, mois, exercice, rubrique, montant) "
                                "VALUES (?,?,?,?,?)",
                                (agence_id, mois, exercice, rubrique, val),
                            )
                            stats['ca_specifiques'] += 1
                        else:
                            # Volumes : si c'est une sheet CA on ne réinjecte PAS
                            # le volume (la valeur en FCFA n'est pas un volume).
                            if not is_vol:
                                continue
                            db.execute(
                                "INSERT OR REPLACE INTO volumes "
                                "(agence_id, mois, exercice, categorie, valeur) "
                                "VALUES (?,?,?,?,?)",
                                (agence_id, mois, exercice, rubrique, val),
                            )
                            stats['volumes'] += 1
                stats['agences'] += 1

        elif file_type == 'encaissements':
            for sname in wb.sheet_names:
                if dr_code and _is_dr_recap_sheet(sname, dr_code):
                    continue
                if _slug(sname) in ('feuil1', 'feuille1', 'sheet1'):
                    continue
                agence_id = _resolve_agence(sname, agence_idx, dr_code)
                if not agence_id:
                    warnings.append(f"Feuille '{sname}' : agence non reconnue, ignorée")
                    continue
                sh = wb.sheet(sname)
                data = _parse_encaissements_sheet(sh)
                for rubrique, mois_data in data.items():
                    for mois, val in mois_data.items():
                        if mois not in mois_set:
                            continue
                        db.execute(
                            "INSERT OR REPLACE INTO encaissements "
                            "(agence_id, mois, exercice, section, rubrique, montant) "
                            "VALUES (?,?,?,?,?,?)",
                            (agence_id, mois, exercice, 'Cciale', rubrique, val),
                        )
                        stats['encaissements'] += 1
                stats['agences'] += 1

        elif file_type == 'impayes':
            dr_targets = [dr_code] if dr_code else _DR_CODES
            for dr_t in dr_targets:
                feuille_dr = None
                for sn in wb.sheet_names:
                    if _slug(sn) == _slug(dr_t):
                        feuille_dr = sn
                        break
                if not feuille_dr:
                    continue
                sh = wb.sheet(feuille_dr)
                if not dr_code:
                    agence_idx_dr = _build_agence_index(db, dr_code=dr_t)
                else:
                    agence_idx_dr = agence_idx
                data = _parse_impayes_dr_sheet(sh, agence_idx_dr, dr_t)
                for agence_id, mois_data in data.items():
                    for mois, fields in mois_data.items():
                        if mois not in mois_set:
                            continue
                        # Compose la requête INSERT OR REPLACE
                        cols = ['agence_id', 'mois', 'exercice'] + list(fields.keys())
                        vals = [agence_id, mois, exercice] + list(fields.values())
                        placeholders = ','.join(['?'] * len(cols))
                        sql = (f"INSERT OR REPLACE INTO impayes "
                               f"({','.join(cols)}) VALUES ({placeholders})")
                        db.execute(sql, vals)
                        stats['impayes'] += 1
                    stats['agences'] += 1

        else:
            db.execute('ROLLBACK')
            wb.close()
            db.close()
            return {'success': False,
                    'error': f"Type de fichier non supporté : {file_type}",
                    'stats': stats}

        db.execute('COMMIT')

    except Exception as exc:
        try:
            db.execute('ROLLBACK')
        except Exception:
            pass
        wb.close()
        db.close()
        return {'success': False, 'error': f"Erreur écriture base : {exc}",
                'stats': stats, 'errors': errors, 'warnings': warnings}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    db.close()

    total = stats['volumes'] + stats['ca_specifiques'] + stats['encaissements'] + stats['impayes']
    return {
        'success': True,
        'stats': stats,
        'total_lignes': total,
        'errors': errors,
        'warnings': warnings,
        'message': (
            f"{total} cellule(s) importée(s) — "
            f"vol:{stats['volumes']}, ca_spec:{stats['ca_specifiques']}, "
            f"enc:{stats['encaissements']}, imp:{stats['impayes']}"
        ),
    }
